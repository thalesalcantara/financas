from __future__ import annotations

# ============ Stdlib ============
import os, io, csv, re, json, time, difflib, unicodedata
from collections import defaultdict, namedtuple
from functools import wraps
from types import SimpleNamespace

# MIME types (fixes p/ Office)
import mimetypes
mimetypes.add_type("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", ".xlsx")
mimetypes.add_type("application/vnd.ms-excel", ".xls")
mimetypes.add_type("application/vnd.openxmlformats-officedocument.wordprocessingml.document", ".docx")
mimetypes.add_type("application/msword", ".doc")
mimetypes.add_type("application/vnd.openxmlformats-officedocument.presentationml.presentation", ".pptx")
mimetypes.add_type("application/vnd.ms-powerpoint", ".ppt")

# ============ Terceiros ============
from flask import (
    Flask, render_template, request, redirect, url_for, session,
    flash, send_file, abort, jsonify, current_app
)
from flask_sqlalchemy import SQLAlchemy
from flask_login import UserMixin
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from dateutil.relativedelta import relativedelta
from sqlalchemy import func, text as sa_text, or_, and_, case
from sqlalchemy.inspection import inspect as sa_inspect
from sqlalchemy.pool import QueuePool
from sqlalchemy import event
from sqlalchemy.engine import Engine
from sqlalchemy.exc import OperationalError, SQLAlchemyError, IntegrityError
from sqlalchemy import delete as sa_delete

# ðŸ‘‰ Novo: para gerar XLSX em memÃ³ria
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from urllib.parse import urlparse, parse_qsl, urlencode, urlunparse

BASE_DIR = os.path.abspath(os.path.dirname(__file__))

UPLOAD_DIR = os.path.join(BASE_DIR, "static", "uploads")
DOCS_DIR   = os.path.join(UPLOAD_DIR, "docs")
os.makedirs(DOCS_DIR, exist_ok=True)

# PersistÃªncia real (Render Disk)
PERSIST_ROOT = os.environ.get("PERSIST_ROOT", "/var/data")
if not os.path.isdir(PERSIST_ROOT):
    PERSIST_ROOT = os.path.join(BASE_DIR, "data")
os.makedirs(PERSIST_ROOT, exist_ok=True)
TABELAS_DIR = os.path.join(PERSIST_ROOT, "tabelas")
os.makedirs(TABELAS_DIR, exist_ok=True)
STATIC_TABLES = os.path.join(BASE_DIR, "static", "uploads", "tabelas")
os.makedirs(STATIC_TABLES, exist_ok=True)
DOCS_PERSIST_DIR = os.path.join(PERSIST_ROOT, "docs")
os.makedirs(DOCS_PERSIST_DIR, exist_ok=True)

def _build_db_uri() -> str:
    url = os.environ.get("DATABASE_URL")
    if not url:
        # Rodando local: SQLite simples
        return "sqlite:///" + os.path.join(BASE_DIR, "app.db")
    # Se for postgres no Render
    if url.startswith("postgres://"):
        url = url.replace("postgres://", "postgresql+psycopg://", 1)
    elif url.startswith("postgresql://") and "+psycopg" not in url:
        url = url.replace("postgresql://", "postgresql+psycopg://", 1)
    # Adiciona sslmode, keepalives, etc **SÃ“ SE NÃƒO FOR SQLITE**
    sep = "&" if "?" in url else "?"
    url = (
        f"{url}{sep}"
        "sslmode=require&"
        "keepalives=1&keepalives_idle=30&keepalives_interval=10&keepalives_count=5"
    )
    return url

app = Flask(__name__, static_folder="static", template_folder="templates")
app.secret_key = os.environ.get("SECRET_KEY", "coopex-secret")

URI = _build_db_uri()

# ðŸš« Guard: impede cair em SQLite em produÃ§Ã£o se DATABASE_URL nÃ£o existir
if "sqlite" in URI and os.environ.get("FLASK_ENV") == "production":
    raise RuntimeError("DATABASE_URL ausente em produÃ§Ã£o")

# ===== Pool â€œsem travarâ€, mas seguro =====
workers = int(os.environ.get("WEB_CONCURRENCY", "1") or "1")
target_total = int(os.environ.get("DB_TARGET_CONNS", "60") or "60")
per_worker_target = max(5, target_total // max(1, workers))

pool_size = int(os.environ.get("DB_POOL_SIZE", str(per_worker_target)))
max_overflow = int(os.environ.get("DB_MAX_OVERFLOW", str(max(5, per_worker_target // 2))))

# --- Engine Options adaptados para local x produÃ§Ã£o ---
def _get_engine_options():
    # SÃ³ usa connect_args extra se NÃƒO for SQLite
    if URI.startswith("sqlite"):
        return {
            "poolclass": QueuePool,
            "pool_size": pool_size,
            "max_overflow": max_overflow,
            "pool_timeout": 15,
            "pool_pre_ping": True,
            "pool_use_lifo": True,
        }
    else:
        return {
            "poolclass": QueuePool,
            "pool_size": pool_size,
            "max_overflow": max_overflow,
            "pool_timeout": 15,
            "pool_pre_ping": True,
            "pool_use_lifo": True,
            "connect_args": {
                "connect_timeout": 5,
                "options": "-c statement_timeout=15000",
            },
        }

app.config.update(
    SQLALCHEMY_DATABASE_URI=URI,
    SQLALCHEMY_TRACK_MODIFICATIONS=False,
    JSON_SORT_KEYS=False,
    MAX_CONTENT_LENGTH=32 * 1024 * 1024,  # 32MB
    SESSION_COOKIE_HTTPONLY=True,
    REMEMBER_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=os.environ.get("FLASK_SECURE_COOKIES", "1") == "1",
    PERMANENT_SESSION_LIFETIME=timedelta(hours=12),
    SQLALCHEMY_ENGINE_OPTIONS=_get_engine_options(),
)

db = SQLAlchemy()`r`n`r`ndb.init_app(app)  # <<<<<<<<<<<<<<<  DEFINA O db AQUI ANTES DE USAR NOS MODELS

# Health checks
@app.get("/healthz")
def healthz():
    return "ok", 200

@app.get("/readyz")
def readyz():
    try:
        db.session.execute(sa_text("SELECT 1"))
        return "ready", 200
    except Exception:
        return "not-ready", 503

# Liga foreign_keys no SQLite
@event.listens_for(Engine, "connect")
def _set_sqlite_pragma(dbapi_con, con_record):
    try:
        if app.config["SQLALCHEMY_DATABASE_URI"].startswith("sqlite"):
            cur = dbapi_con.cursor()
            cur.execute("PRAGMA foreign_keys=ON")
            cur.close()
    except Exception:
        pass

def _is_sqlite() -> bool:
    try:
        return db.session.get_bind().dialect.name == "sqlite"
    except Exception:
        return "sqlite" in (app.config.get("SQLALCHEMY_DATABASE_URI") or "")

# ==================
# MODELS VÃƒO AQUI!
# ==================
class Usuario(db.Model, UserMixin):
    __tablename__ = "usuarios"
    id = db.Column(db.Integer, primary_key=True)
    usuario = db.Column(db.String(80), unique=True, nullable=False)
    senha_hash = db.Column(db.String(200), nullable=False)
    tipo = db.Column(db.String(20), nullable=False)  # admin | cooperado | restaurante | farmacia
    ativo = db.Column(db.Boolean, nullable=False, default=True)

    def set_password(self, raw: str):
        self.senha_hash = generate_password_hash(raw)

    def check_password(self, raw: str) -> bool:
        return check_password_hash(self.senha_hash, raw)

        
class Cooperado(db.Model):
    __tablename__ = "cooperados"
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(120), nullable=False)
    usuario_id = db.Column(db.Integer, db.ForeignKey("usuarios.id"), nullable=False)
    usuario_ref = db.relationship("Usuario", backref="coop_account", uselist=False)

    # NOVO
    telefone = db.Column(db.String(30))

    # Foto no banco
    foto_bytes = db.Column(db.LargeBinary)
    foto_mime = db.Column(db.String(100))
    foto_filename = db.Column(db.String(255))
    foto_url = db.Column(db.String(255))
    cnh_numero = db.Column(db.String(50))
    cnh_validade = db.Column(db.Date)
    placa = db.Column(db.String(20))
    placa_validade = db.Column(db.Date)
    ultima_atualizacao = db.Column(db.DateTime)


class Restaurante(db.Model):
    __tablename__ = "restaurantes"

    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(120), nullable=False)
    periodo = db.Column(db.String(20), nullable=False)  # seg-dom | sab-sex | sex-qui

    # VÃNCULO COM USUÃRIO
    usuario_id = db.Column(db.Integer, db.ForeignKey("usuarios.id"), nullable=False)
    # mantÃ©m seu backref atual ("rest_account") â€” o cÃ³digo usa r.usuario_ref normalmente
    usuario_ref = db.relationship("Usuario", backref="rest_account", uselist=False)

    # FOTO NO BANCO
    foto_bytes = db.Column(db.LargeBinary)
    foto_mime = db.Column(db.String(100))
    foto_filename = db.Column(db.String(255))
    # compatibilidade (seu cÃ³digo jÃ¡ usa isso em alguns pontos)
    foto_url = db.Column(db.String(255))

    # >>> NOVO: flag para distinguir restaurante comum de farmÃ¡cia
    farmacia = db.Column(db.Boolean, nullable=False, default=False, server_default=db.text("0"))

    # (opcional) Ã­ndice para filtros/consultas por farmÃ¡cia
    __table_args__ = (
        db.Index("ix_restaurantes_farmacia", "farmacia"),
    )

    def __repr__(self):
        return f"<Restaurante {self.id} {self.nome} farmacia={self.farmacia}>"


class Lancamento(db.Model):
    __tablename__ = "lancamentos"
    id = db.Column(db.Integer, primary_key=True)
    restaurante_id = db.Column(db.Integer, db.ForeignKey("restaurantes.id"), nullable=False)
    cooperado_id = db.Column(db.Integer, db.ForeignKey("cooperados.id"), nullable=False)
    restaurante = db.relationship("Restaurante")
    cooperado = db.relationship("Cooperado")
    descricao = db.Column(db.String(200))
    valor = db.Column(db.Float, default=0.0)
    data = db.Column(db.Date)
    hora_inicio = db.Column(db.String(10))
    hora_fim = db.Column(db.String(10))
    qtd_entregas = db.Column(db.Integer)

    # ===== CAMPOS NOVOS =====
    status_entrega = db.Column(db.String(20), default="entregue", nullable=False)  # entregue, pendente, nao_entregue
    motivo_nao_entregue = db.Column(db.String(200))
    recebido_por = db.Column(db.String(100))

# === AVALIAÃ‡Ã•ES DE COOPERADO (NOVO) =========================================
class AvaliacaoCooperado(db.Model):
    __tablename__ = "avaliacoes"
    id = db.Column(db.Integer, primary_key=True)

    restaurante_id = db.Column(db.Integer, db.ForeignKey("restaurantes.id"), nullable=False)
    cooperado_id   = db.Column(db.Integer, db.ForeignKey("cooperados.id"),  nullable=False)

    # ðŸ”´ IMPORTANTE: CASCADE na FK para o lanÃ§amento
    lancamento_id  = db.Column(
        db.Integer,
        db.ForeignKey("lancamentos.id", ondelete="CASCADE"),
        nullable=True  # deixe True para permitir trocar para SET NULL futuramente, se quiser
    )

    # notas 1..5
    estrelas_geral         = db.Column(db.Integer)
    estrelas_pontualidade  = db.Column(db.Integer)
    estrelas_educacao      = db.Column(db.Integer)
    estrelas_eficiencia    = db.Column(db.Integer)
    estrelas_apresentacao  = db.Column(db.Integer)  # "Bem apresentado"

    comentario       = db.Column(db.Text)

    # IA/heurÃ­sticas
    media_ponderada  = db.Column(db.Float)
    sentimento       = db.Column(db.String(12))     # positivo | neutro | negativo
    temas            = db.Column(db.String(255))    # palavras-chave resumidas
    alerta_crise     = db.Column(db.Boolean, default=False)
    feedback_motoboy = db.Column(db.Text)

    criado_em = db.Column(db.DateTime, default=datetime.utcnow)


class ReceitaCooperativa(db.Model):
    __tablename__ = "receitas_coop"
    id = db.Column(db.Integer, primary_key=True)
    descricao = db.Column(db.String(200), nullable=False)
    valor_total = db.Column(db.Float, default=0.0)
    data = db.Column(db.Date, nullable=True)


class DespesaCooperativa(db.Model):
    __tablename__ = "despesas_coop"
    id = db.Column(db.Integer, primary_key=True)
    descricao = db.Column(db.String(200), nullable=False)
    valor = db.Column(db.Float, default=0.0)
    data = db.Column(db.Date)


class ReceitaCooperado(db.Model):
    __tablename__ = "receitas_cooperado"
    id = db.Column(db.Integer, primary_key=True)
    cooperado_id = db.Column(db.Integer, db.ForeignKey("cooperados.id"), nullable=False)
    cooperado = db.relationship("Cooperado")
    descricao = db.Column(db.String(200), nullable=False)
    valor = db.Column(db.Float, default=0.0)
    data = db.Column(db.Date)


# =========================
# Models
# =========================
class DespesaCooperado(db.Model):
    __tablename__ = "despesas_cooperado"
    id = db.Column(db.Integer, primary_key=True)
    cooperado_id = db.Column(db.Integer, db.ForeignKey("cooperados.id"), nullable=True)  # None = Todos
    cooperado = db.relationship("Cooperado")
    descricao = db.Column(db.String(200), nullable=False)
    valor = db.Column(db.Float, default=0.0)
    # legado (pontual)
    data = db.Column(db.Date)

    # novo (perÃ­odo)
    data_inicio = db.Column(db.Date)
    data_fim    = db.Column(db.Date)

    # NOVO: vÃ­nculo forte para sabermos quem gerou a despesa
    beneficio_id = db.Column(
        db.Integer,
        db.ForeignKey("beneficios_registro.id", ondelete="CASCADE"),
        index=True,
        nullable=True  # deixa True para migrar suave
    )


class BeneficioRegistro(db.Model):
    __tablename__ = "beneficios_registro"
    id = db.Column(db.Integer, primary_key=True)
    data_inicial = db.Column(db.Date, nullable=False)
    data_final = db.Column(db.Date, nullable=False)
    data_lancamento = db.Column(db.Date)
    tipo = db.Column(db.String(40), nullable=False)  # hospitalar | farmaceutico | alimentar
    valor_total = db.Column(db.Float, default=0.0)
    recebedores_nomes = db.Column(db.Text)  # nomes separados por ';'
    recebedores_ids = db.Column(db.Text)    # ids separados por ';'

    # NOVO: relacionamento reverso (nÃ£o remove nada seu)
    despesas = db.relationship(
        "DespesaCooperado",
        primaryjoin="BeneficioRegistro.id==DespesaCooperado.beneficio_id",
        cascade="all, delete-orphan",
        passive_deletes=True
    )

# =========================
# CLI helpers + Master Admin
# =========================
import click  # << necessÃ¡rio para @click.argument

def find_user_by_login(login_raw: str) -> Usuario | None:
    """
    Procura usuÃ¡rio por:
      1) Usuario.usuario (case-insensitive + normalizaÃ§Ã£o se _norm_login existir)
      2) Restaurante.nome -> usa r.usuario_ref (login pelo nome do restaurante)
    """
    if not login_raw:
        return None

    # normalizaÃ§Ã£o forte se _norm_login existir no mÃ³dulo
    try:
        key = _norm_login(login_raw)  # definido mais abaixo no arquivo
    except NameError:
        key = login_raw.strip().lower()

    # 1) match direto case-insensitive
    try:
        u = Usuario.query.filter(func.lower(Usuario.usuario) == login_raw.lower()).first()
        if u:
            return u
    except Exception:
        pass

    # 1b) varrer ativos e comparar por normalizaÃ§Ã£o
    try:
        for u in Usuario.query.filter(Usuario.ativo == True).all():
            try:
                if _norm_login(u.usuario) == key:  # se _norm_login existir
                    return u
            except NameError:
                if (u.usuario or "").strip().lower() == key:
                    return u
    except Exception:
        pass

    # 2) permitir login pelo NOME do restaurante
    try:
        r = Restaurante.query.filter(func.lower(Restaurante.nome) == login_raw.lower()).first()
        if not r:
            # fallback normalizado
            for r0 in Restaurante.query.all():
                try:
                    if _norm_login(r0.nome) == key:
                        r = r0
                        break
                except NameError:
                    if (r0.nome or "").strip().lower() == key:
                        r = r0
                        break
        if r and r.usuario_ref:
            return r.usuario_ref
    except Exception:
        pass

    return None

def ensure_master_admin():
    """
    Garante a existÃªncia do admin mestre.
    Por padrÃ£o: usuario=coopex, senha=05289 (pode sobrescrever via env).
    Se FORCE_MASTER_ADMIN=1, reseta a senha mesmo se jÃ¡ existir.
    """
    user_name = os.environ.get("MASTER_ADMIN_USER", "coopex")
    user_pass = os.environ.get("MASTER_ADMIN_PASS", "05289")
    force = os.environ.get("FORCE_MASTER_ADMIN", "0") == "1"

    adm = Usuario.query.filter_by(usuario=user_name).first()
    if adm:
        if force:
            adm.set_password(user_pass)
            adm.tipo = "admin"
            adm.ativo = True
            db.session.commit()
        return

    adm = Usuario(usuario=user_name, tipo="admin", ativo=True, senha_hash="")
    adm.set_password(user_pass)
    db.session.add(adm)
    db.session.commit()

@app.cli.command("list-users")
def cli_list_users():
    """Lista id, usuario e tipo de todos os usuÃ¡rios."""
    with app.app_context():
        us = Usuario.query.order_by(Usuario.id.asc()).all()
        for u in us:
            print(f"{u.id:>3}  {u.usuario:<20}  tipo={u.tipo}  ativo={u.ativo}")

@app.cli.command("set-pass")
@click.argument("login")
@click.argument("senha")
def cli_set_pass(login, senha):
    """Reseta a senha de um usuÃ¡rio (por login OU nome do restaurante)."""
    with app.app_context():
        u = find_user_by_login(login)
        if not u:
            print("UsuÃ¡rio nÃ£o encontrado.")
            return
        u.set_password(senha)
        u.ativo = True
        db.session.commit()
        print(f"Senha atualizada para '{u.usuario}' (tipo={u.tipo}).")

@app.cli.command("ensure-master")
def cli_ensure_master():
    """Garante a existÃªncia/atualizaÃ§Ã£o do admin mestre."""
    with app.app_context():
        ensure_master_admin()
        print("Master admin garantido.")

# Garante o master admin no import (idempotente)
try:
    with app.app_context():
        ensure_master_admin()
except Exception:
    pass

# =========================
# Semana segâ†’dom + NormalizaÃ§Ã£o automÃ¡tica
# =========================
from sqlalchemy import event

def semana_bounds(d: date):
    """
    Para uma data 'd', retorna (segunda, domingo) da mesma semana.
    0=segunda .. 6=domingo
    """
    dow = d.weekday()
    ini = d - timedelta(days=dow)     # segunda
    fim = ini + timedelta(days=6)     # domingo
    return ini, fim

def normaliza_periodo(data: date|None, data_inicio: date|None, data_fim: date|None):
    """
    Regra global para DespesaCooperado:
      - data_inicio = segunda
      - data_fim    = domingo
      - data        = domingo (== data_fim)
    Se vier sÃ³ 'data', deriva o perÃ­odo pela semana dessa data.
    Se vier inÃ­cio/fim, ajusta para seg/dom da(s) semana(s) informada(s).
    """
    if data_inicio and data_fim:
        # garante ordem e aplica bordas semanais
        if data_fim < data_inicio:
            data_inicio, data_fim = data_fim, data_inicio
        ini, _ = semana_bounds(data_inicio)
        _, fim = semana_bounds(data_fim)
        return fim, ini, fim

    base = data or date.today()
    ini, fim = semana_bounds(base)
    return fim, ini, fim

# === NormalizaÃ§Ã£o automÃ¡tica de perÃ­odo em DespesaCooperado ===
def _ajusta_semana(target: DespesaCooperado):
    d, di, df = normaliza_periodo(target.data, target.data_inicio, target.data_fim)
    target.data = d
    target.data_inicio = di
    target.data_fim = df

@event.listens_for(DespesaCooperado, "before_insert")
def _desp_before_insert(mapper, connection, target):
    _ajusta_semana(target)

@event.listens_for(DespesaCooperado, "before_update")
def _desp_before_update(mapper, connection, target):
    _ajusta_semana(target)



class Escala(db.Model):
    __tablename__ = "escalas"
    id = db.Column(db.Integer, primary_key=True)
    cooperado_id = db.Column(
        db.Integer,
        db.ForeignKey("cooperados.id", ondelete="CASCADE"),
        nullable=True  # pode nÃ£o ter cadastro
    )
    restaurante_id = db.Column(
        db.Integer,
        db.ForeignKey("restaurantes.id", ondelete="CASCADE"),
        nullable=True
    )

    data = db.Column(db.String(40))
    turno = db.Column(db.String(50))
    horario = db.Column(db.String(50))
    contrato = db.Column(db.String(80))
    cor = db.Column(db.String(200))
    cooperado_nome = db.Column(db.String(120))  # nome bruto da planilha quando nÃ£o hÃ¡ cadastro


class TrocaSolicitacao(db.Model):
    __tablename__ = "trocas"
    id = db.Column(db.Integer, primary_key=True)
    solicitante_id = db.Column(
        db.Integer,
        db.ForeignKey("cooperados.id", ondelete="CASCADE"),
        nullable=False
    )
    destino_id = db.Column(
        db.Integer,
        db.ForeignKey("cooperados.id", ondelete="CASCADE"),
        nullable=False
    )
    origem_escala_id = db.Column(
        db.Integer,
        db.ForeignKey("escalas.id", ondelete="CASCADE"),
        nullable=False
    )
    mensagem = db.Column(db.Text)
    status = db.Column(db.String(20), default="pendente")
    criada_em = db.Column(db.DateTime, default=datetime.utcnow)
    aplicada_em = db.Column(db.DateTime)


class Config(db.Model):
    __tablename__ = "config"
    id = db.Column(db.Integer, primary_key=True)
    salario_minimo = db.Column(db.Float, default=0.0)


class Documento(db.Model):
    __tablename__ = "documentos"
    id = db.Column(db.Integer, primary_key=True)
    titulo = db.Column(db.String(200), nullable=False)
    categoria = db.Column(db.String(40))
    descricao = db.Column(db.String(255))
    arquivo_url = db.Column(db.String(255), nullable=False)
    arquivo_nome = db.Column(db.String(255))
    enviado_em = db.Column(db.DateTime, default=datetime.utcnow)


class Tabela(db.Model):

    id = db.Column(db.Integer, primary_key=True)
    titulo = db.Column(db.String(200), nullable=False)
    categoria = db.Column(db.String(40))
    descricao = db.Column(db.String(255))
    arquivo_url = db.Column(db.String(255), nullable=False)
    arquivo_nome = db.Column(db.String(255))
    enviado_em = db.Column(db.DateTime, default=datetime.utcnow)


# ---------- AVISOS (NOVO) ----------
aviso_restaurantes = db.Table(
    "aviso_restaurantes",
    db.Column("aviso_id", db.Integer, db.ForeignKey("avisos.id"), primary_key=True),
    db.Column("restaurante_id", db.Integer, db.ForeignKey("restaurantes.id"), primary_key=True),
)

class Aviso(db.Model):
    __tablename__ = "avisos"
    id = db.Column(db.Integer, primary_key=True)
    titulo = db.Column(db.String(140), nullable=False)
    corpo = db.Column(db.Text, nullable=False)
    # escopo: global | restaurante | cooperado
    tipo = db.Column(db.String(20), nullable=False, default="global")

    # destino individual (opcional)
    destino_cooperado_id = db.Column(db.Integer, db.ForeignKey("cooperados.id"))
    destino_cooperado = db.relationship("Cooperado", foreign_keys=[destino_cooperado_id])

    # destino por restaurante (opcional, N:N)
    restaurantes = db.relationship("Restaurante", secondary=aviso_restaurantes, backref="avisos")

    prioridade = db.Column(db.String(10), default="normal")  # normal | alta
    fixado = db.Column(db.Boolean, default=False)
    ativo = db.Column(db.Boolean, default=True)
    inicio_em = db.Column(db.DateTime)  # janela de exibiÃ§Ã£o opcional
    fim_em = db.Column(db.DateTime)

    criado_por_id = db.Column(db.Integer, db.ForeignKey("usuarios.id"), nullable=False)
    criado_em = db.Column(db.DateTime, default=datetime.utcnow)
    atualizado_em = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class AvisoLeitura(db.Model):
    __tablename__ = "aviso_leituras"
    id = db.Column(db.Integer, primary_key=True)
    aviso_id = db.Column(db.Integer, db.ForeignKey("avisos.id"), nullable=False, index=True)
    cooperado_id = db.Column(db.Integer, db.ForeignKey("cooperados.id"), nullable=True, index=True)
    restaurante_id = db.Column(db.Integer, db.ForeignKey("restaurantes.id"), nullable=True, index=True)
    lido_em = db.Column(db.DateTime, default=datetime.utcnow)
    __table_args__ = (db.UniqueConstraint("aviso_id", "cooperado_id", "restaurante_id", name="uq_aviso_dest"), )

# =========================
# Helpers
# =========================
def _is_sqlite() -> bool:
    try:
        return db.session.get_bind().dialect.name == "sqlite"
    except Exception:
        return "sqlite" in (app.config.get("SQLALCHEMY_DATABASE_URI") or "")


# =========================
# Init DB / MigraÃ§Ã£o leve
# =========================
def init_db():
    """
    VersÃ£o unificada e idempotente:
      1) Ajustes de performance para SQLite (WAL/synchronous)
      2) CriaÃ§Ã£o de todas as tabelas (create_all)
      3) Ãndices Ãºteis (cooperado/restaurante/criado_em)
      4) MigraÃ§Ãµes leves (qtd_entregas, ativo em usuarios, colunas de escalas, fotos, tabela avaliacoes_restaurante)
      5) Bootstrap mÃ­nimo (admin e config) â€” sÃ³ se os modelos existirem
    """
    # 1) Perf no SQLite
    try:
        if _is_sqlite():
            db.session.execute(sa_text("PRAGMA journal_mode=WAL;"))
            db.session.execute(sa_text("PRAGMA synchronous=NORMAL;"))
            db.session.commit()
    except Exception:
        db.session.rollback()

    # 2) Tabelas/mapeamentos
    try:
        db.create_all()
    except Exception:
        db.session.rollback()

    # 3) Ãndices de performance (idempotentes)
    try:
        if _is_sqlite():
            db.session.execute(sa_text("""
                CREATE INDEX IF NOT EXISTS ix_avaliacoes_criado_em        ON avaliacoes (criado_em);
                CREATE INDEX IF NOT EXISTS ix_avaliacoes_rest_criado      ON avaliacoes (restaurante_id, criado_em);
                CREATE INDEX IF NOT EXISTS ix_avaliacoes_coop_criado      ON avaliacoes (cooperado_id,  criado_em);

                CREATE INDEX IF NOT EXISTS ix_av_rest_criado_em           ON avaliacoes_restaurante (criado_em);
                CREATE INDEX IF NOT EXISTS ix_av_rest_rest_criado         ON avaliacoes_restaurante (restaurante_id, criado_em);
                CREATE INDEX IF NOT EXISTS ix_av_rest_coop_criado         ON avaliacoes_restaurante (cooperado_id,  criado_em);
            """))
        else:
            db.session.execute(sa_text("""
                CREATE INDEX IF NOT EXISTS ix_avaliacoes_criado_em        ON public.avaliacoes (criado_em);
                CREATE INDEX IF NOT EXISTS ix_avaliacoes_rest_criado      ON public.avaliacoes (restaurante_id, criado_em);
                CREATE INDEX IF NOT EXISTS ix_avaliacoes_coop_criado      ON public.avaliacoes (cooperado_id,  criado_em);

                CREATE INDEX IF NOT EXISTS ix_av_rest_criado_em           ON public.avaliacoes_restaurante (criado_em);
                CREATE INDEX IF NOT EXISTS ix_av_rest_rest_criado         ON public.avaliacoes_restaurante (restaurante_id, criado_em);
                CREATE INDEX IF NOT EXISTS ix_av_rest_coop_criado         ON public.avaliacoes_restaurante (cooperado_id,  criado_em);
            """))
        db.session.commit()
    except Exception:
        db.session.rollback()

    # 4) MigraÃ§Ã£o leve: garantir coluna qtd_entregas em lancamentos
    try:
        if _is_sqlite():
            cols = db.session.execute(sa_text("PRAGMA table_info(lancamentos);")).fetchall()
            colnames = {row[1] for row in cols}
            if "qtd_entregas" not in colnames:
                db.session.execute(sa_text("ALTER TABLE lancamentos ADD COLUMN qtd_entregas INTEGER"))
            db.session.commit()
        else:
            db.session.execute(sa_text(
                "ALTER TABLE IF EXISTS public.lancamentos "
                "ADD COLUMN IF NOT EXISTS qtd_entregas INTEGER"
            ))
            db.session.commit()
    except Exception:
        db.session.rollback()

       # 4.x) perÃ­odo em despesas_cooperado (data_inicio / data_fim) + backfill
    try:
        if _is_sqlite():
            cols = db.session.execute(sa_text("PRAGMA table_info(despesas_cooperado);")).fetchall()
            colnames = {row[1] for row in cols}

            if "data_inicio" not in colnames:
                db.session.execute(sa_text("ALTER TABLE despesas_cooperado ADD COLUMN data_inicio DATE"))
            if "data_fim" not in colnames:
                db.session.execute(sa_text("ALTER TABLE despesas_cooperado ADD COLUMN data_fim DATE"))
            db.session.commit()

            # Retropreenche linhas antigas
            db.session.execute(sa_text("""
                UPDATE despesas_cooperado
                   SET data_inicio = COALESCE(data_inicio, data),
                       data_fim    = COALESCE(data_fim,    data)
                 WHERE data IS NOT NULL
                   AND (data_inicio IS NULL OR data_fim IS NULL)
            """))
            db.session.commit()
        else:
            db.session.execute(sa_text("""
                ALTER TABLE IF NOT EXISTS public.despesas_cooperado
                ADD COLUMN IF NOT EXISTS data_inicio DATE
            """))
            db.session.execute(sa_text("""
                ALTER TABLE IF NOT EXISTS public.despesas_cooperado
                ADD COLUMN IF NOT EXISTS data_fim DATE
            """))
            db.session.commit()

            db.session.execute(sa_text("""
                UPDATE public.despesas_cooperado
                   SET data_inicio = COALESCE(data_inicio, data),
                       data_fim    = COALESCE(data_fim,    data)
                 WHERE data IS NOT NULL
                   AND (data_inicio IS NULL OR data_fim IS NULL)
            """))
            db.session.commit()
    except Exception:
        db.session.rollback()

    # 4.y) beneficio_id em despesas_cooperado (FK p/ beneficios_registro)
    try:
        if _is_sqlite():
           cols = db.session.execute(sa_text("PRAGMA table_info(despesas_cooperado);")).fetchall()
           colnames = {row[1] for row in cols}
           if "beneficio_id" not in colnames:
               db.session.execute(sa_text("ALTER TABLE despesas_cooperado ADD COLUMN beneficio_id INTEGER"))
           db.session.commit()
        # OBS: SQLite nÃ£o permite adicionar uma FK com ON DELETE CASCADE via ALTER TABLE;
        # para ter a FK de fato, teria que recriar a tabela. Em dev, costuma bastar sÃ³ a coluna.
        else:
            db.session.execute(sa_text("""
                ALTER TABLE IF NOT EXISTS public.despesas_cooperado
                ADD COLUMN IF NOT EXISTS beneficio_id INTEGER
         """))
            db.session.execute(sa_text("""
            CREATE INDEX IF NOT EXISTS ix_despesas_beneficio_id
            ON public.despesas_cooperado (beneficio_id)
         """))
            db.session.execute(sa_text("""
            ALTER TABLE public.despesas_cooperado
            DROP CONSTRAINT IF EXISTS despesas_cooperado_beneficio_id_fkey
        """))
        db.session.execute(sa_text("""
            ALTER TABLE public.despesas_cooperado
            ADD CONSTRAINT despesas_cooperado_beneficio_id_fkey
            FOREIGN KEY (beneficio_id) REFERENCES public.beneficios_registro (id)
            ON DELETE CASCADE
        """))
        db.session.commit()
    except Exception:
        db.session.rollback()

    # 4.1) cooperado_nome em escalas
    try:
        if _is_sqlite():
            cols = db.session.execute(sa_text("PRAGMA table_info(escalas);")).fetchall()
            colnames = {row[1] for row in cols}
            if "cooperado_nome" not in colnames:
                db.session.execute(sa_text("ALTER TABLE escalas ADD COLUMN cooperado_nome VARCHAR(120)"))
            db.session.commit()
        else:
            db.session.execute(sa_text(
                "ALTER TABLE IF NOT EXISTS escalas "
                "ADD COLUMN IF NOT EXISTS cooperado_nome VARCHAR(120)"
            ))
            db.session.commit()
    except Exception:
        db.session.rollback()

    # 4.2) restaurante_id em escalas
    try:
        if _is_sqlite():
            cols = db.session.execute(sa_text("PRAGMA table_info(escalas);")).fetchall()
            colnames = {row[1] for row in cols}
            if "restaurante_id" not in colnames:
                db.session.execute(sa_text("ALTER TABLE escalas ADD COLUMN restaurante_id INTEGER"))
            db.session.commit()
        else:
            db.session.execute(sa_text(
                "ALTER TABLE IF NOT EXISTS escalas "
                "ADD COLUMN IF NOT EXISTS restaurante_id INTEGER"
            ))
            db.session.commit()
    except Exception:
        db.session.rollback()

    # 4.3) fotos no banco (cooperados)
    try:
        if _is_sqlite():
            cols = db.session.execute(sa_text("PRAGMA table_info(cooperados);")).fetchall()
            colnames = {row[1] for row in cols}
            if "foto_bytes" not in colnames:
                db.session.execute(sa_text("ALTER TABLE cooperados ADD COLUMN foto_bytes BLOB"))
            if "foto_mime" not in colnames:
                db.session.execute(sa_text("ALTER TABLE cooperados ADD COLUMN foto_mime VARCHAR(100)"))
            if "foto_filename" not in colnames:
                db.session.execute(sa_text("ALTER TABLE cooperados ADD COLUMN foto_filename VARCHAR(255)"))
            if "foto_url" not in colnames:
                db.session.execute(sa_text("ALTER TABLE cooperados ADD COLUMN foto_url VARCHAR(255)"))
            db.session.commit()
        else:
            db.session.execute(sa_text(
                "ALTER TABLE IF NOT EXISTS cooperados ADD COLUMN IF NOT EXISTS foto_bytes BYTEA"))
            db.session.execute(sa_text(
                "ALTER TABLE IF NOT EXISTS cooperados ADD COLUMN IF NOT EXISTS foto_mime VARCHAR(100)"))
            db.session.execute(sa_text(
                "ALTER TABLE IF NOT EXISTS cooperados ADD COLUMN IF NOT EXISTS foto_filename VARCHAR(255)"))
            db.session.execute(sa_text(
                "ALTER TABLE IF NOT EXISTS cooperados ADD COLUMN IF NOT EXISTS foto_url VARCHAR(255)"))
            db.session.commit()
    except Exception:
        db.session.rollback()

    # 4.3.x) telefone em cooperados
    try:
        if _is_sqlite():
            cols = db.session.execute(sa_text("PRAGMA table_info(cooperados);")).fetchall()
            colnames = {row[1] for row in cols}
            if "telefone" not in colnames:
                db.session.execute(sa_text("ALTER TABLE cooperados ADD COLUMN telefone VARCHAR(30)"))
            db.session.commit()
        else:
            db.session.execute(sa_text(
                "ALTER TABLE IF NOT EXISTS cooperados "
                "ADD COLUMN IF NOT EXISTS telefone VARCHAR(30)"
            ))
            db.session.commit()
    except Exception:
        db.session.rollback()

    # 4.4) tabela avaliacoes_restaurante (se nÃ£o existir)
    try:
        if _is_sqlite():
            db.session.execute(sa_text("""
                CREATE TABLE IF NOT EXISTS avaliacoes_restaurante (
                  id INTEGER PRIMARY KEY,
                  restaurante_id INTEGER NOT NULL,
                  cooperado_id   INTEGER NOT NULL,
                  lancamento_id  INTEGER UNIQUE,
                  estrelas_geral INTEGER,
                  estrelas_pontualidade INTEGER,
                  estrelas_educacao INTEGER,
                  estrelas_eficiencia INTEGER,
                  estrelas_apresentacao INTEGER,
                  comentario TEXT,
                  media_ponderada REAL,
                  sentimento TEXT,
                  temas TEXT,
                  alerta_crise INTEGER DEFAULT 0,
                  criado_em TIMESTAMP
                );
            """))
            db.session.execute(sa_text(
                "CREATE INDEX IF NOT EXISTS ix_av_rest_rest ON avaliacoes_restaurante(restaurante_id, criado_em)"))
            db.session.execute(sa_text(
                "CREATE INDEX IF NOT EXISTS ix_av_rest_coop ON avaliacoes_restaurante(cooperado_id)"))
            db.session.commit()
        else:
            db.session.execute(sa_text("""
                CREATE TABLE IF NOT EXISTS avaliacoes_restaurante (
                  id SERIAL PRIMARY KEY,
                  restaurante_id INTEGER NOT NULL,
                  cooperado_id   INTEGER NOT NULL,
                  lancamento_id  INTEGER UNIQUE,
                  estrelas_geral INTEGER,
                  estrelas_pontualidade INTEGER,
                  estrelas_educacao INTEGER,
                  estrelas_eficiencia INTEGER,
                  estrelas_apresentacao INTEGER,
                  comentario TEXT,
                  media_ponderada DOUBLE PRECISION,
                  sentimento VARCHAR(12),
                  temas VARCHAR(255),
                  alerta_crise BOOLEAN DEFAULT FALSE,
                  criado_em TIMESTAMP
                );
            """))
            db.session.execute(sa_text(
                "CREATE INDEX IF NOT EXISTS ix_av_rest_rest ON avaliacoes_restaurante(restaurante_id, criado_em)"))
            db.session.execute(sa_text(
                "CREATE INDEX IF NOT EXISTS ix_av_rest_coop ON avaliacoes_restaurante(cooperado_id)"))
            db.session.commit()
    except Exception:
        db.session.rollback()

    # 4.5) fotos no banco (restaurantes)
    try:
        if _is_sqlite():
            cols = db.session.execute(sa_text("PRAGMA table_info(restaurantes);")).fetchall()
            colnames = {row[1] for row in cols}
            if "foto_bytes" not in colnames:
                db.session.execute(sa_text("ALTER TABLE restaurantes ADD COLUMN foto_bytes BLOB"))
            if "foto_mime" not in colnames:
                db.session.execute(sa_text("ALTER TABLE restaurantes ADD COLUMN foto_mime VARCHAR(100)"))
            if "foto_filename" not in colnames:
                db.session.execute(sa_text("ALTER TABLE restaurantes ADD COLUMN foto_filename VARCHAR(255)"))
            if "foto_url" not in colnames:
                db.session.execute(sa_text("ALTER TABLE restaurantes ADD COLUMN foto_url VARCHAR(255)"))
            db.session.commit()
        else:
            db.session.execute(sa_text(
                "ALTER TABLE IF NOT EXISTS restaurantes ADD COLUMN IF NOT EXISTS foto_bytes BYTEA"))
            db.session.execute(sa_text(
                "ALTER TABLE IF NOT EXISTS restaurantes ADD COLUMN IF NOT EXISTS foto_mime VARCHAR(100)"))
            db.session.execute(sa_text(
                "ALTER TABLE IF NOT EXISTS restaurantes ADD COLUMN IF NOT EXISTS foto_filename VARCHAR(255)"))
            db.session.execute(sa_text(
                "ALTER TABLE IF NOT EXISTS restaurantes ADD COLUMN IF NOT EXISTS foto_url VARCHAR(255)"))
            db.session.commit()
    except Exception:
        db.session.rollback()


    # 5) Bootstrap mÃ­nimo (admin e config) â€” sÃ³ se os modelos existirem
    try:
        # se houver um model Config definido em outro lugar, este bloco funciona;
        # senÃ£o, apenas ignora silenciosamente
        _ = Usuario  # garante que o model estÃ¡ acessÃ­vel

        # Admin
        try:
            tem_admin = Usuario.query.filter_by(tipo="admin").first()
        except Exception:
            tem_admin = None

        if not tem_admin:
            admin_user = os.environ.get("ADMIN_USER", "admin")
            admin_pass = os.environ.get("ADMIN_PASS", os.urandom(8).hex())
            admin = Usuario(usuario=admin_user, tipo="admin", senha_hash="")
            try:
                admin.set_password(admin_pass)
            except Exception:
                try:
                    admin.senha_hash = generate_password_hash(admin_pass)
                except Exception:
                    pass
            db.session.add(admin)
            db.session.commit()
    except Exception:
        db.session.rollback()

    try:
        # Se o model Config existir, cria default
        from sqlalchemy.exc import InvalidRequestError
        try:
            Config  # type: ignore[name-defined]
            has_config_model = True
        except NameError:
            has_config_model = False

        if has_config_model:
            if not db.session.get(Config, 1):  # type: ignore[name-defined]
                db.session.add(Config(id=1, salario_minimo=0.0))  # type: ignore[name-defined]
                db.session.commit()
    except Exception:
        db.session.rollback()


# === Bootstrap do banco no start (Render/Gunicorn) ===
try:
    if os.environ.get("INIT_DB_ON_START", "1") == "1":
        _t0 = datetime.now(timezone.utc)
        with app.app_context():
            init_db()
        try:
            app.logger.info(f"init_db concluÃ­do em {(datetime.now(timezone.utc) - _t0).total_seconds():.2f}s")
        except Exception:
            pass
    else:
        try:
            app.logger.info("INIT_DB_ON_START=0: pulando init_db no boot.")
        except Exception:
            pass
except Exception as e:
    try:
        app.logger.warning(f"init_db falhou/pulado: {e}")
    except Exception:
        pass


def get_config() -> Config:
    cfg = db.session.get(Config, 1)
    if not cfg:
        cfg = Config(id=1, salario_minimo=0.0)
        db.session.add(cfg)
        db.session.commit()
    return cfg


# ============================================================
# Decorator de papel/tipo de usuÃ¡rio com mÃºltiplos tipos
# ============================================================
from functools import wraps
from flask import session, redirect, url_for, flash, abort

def role_required(*allowed_types):
    """
    Uso:
      @role_required("admin")
      @role_required("restaurante")
      @role_required("restaurante", "farmacia")   # mÃºltiplos
      @role_required                             # apenas exige estar logado
    """
    # Suporte opcional ao uso sem argumentos: @role_required
    if len(allowed_types) == 1 and callable(allowed_types[0]):
        fn = allowed_types[0]
        @wraps(fn)
        def wrapper(*args, **kwargs):
            u_id = session.get("user_id")
            if not u_id:
                flash("FaÃ§a login para continuar.", "warning")
                return redirect(url_for("login"))
            return fn(*args, **kwargs)
        return wrapper

    # Uso normal: @role_required("tipo1", "tipo2", ...)
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            u_id = session.get("user_id")
            if not u_id:
                flash("FaÃ§a login para continuar.", "warning")
                return redirect(url_for("login"))

            user = db.session.get(Usuario, u_id)
            if not user or not getattr(user, "ativo", True):
                session.clear()
                flash("SessÃ£o invÃ¡lida. Entre novamente.", "warning")
                return redirect(url_for("login"))

            if allowed_types and user.tipo not in allowed_types:
                abort(403)

            return fn(*args, **kwargs)
        return wrapper
    return decorator


def admin_required(fn):
    return role_required("admin")(fn)


def _normalize_name(s: str) -> list[str]:
    s = unicodedata.normalize("NFD", s or "")
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = re.sub(r"[^a-zA-Z0-9\s]", " ", s)
    parts = [p.lower() for p in s.split() if p.strip()]
    return parts


def _norm_login(s: str) -> str:
    # remove acento, minÃºsculo e sem espaÃ§os
    s = unicodedata.normalize("NFD", s or "")
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = s.lower().strip()
    s = re.sub(r"\s+", "", s)
    return s

def _match_cooperado_by_login(login_planilha: str, cooperados: list[Cooperado]) -> Cooperado | None:
    """Casa EXATAMENTE com Usuario.usuario apÃ³s normalizaÃ§Ã£o."""
    key = _norm_login(login_planilha)
    if not key:
        return None
    for c in cooperados:
        # c.usuario_ref.usuario Ã© o login usado no sistema
        login = getattr(c.usuario_ref, "usuario", "") or ""
        if _norm_login(login) == key:
            return c
    return None


def _match_restaurante_id(contrato_txt: str) -> int | None:
    alvo = " ".join(_normalize_name(contrato_txt or ""))
    if not alvo:
        return None
    restaurantes = Restaurante.query.order_by(Restaurante.nome.asc()).all()

    for r in restaurantes:
        rn = " ".join(_normalize_name(r.nome))
        if alvo == rn or alvo in rn or rn in alvo:
            return r.id

    try:
        nomes_norm = [" ".join(_normalize_name(r.nome)) for r in restaurantes]
        close = difflib.get_close_matches(alvo, nomes_norm, n=1, cutoff=0.85)
        if close:
            alvo_norm = close[0]
            for r in restaurantes:
                if " ".join(_normalize_name(r.nome)) == alvo_norm:
                    return r.id
    except Exception:
        pass
    return None


def _match_cooperado_by_name(nome_planilha: str, cooperados: list[Cooperado]) -> Cooperado | None:
    def norm_join(s: str) -> str:
        return " ".join(_normalize_name(s))

    sheet_tokens = _normalize_name(nome_planilha)
    sheet_norm = " ".join(sheet_tokens)
    if not sheet_norm:
        return None

    for c in cooperados:
        c_norm = norm_join(c.nome)
        if sheet_norm == c_norm or sheet_norm in c_norm or c_norm in sheet_norm:
            return c

    parts_sheet = set(sheet_tokens)
    best, best_count = None, 0
    for c in cooperados:
        parts_c = set(_normalize_name(c.nome))
        inter = parts_sheet & parts_c
        if len(inter) > best_count:
            best, best_count = c, len(inter)
    if best and best_count >= 2:
        return best

    if len(sheet_tokens) == 1 and len(sheet_tokens[0]) >= 3:
        token = sheet_tokens[0]
        hits = [c for c in cooperados if token in set(_normalize_name(c.nome))]
        if os.environ.get("INIT_DB_ON_START", "0") == "1":
            return hits[0]

    names_norm = [norm_join(c.nome) for c in cooperados]
    close = difflib.get_close_matches(sheet_norm, names_norm, n=1, cutoff=0.85)
    if close:
        target = close[0]
        for c in cooperados:
            if norm_join(c.nome) == target:
                return c
    return None


def _build_docinfo(c: Cooperado) -> dict:
    today = date.today()
    cnh_ok = (c.cnh_validade is not None and c.cnh_validade >= today)
    placa_ok = (c.placa_validade is not None and c.placa_validade >= today)
    return {"cnh": {"ok": cnh_ok}, "placa": {"ok": placa_ok}}


def _save_upload(file_storage) -> str | None:
    # Mantido para compatibilidade com outras partes do app (ex.: uploads de xlsx)
    if not file_storage:
        return None
    fname = secure_filename(file_storage.filename or "")
    if not fname:
        return None
    path = os.path.join(UPLOAD_DIR, fname)
    file_storage.save(path)
    return f"/static/uploads/{fname}"

def salvar_tabela_upload(file_storage) -> str | None:
    """
    Salva o arquivo de TABELA dentro do diretÃ³rio persistente (TABELAS_DIR)
    e retorna APENAS o nome do arquivo (para guardar no banco em Tabela.arquivo_nome).
    """
    if not file_storage or not file_storage.filename:
        return None
    fname = secure_filename(file_storage.filename)
    base, ext = os.path.splitext(fname)
    unique = f"{base}_{time.strftime('%Y%m%d_%H%M%S')}{ext.lower()}"
    destino = os.path.join(TABELAS_DIR, unique)
    file_storage.save(destino)
    return unique  # <- guarde este em Tabela.arquivo_nome


def resolve_tabela_path(nome_arquivo: str) -> str | None:
    """
    Resolve o caminho real de uma TABELA:
      1) /var/data/tabelas   (persistente)
      2) static/uploads/...  (legado)
    """
    if not nome_arquivo:
        return None
    candidatos = [
        os.path.join(TABELAS_DIR, nome_arquivo),
        os.path.join(STATIC_TABLES, nome_arquivo),  # legado
        # Ãºltimo fallback: se por acaso gravaram caminho completo em arquivo_url
        _abs_path_from_url(nome_arquivo) if nome_arquivo.startswith("/") else None,
    ]
    for p in candidatos:
        if p and os.path.isfile(p):
            return p
    # log amigÃ¡vel (vai parar com o WARNING que vocÃª viu)
    app.logger.warning("Arquivo de Tabela nÃ£o encontrado. nome='%s' tents=%s",
                       nome_arquivo, [c for c in candidatos if c])
    return None

def _save_foto_to_db(entidade, file_storage, *, is_cooperado: bool) -> str | None:
    """
    Salva o arquivo enviado diretamente no banco (bytea/Blob) e
    retorna uma URL interna (/media/coop/<id> ou /media/rest/<id>).
    """
    if not file_storage or not file_storage.filename:
        return getattr(entidade, "foto_url", None)
    data = file_storage.read()
    if not data:
        return getattr(entidade, "foto_url", None)
    entidade.foto_bytes = data
    entidade.foto_mime = (file_storage.mimetype or "application/octet-stream")
    entidade.foto_filename = secure_filename(file_storage.filename)
    # garante que temos ID
    db.session.flush()
    if is_cooperado:
        url = url_for("media_coop", coop_id=entidade.id)
    else:
        url = url_for("media_rest", rest_id=entidade.id)
    entidade.foto_url = f"{url}?v={int(datetime.now(timezone.utc).timestamp())}"
    return entidade.foto_url

def _abs_path_from_url(rel_url: str) -> str:
    """
    Converte '/static/uploads/arquivo.pdf' para o caminho absoluto no disco.
    """
    if not rel_url:
        return ""
    # caminho padrÃ£o: /static/uploads/...
    if rel_url.startswith("/"):
        rel_url = rel_url.lstrip("/")
    return os.path.join(BASE_DIR, rel_url.replace("/", os.sep))

def _serve_uploaded(rel_url: str, *, download_name: str | None = None, force_download: bool = False):
    """
    Entrega um arquivo salvo em /static/uploads com mimetype correto.
    - PDFs abrem inline (no navegador) por padrÃ£o.
    - Se quiser forÃ§ar download, passe force_download=True.
    """
    if not rel_url:
        abort(404)
    abs_path = _abs_path_from_url(rel_url)
    if not os.path.exists(abs_path):
        abort(404)

    mime, _ = mimetypes.guess_type(abs_path)
    is_pdf = (mime == "application/pdf") or abs_path.lower().endswith(".pdf")
    return send_file(
        abs_path,
        mimetype=mime or "application/octet-stream",
        as_attachment=(force_download or not is_pdf),
        download_name=(download_name or os.path.basename(abs_path)),
        conditional=True,     # ajuda visualizaÃ§Ã£o/retomar download
    )


# ========= Helpers de DOCUMENTOS (PDFs, etc.) =========
def salvar_documento_upload(file_storage) -> str | None:
    """
    Salva o arquivo em disco persistente (/var/data/docs ou BASE_DIR/data/docs)
    e retorna APENAS o nome do arquivo (para guardar no banco em Documento.arquivo_nome).
    Requer que DOCS_PERSIST_DIR jÃ¡ exista (criado no Passo 1).
    """
    if not file_storage or not file_storage.filename:
        return None
    fname = secure_filename(file_storage.filename)
    base, ext = os.path.splitext(fname)
    unique = f"{base}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}{ext.lower()}"
    destino = os.path.join(DOCS_PERSIST_DIR, unique)
    file_storage.save(destino)
    return unique  # <- gravar em Documento.arquivo_nome

def resolve_documento_path(nome_arquivo: str) -> str | None:
    """
    Resolve o caminho real do documento nesta ordem:
      1) persistente (/var/data/docs ou BASE_DIR/data/docs)
      2) legado (DOCS_DIR -> static/uploads/docs)
      3) caminho absoluto derivado de '/static/...'
    """
    if not nome_arquivo:
        return None
    candidatos = [
        os.path.join(DOCS_PERSIST_DIR, nome_arquivo),   # persistente
        os.path.join(DOCS_DIR, nome_arquivo),           # legado
        _abs_path_from_url(nome_arquivo) if str(nome_arquivo).startswith("/") else None,
    ]
    for p in candidatos:
        if p and os.path.isfile(p):
            return p
    app.logger.warning("Documento nÃ£o encontrado. nome='%s' tents=%s",
                       nome_arquivo, [c for c in candidatos if c])
    return None

def _assert_cooperado_ativo(cooperado_id: int):
    c = (Cooperado.query
         .join(Usuario, Cooperado.usuario_id == Usuario.id)
         .filter(Cooperado.id == cooperado_id, Usuario.ativo.is_(True))
         .first())
    if not c:
        abort(400, description="Cooperado inativo ou inexistente.")
    return c

# ========= ROTA: /docs/<nome> (abre inline PDF; baixa outros tipos) =========
@app.get("/docs/<path:nome>")
def serve_documento(nome: str):
    """
    Abre PDFs inline (no navegador) e forÃ§a download para outros tipos.
    Busca primeiro no disco persistente e faz fallback pro legado.
    """
    path = resolve_documento_path(nome)
    if not path:
        abort(404)

    mime, _ = mimetypes.guess_type(path)
    is_pdf = (mime == "application/pdf") or path.lower().endswith(".pdf")

    return send_file(
        path,
        mimetype=mime or "application/octet-stream",
        as_attachment=not is_pdf,                  # PDF inline, outros baixam
        download_name=os.path.basename(path),
        conditional=True,
    )

def _prox_ocorrencia_anual(dt: date | None) -> date | None:
    if not dt:
        return None
    hoje = date.today()
    alvo = date(hoje.year, dt.month, dt.day)
    if alvo < hoje:
        alvo = date(hoje.year + 1, dt.month, dt.day)
    return alvo
def _parse_date(s: str | None) -> date | None:
    """
    Converte diversas entradas em datetime.date.
    Aceita:
      - objetos date (retorna como estÃ¡)
      - strings nos formatos mais comuns:
        YYYY-MM-DD, DD/MM/YYYY, DD-MM-YYYY, YYYY/MM/DD, YYYY.MM.DD, DD.MM.YYYY
      - nÃºmeros com 8 dÃ­gitos: YYYYMMDD ou DDMMYYYY (heurÃ­stica)
      - palavras-chave: 'hoje', 'ontem', 'amanha'/'amanhÃ£'
    Retorna None se nÃ£o conseguir interpretar.
    """
    if s is None:
        return None
    if isinstance(s, date):
        return s

    txt = str(s).strip()
    if not txt:
        return None

    low = txt.lower()
    if low in {"hoje", "today"}:
        return date.today()
    if low in {"ontem", "yesterday"}:
        return date.today().fromordinal(date.today().toordinal() - 1)
    if low in {"amanha", "amanhÃ£", "tomorrow"}:
        return date.today().fromordinal(date.today().toordinal() + 1)

    # Tenta formatos comuns
    fmts = [
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y/%m/%d",
        "%Y.%m.%d",
        "%d.%m.%Y",
    ]
    for fmt in fmts:
        try:
            return datetime.strptime(txt, fmt).date()
        except Exception:
            pass

    # HeurÃ­stica para 8 dÃ­gitos (YYYYMMDD ou DDMMYYYY)
    digits = "".join(ch for ch in txt if ch.isdigit())
    if len(digits) == 8:
        y, m, d = digits[:4], digits[4:6], digits[6:]  # tenta YYYYMMDD
        try:
            return date(int(y), int(m), int(d))
        except Exception:
            # tenta DDMMYYYY
            d, m, y = digits[:2], digits[2:4], digits[4:]
            try:
                return date(int(y), int(m), int(d))
            except Exception:
                pass

    return None
    

def _parse_ymd(s):
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None

def _bounds_mes(yyyy_mm: str):
    # "2025-06" -> [2025-06-01, 2025-07-01)
    y, m = map(int, yyyy_mm.split("-"))
    ini = date(y, m, 1)
    fim = (ini + relativedelta(months=1))
    return ini, fim
    

def _parse_data_ymd(s):
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None

def _fmt_br(d: date | None) -> str:
    return d.strftime("%d/%m/%Y") if d else ""


def _dow(dt: date) -> str:
    return str((dt.weekday() % 7) + 1)


# === Helpers de AvaliaÃ§Ã£o (NLP leve + mÃ©tricas) =============================
def _clamp_star(v):
    try:
        v = int(v)
    except Exception:
        return None
    return min(5, max(1, v))

def _media_ponderada(geral, pont, educ, efic, apres):
    """
    PonderaÃ§Ã£o (soma=1.0):
      Geral 0.40 + Pontualidade 0.15 + EducaÃ§Ã£o 0.15 + EficiÃªncia 0.15 + Bem Apresentado 0.15
    Calcula sÃ³ com os campos presentes (ignora None) e renormaliza pesos.
    """
    pares = [
        (geral, 0.40),
        (pont,  0.15),
        (educ,  0.15),
        (efic,  0.15),
        (apres, 0.15),
    ]
    num = 0.0
    den = 0.0
    for nota, w in pares:
        if nota is not None:
            num += float(nota) * w
            den += w
    return round(num / den, 2) if den > 0 else None

_POS = set("""
bom Ã³tima otimo excelente parabÃ©ns educado gentil atencioso cordial limpo cheiroso organizado rÃ¡pido rapida rapido pontual
""".split())
_NEG = set("""
ruim pÃ©ssimo pessimo horrÃ­vel horrivel sujo atrasado grosseiro mal educado agressivo impaciente amassado quebrado frio derramou
""".split())

def _analise_sentimento(txt: str | None) -> str:
    if not txt:
        return "neutro"
    t = (txt or "").lower()
    # contagem bem simples
    pos = sum(1 for w in _POS if w in t)
    neg = sum(1 for w in _NEG if w in t)
    if neg > pos + 0: return "negativo"
    if pos > neg + 0: return "positivo"
    return "neutro"

# mapeia temas por palavras-chave simples
_TEMAS = {
    "Pontualidade":  ["pontual", "atras", "horario", "horÃ¡rio", "demor", "rÃ¡pido", "rapido", "lent"],
    "EducaÃ§Ã£o":      ["educad", "grosseir", "simpat", "antipatic", "mal trat", "sem paciencia", "sem paciÃªncia", "atencios"],
    "EficiÃªncia":    ["amass", "vazou", "quebrad", "frio", "bagunÃ§a", "bagunca", "cuidado", "eficien", "desorgan"],
    "Bem apresentado": ["uniform", "higien", "apresenta", "limpo", "cheiroso", "aparencia", "aparÃªncia"],
}

def _identifica_temas(txt: str | None) -> list[str]:
    if not txt:
        return []
    t = (txt or "").lower()
    hits = []
    for tema, keys in _TEMAS.items():
        if any(k in t for k in keys):
            hits.append(tema)
    return hits[:4]

_RISCO = ["ameaÃ§a","ameaca","acidente","quebrado","agress","roubo","violÃªn","violenc","lesÃ£o","lesao","sangue","caiu","bateu","droga","alcool","Ã¡lcool"]

def _sinaliza_crise(nota_geral: int | None, txt: str | None) -> bool:
    if nota_geral == 1 and txt:
        low = txt.lower()
        return any(k in low for k in _RISCO)
    return False

def _gerar_feedback(pont, educ, efic, apres, comentario, sentimento):
    partes = []
    def badge(nome, nota):
        return f"{nome}: {nota} â˜…" if nota is not None else None

    for nome, nota in (("Pontualidade", pont), ("EducaÃ§Ã£o", educ), ("EficiÃªncia", efic), ("ApresentaÃ§Ã£o", apres)):
        b = badge(nome, nota)
        if b: partes.append(b)

    dicas = []
    if educ is not None and educ <= 2: dicas.append("melhore a abordagem/educaÃ§Ã£o ao falar com o cliente")
    if pont is not None and pont <= 2: dicas.append("tente chegar no horÃ¡rio combinado")
    if efic is not None and efic <= 2: dicas.append("redobre o cuidado com o pedido durante o transporte")
    if apres is not None and apres <= 2: dicas.append("capriche na apresentaÃ§Ã£o pessoal (higiene/uniforme)")

    txt = f"Notas â€” " + " | ".join(partes) if partes else "Obrigado pelo trabalho!"
    if comentario:
        txt += f". Cliente comentou: \"{comentario.strip()}\""
    if dicas:
        txt += ". Dica: " + "; ".join(dicas) + "."
    if sentimento == "positivo":
        txt += " ðŸ‘"
    return txt[:1000]
from sqlalchemy import or_, and_

# routes_avisos.py
from flask import Blueprint, render_template, redirect, request, url_for, abort
from flask_login import current_user


def _cooperado_atual() -> Cooperado | None:
    """
    Retorna o Cooperado do usuÃ¡rio logado, usando a sessÃ£o da aplicaÃ§Ã£o.
    """
    uid = session.get("user_id")
    if not uid:
        return None
    return Cooperado.query.filter_by(usuario_id=uid).first()

    # --- Blueprint Portal (topo do arquivo, depois de criar `app`) ---
portal_bp = Blueprint("portal", __name__, url_prefix="/portal")

@portal_bp.get("/avisos", endpoint="portal_cooperado_avisos")
@role_required("cooperado")
def avisos_list():
    coop = _cooperado_atual()
    if not coop:
        abort(403)

    # pega todos os avisos que se aplicam ao cooperado (seu helper)
    avisos = get_avisos_for_cooperado(coop)

    # busca leituras de uma vez (evita N+1)
    lidos_ids = {
        r.aviso_id
        for r in AvisoLeitura.query.filter_by(cooperado_id=coop.id).all()
    }

    # injeta flag lido para o template (sem tocar no banco)
    for a in avisos:
        a.lido = (a.id in lidos_ids)

    avisos_nao_lidos_count = sum(1 for a in avisos if not getattr(a, "lido", False))
    current_year = datetime.now().year

    return render_template(
        "portal_cooperado_avisos.html",
        avisos=avisos,
        avisos_nao_lidos_count=avisos_nao_lidos_count,
        current_year=current_year
    )

# === AVALIAÃ‡Ã•ES: Cooperado -> Restaurante (NOVO) =============================
class AvaliacaoRestaurante(db.Model):
    __tablename__ = "avaliacoes_restaurante"
    id = db.Column(db.Integer, primary_key=True)

    restaurante_id = db.Column(db.Integer, db.ForeignKey("restaurantes.id"), nullable=False, index=True)
    cooperado_id   = db.Column(db.Integer, db.ForeignKey("cooperados.id"),   nullable=False, index=True)

    # ðŸ”´ IMPORTANTE: CASCADE ao apagar o lanÃ§amento
    lancamento_id  = db.Column(
        db.Integer,
        db.ForeignKey("lancamentos.id", ondelete="CASCADE"),
        unique=True,
        index=True,
        nullable=True
    )

    # mesmas mÃ©tricas 1..5
    estrelas_geral        = db.Column(db.Integer)
    estrelas_pontualidade = db.Column(db.Integer)
    estrelas_educacao     = db.Column(db.Integer)
    estrelas_eficiencia   = db.Column(db.Integer)
    estrelas_apresentacao = db.Column(db.Integer)

    comentario      = db.Column(db.Text)
    media_ponderada = db.Column(db.Float)
    sentimento      = db.Column(db.String(12))
    temas           = db.Column(db.String(255))
    alerta_crise    = db.Column(db.Boolean, default=False)

    criado_em = db.Column(db.DateTime, default=datetime.utcnow, index=True)

@portal_bp.post("/avisos/<int:aviso_id>/lido", endpoint="marcar_aviso_lido")
@role_required("cooperado")
def avisos_marcar_lido(aviso_id: int):
    coop = _cooperado_atual()
    if not coop:
        abort(403)

    aviso = Aviso.query.get_or_404(aviso_id)

    # idempotente: sÃ³ cria se ainda nÃ£o houver leitura
    ja_leu = AvisoLeitura.query.filter_by(
        cooperado_id=coop.id,
        aviso_id=aviso.id
    ).first()

    if not ja_leu:
        db.session.add(AvisoLeitura(
            cooperado_id=coop.id,
            aviso_id=aviso.id,
            lido_em=datetime.now(timezone.utc)
        ))
        db.session.commit()

    next_url = request.form.get("next") or (url_for("portal.portal_cooperado_avisos") + f"#aviso-{aviso.id}")
    return redirect(next_url)

@portal_bp.post("/avisos/marcar-todos", endpoint="marcar_todos_avisos_lidos")
@role_required("cooperado")
def avisos_marcar_todos():
    coop = _cooperado_atual()
    if not coop:
        abort(403)

    avisos = get_avisos_for_cooperado(coop)
    if not avisos:
        return redirect(url_for("portal.portal_cooperado_avisos"))
    ids_todos = {a.id for a in avisos}
    ids_ja_lidos = {
        r.aviso_id
        for r in AvisoLeitura.query.filter_by(cooperado_id=coop.id).all()
    }
    ids_pendentes = list(ids_todos - ids_ja_lidos)

    if ids_pendentes:
        db.session.bulk_save_objects([
            AvisoLeitura(cooperado_id=coop.id, aviso_id=aid, lido_em=datetime.now(timezone.utc))
            for aid in ids_pendentes
        ])
        db.session.commit()

    return redirect(url_for("portal.portal_cooperado_avisos"))

  # --- Registro do blueprint 'portal' (uma Ãºnica vez, apÃ³s definir TODAS as rotas dele)
# --- Registro do blueprint 'portal' (depois de definir TODAS as rotas do blueprint)
def register_blueprints_once(app):
    if "portal" not in app.blueprints:
        app.register_blueprint(portal_bp)

register_blueprints_once(app)

# --- Alias para compatibilidade com o template (endpoint esperado: 'portal_cooperado_avisos')
from flask import redirect, url_for

def _portal_cooperado_avisos_alias():
    # redireciona para a rota real dentro do blueprint 'portal'
    return redirect(url_for("portal.portal_cooperado_avisos"))

# publica a URL "antiga" (ajuste o path se o seu antigo era outro)
app.add_url_rule(
    "/portal/cooperado/avisos",         # caminho acessado
    endpoint="portal_cooperado_avisos", # nome que o template usa no url_for(...)
    view_func=_portal_cooperado_avisos_alias,
    methods=["GET"],
)

# ======== Helpers p/ troca: data/weekday/turno ========
def _parse_data_escala_str(s: str) -> date | None:
    m = re.search(r'(\d{1,2})/(\d{1,2})/(\d{2,4})', str(s or ''))
    if not m:
        return None
    d_, mth, y = map(int, m.groups())
    if y < 100:
        y += 2000
    try:
        return date(y, mth, d_)
    except Exception:
        return None

def _weekday_from_data_str(s: str) -> int | None:
    dt = _parse_data_escala_str(s)
    if dt:
        return (dt.weekday() % 7) + 1
    txt = unicodedata.normalize("NFD", str(s or "").lower())
    txt = "".join(ch for ch in txt if unicodedata.category(ch) != "Mn")
    M = re.search(
        r"\b(seg|segunda|ter|terca|terÃ§a|qua|quarta|qui|quinta|sex|sexta|sab|sabado|sÃ¡bado|dom|domingo)\b",
        txt
    )
    if not M:
        M = re.search(r"\b(seg|ter|qua|qui|sex|sab|dom)\b", txt)
        if not M:
            return None
    token = M.group(1)
    mapa = {
        "seg":1,"segunda":1,
        "ter":2,"terca":2,"terÃ§a":2,
        "qua":3,"quarta":3,
        "qui":4,"quinta":4,
        "sex":5,"sexta":5,
        "sab":6,"sabado":6,"sÃ¡bado":6,
        "dom":7,"domingo":7,
    }
    return mapa.get(token)

def _weekday_abbr(num: int | None) -> str:
    return {1:"SEG",2:"TER",3:"QUA",4:"QUI",5:"SEX",6:"SÃB",7:"DOM"}.get(num or 0, "")

def _turno_bucket(turno: str | None, horario: str | None) -> str:
    t = (turno or "").lower()
    t = unicodedata.normalize("NFD", t)
    t = "".join(ch for ch in t if unicodedata.category(ch) != "Mn")
    if "noite" in t or "noturn" in t:
        return "noite"
    if any(x in t for x in ["dia", "diurn", "manha", "manhÃ£", "tarde"]):
        return "dia"
    m = re.search(r'(\d{1,2}):(\d{2})', str(horario or ""))
    if m:
        h = int(m.group(1))
        return "noite" if (h >= 17 or h <= 6) else "dia"
    return "dia"

def _escala_label(e: Escala | None) -> str:
    if not e:
        return "â€”"
    wd = _weekday_from_data_str(e.data)
    wd_abbr = _weekday_abbr(wd)
    dt = _parse_data_escala_str(e.data)
    if dt:
        data_txt = dt.strftime("%d/%m/%y") + (f"-{wd_abbr}" if wd_abbr else "")
    else:
        data_txt = (str(e.data or "").strip() or wd_abbr)
    turno_txt = (e.turno or "").strip() or _turno_bucket(e.turno, e.horario).upper()
    horario_txt = (e.horario or "").strip()
    contrato_txt = (e.contrato or "").strip()
    parts = [x for x in [data_txt, turno_txt, horario_txt, contrato_txt] if x]
    return " â€¢ ".join(parts)

def _carry_forward_contrato(escalas: list[Escala]) -> dict[int, str]:
    eff = {}
    atual = ""
    for e in escalas:
        raw = (e.contrato or "").strip()
        if raw:
            atual = raw
        eff[e.id] = atual
    return eff

def _parse_linhas_from_msg(msg: str | None) -> list[dict]:
    if not msg:
        return []
    blobs = re.findall(r"__AFETACAO_JSON__\s*[:=]\s*(\{.*?\})\s*$", str(msg), flags=re.DOTALL)
    if not blobs:
        return []
    raw = blobs[-1]
    try:
        payload = json.loads(raw)
    except Exception:
        try:
            payload = json.loads(raw.replace("'", '\"'))
        except Exception:
            return []
    linhas = payload.get("linhas") or payload.get("rows") or []
    out = []
    for r in (linhas if isinstance(linhas, list) else []):
        turno = str(r.get("turno") or "").strip()
        horario = str(r.get("horario") or "").strip()
        turno_horario = (r.get("turno_horario") or " â€¢ ".join(x for x in [turno, horario] if x)).strip()
        out.append({
            "dia": str(r.get("dia") or ""),
            "turno_horario": turno_horario,
            "contrato": str(r.get("contrato") or ""),
            "saiu": str(r.get("saiu") or ""),
            "entrou": str(r.get("entrou") or ""),
        })
    return out

def _strip_afetacao_blob(msg: str | None) -> str:
    if not msg:
        return ""
    return re.sub(r"__AFETACAO_JSON__\s*[:=]\s*\{.*\}\s*$", "", str(msg), flags=re.DOTALL).strip()

def _norm(s: str) -> str:
    s = unicodedata.normalize("NFD", str(s or "").strip().lower())
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = re.sub(r"[^a-z0-9]+", " ", s).strip()
    return s

def to_css_color(v: str) -> str:
    t = str(v or "").strip()
    if not t:
        return ""
    t_low = t.lower().strip()
    if re.fullmatch(r"[0-9a-fA-F]{8}", t):
        return f"#{t[2:8]}"
    if re.fullmatch(r"[0-9a-fA-F]{6}", t):
        return f"#{t}"
    if re.fullmatch(r"#?[0-9a-fA-F]{6,8}", t):
        if not t.startswith("#"):
            t = f"#{t}"
        if len(t) == 9:
            a = int(t[1:3], 16) / 255.0
            r = int(t[3:5], 16)
            g = int(t[5:7], 16)
            b = int(t[7:9], 16)
            return f"rgba({r},{g},{b},{a:.3f})"
        return t
    m = re.fullmatch(r"\s*(\d{1,3})\s*[,;]\s*(\d{1,3})\s*[,;]\s*(\d{1,3})\s*", t)
    if m:
        r, g, b = [max(0, min(255, int(x))) for x in m.groups()]
        return f"rgb({r},{g},{b})"
    mapa = {
        "azul": "blue", "vermelho": "red", "verde": "green",
        "amarelo": "yellow", "cinza": "gray", "preto": "black",
        "branco": "white", "laranja": "orange", "roxo": "purple",
    }
    return mapa.get(t_low, t)

# ---------- AVISOS: helpers ----------
from sqlalchemy import case, or_, and_, func
from sqlalchemy.orm import selectinload

def _avisos_base_query():
    # usa o relÃ³gio do banco; evita divergÃªncia de TZ/UTC da app
    now = func.now()
    return (
        Aviso.query
        .options(selectinload(Aviso.restaurantes))  # evita N+1 no template
        .filter(Aviso.ativo.is_(True))
        .filter(or_(Aviso.inicio_em.is_(None), Aviso.inicio_em <= now))
        .filter(or_(Aviso.fim_em.is_(None),    Aviso.fim_em    >= now))
    )

# PRIORIDADE: "alta" (0), "media"/"mÃ©dia" (1), outras/NULL (2)
_PRIORD = case(
    (func.lower(Aviso.prioridade) == "alta", 0),
    (func.lower(Aviso.prioridade).in_(("media", "mÃ©dia")), 1),
    else_=2,
)


def get_avisos_for_cooperado(coop: Cooperado):
    q = (
        _avisos_base_query()
        .filter(
            or_(
                (Aviso.tipo == "global"),
                and_(
                    Aviso.tipo == "cooperado",
                    or_(
                        Aviso.destino_cooperado_id == coop.id,
                        Aviso.destino_cooperado_id.is_(None)  # broadcast cooperados
                    ),
                ),
            )
        )
        .order_by(
            Aviso.fixado.desc(),
            _PRIORD.asc(),
            Aviso.criado_em.desc(),
        )
    )
    return q.all()

# =========================
# Helper: avisos para um restaurante/farmÃ¡cia
# =========================
def get_avisos_for_restaurante(rest: "Restaurante"):
    """
    Retorna avisos ativos para o estabelecimento:
    - tipo global
    - ou tipo restaurante direcionado especificamente ao rest (M:N) 
      ou sem destinatÃ¡rio explÃ­cito (broadcast).
    - respeita janela (inicio_em/fim_em), se definida.
    """
    now = datetime.now(timezone.utc)

    q = (
        Aviso.query
        .filter(Aviso.ativo.is_(True))
        .filter(
            # janela: (sem inÃ­cio OU inÃ­cio <= agora) E (sem fim OU fim >= agora)
            and_(
                or_(Aviso.inicio_em.is_(None), Aviso.inicio_em <= now),
                or_(Aviso.fim_em.is_(None),    Aviso.fim_em    >= now),
            )
        )
        .filter(
            or_(
                Aviso.tipo == "global",
                and_(
                    Aviso.tipo == "restaurante",
                    or_(
                        # avisos endereÃ§ados explicitamente a este restaurante
                        Aviso.restaurantes.any(Restaurante.id == rest.id),
                        # ou avisos de restaurante sem lista (broadcast)
                        ~Aviso.restaurantes.any()
                    )
                )
            )
        )
        .order_by(Aviso.fixado.desc(), Aviso.criado_em.desc())
    )
    return q.all()

# =========================
# Rotas de mÃ­dia (fotos armazenadas no banco)
# =========================
from flask import Response

def _send_bytes_with_cache(data: bytes, mime: str, filename: str):
    """Envia bytes com Cache-Control/ETag para evitar hits repetidos no banco."""
    if not data:
        abort(404)
    rv = send_file(
        io.BytesIO(data),
        mimetype=(mime or "application/octet-stream"),
        as_attachment=False,
        download_name=filename,
        max_age=60 * 60 * 24 * 7,  # 7 dias
        conditional=True,          # habilita ETag/If-None-Match
        last_modified=None,
        etag=True,
    )
    # Cache explÃ­cito (defensivo)
    rv.headers["Cache-Control"] = "public, max-age=604800, immutable"
    return rv

@app.get("/media/coop/<int:coop_id>")
def media_coop(coop_id: int):
    # Retry simples para conexÃµes quebradas (evita 500 pontual)
    try:
        c = Cooperado.query.get_or_404(coop_id)
    except OperationalError:
        db.session.rollback()
        c = Cooperado.query.get_or_404(coop_id)

    if c.foto_bytes:
        return _send_bytes_with_cache(
            c.foto_bytes,
            c.foto_mime or "image/jpeg",
            c.foto_filename or f"coop_{coop_id}.jpg",
        )
    # fallback para imagem padrÃ£o (nÃ£o bate no banco novamente)
    return redirect(url_for("static", filename="img/default.png"))

@app.get("/media/rest/<int:rest_id>")
def media_rest(rest_id: int):
    try:
        r = Restaurante.query.get_or_404(rest_id)
    except OperationalError:
        db.session.rollback()
        r = Restaurante.query.get_or_404(rest_id)

    if r.foto_bytes:
        return _send_bytes_with_cache(
            r.foto_bytes,
            r.foto_mime or "image/jpeg",
            r.foto_filename or f"rest_{rest_id}.jpg",
        )
    return redirect(url_for("static", filename="img/default.png"))

# =========================
# Rota raiz
# =========================
@app.route("/")
def index():
    uid = session.get("user_id")
    if not uid:
        return redirect(url_for("login"))

    u = db.session.get(Usuario, uid)
    if not u or not getattr(u, "ativo", True):
        session.clear()
        return redirect(url_for("login"))

    def _try_redirect(*endpoints):
        for ep in endpoints:
            try:
                return redirect(url_for(ep))
            except Exception:
                pass
        return None

    if u.tipo == "admin":
        r = _try_redirect("admin", "dashboard_admin")
        return r or redirect(url_for("login"))

    if u.tipo == "cooperado":
        r = _try_redirect("painel_cooperado", "portal_cooperado", "portal_cooperado_home")
        return r or redirect(url_for("login"))

    if u.tipo in ("restaurante", "farmacia"):
        r = _try_redirect("painel_restaurante", "portal_restaurante", "restaurante_painel")
        return r or redirect(url_for("login"))

    return redirect(url_for("login"))

# =========================
# Auth
# =========================
# =========================
# Auth
# =========================
def _safe_norm(s: str) -> str:
    """Normaliza para comparar logins: sem acento, minÃºsculo, sem espaÃ§os extras."""
    try:
        return _norm_login(s)  # usa a sua funÃ§Ã£o jÃ¡ definida no arquivo
    except NameError:
        import unicodedata, re
        s = unicodedata.normalize("NFD", s or "")
        s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
        s = s.lower().strip()
        s = re.sub(r"\s+", "", s)
        return s

def _find_user_for_login(login_input: str) -> Usuario | None:
    """
    Procura o usuÃ¡rio para autenticaÃ§Ã£o:
      1) match direto por Usuario.usuario (case-insensitive)
      2) match por normalizaÃ§Ã£o (_norm_login/_safe_norm)
      3) fallback: login pelo NOME do restaurante (exato ou normalizado)
    """
    if not login_input:
        return None

    raw = (login_input or "").strip()
    key = _safe_norm(raw)

    # 1) direto (case-insensitive)
    try:
        u = Usuario.query.filter(func.lower(Usuario.usuario) == raw.lower()).first()
        if u:
            return u
    except Exception:
        pass

    # 2) por normalizaÃ§Ã£o forte
    try:
        for cand in Usuario.query.filter(Usuario.ativo == True).all():
            if _safe_norm(cand.usuario) == key:
                return cand
    except Exception:
        pass

    # 3) pelo restaurante (exato case-insensitive)
    try:
        r = Restaurante.query.filter(func.lower(Restaurante.nome) == raw.lower()).first()
        if r and r.usuario_ref:
            return r.usuario_ref
    except Exception:
        r = None

    # 3b) pelo restaurante (normalizado)
    try:
        for r0 in Restaurante.query.all():
            if _safe_norm(r0.nome) == key and r0.usuario_ref:
                return r0.usuario_ref
    except Exception:
        pass

    return None


# =========================
# Helpers de login
# =========================

# Garante que existe um normalizador de login (sem acentos/espaÃ§os)
if "_norm_login" not in globals():
    def _norm_login(s: str) -> str:
        import unicodedata, re
        s = unicodedata.normalize("NFD", s or "")
        s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
        s = s.lower().strip()
        s = re.sub(r"\s+", "", s)
        return s

def _find_user_for_login(usuario_input: str) -> "Usuario | None":
    """
    Tenta localizar o usuÃ¡rio de login por:
      1) Usuario.usuario (case-insensitive)
      2) Usuario.usuario (normalizado por _norm_login)
      3) Nome do Restaurante (direto e normalizado) -> retorna r.usuario_ref
    """
    if not usuario_input:
        return None

    key_raw  = usuario_input.strip()
    key_norm = _norm_login(key_raw)

    # 1) match direto case-insensitive
    try:
        u = Usuario.query.filter(func.lower(Usuario.usuario) == key_raw.lower()).first()
        if u:
            return u
    except Exception:
        pass

    # 2) varrer usuÃ¡rios ativos e comparar por normalizaÃ§Ã£o
    try:
        for u in Usuario.query.filter(Usuario.ativo == True).all():
            if _norm_login(u.usuario) == key_norm:
                return u
    except Exception:
        pass

    # 3) permitir login pelo NOME do restaurante
    try:
        # direto (case-insensitive)
        r = Restaurante.query.filter(func.lower(Restaurante.nome) == key_raw.lower()).first()
        if not r:
            # normalizado
            for r0 in Restaurante.query.all():
                if _norm_login(r0.nome) == key_norm:
                    r = r0
                    break
        if r and r.usuario_ref:
            return r.usuario_ref
    except Exception:
        pass

    return None

def _inferir_e_corrigir_tipo_usuario(u: "Usuario") -> "str | None":
    """
    Se o tipo estiver vazio/desconhecido, tenta inferir pelo vÃ­nculo e corrige no banco.
    Retorna o tipo final ('admin' | 'cooperado' | 'restaurante' | 'farmacia') ou None.
    """
    if not u:
        return None

    t = (u.tipo or "").strip().lower()
    if t in {"admin", "cooperado", "restaurante", "farmacia"}:
        return t

    try:
        if Restaurante.query.filter_by(usuario_id=u.id).first():
            u.tipo = "restaurante"
            db.session.commit()
            return "restaurante"
        if Cooperado.query.filter_by(usuario_id=u.id).first():
            u.tipo = "cooperado"
            db.session.commit()
            return "cooperado"
    except Exception:
        db.session.rollback()
        return None

    return None


# =========================
# Auth
# =========================
@app.route("/login", methods=["GET", "POST"])
def login():
    erro_login = None

    if request.method == "POST":
        usuario = (request.form.get("usuario") or "").strip()
        senha   = request.form.get("senha") or ""

        u = _find_user_for_login(usuario)

        if u and u.check_password(senha):
            # bloqueia contas inativas
            if not getattr(u, "ativo", True):
                flash("Conta desativada. Fale com o administrador.", "danger")
                return redirect(url_for("login"))

            # autentica
            session.clear()
            session.permanent = True
            session["user_id"] = u.id
            session["user_tipo"] = u.tipo

            # direciona por tipo (inclui farmacia) + fallback de inferÃªncia
            tipo = (u.tipo or "").strip().lower()
            dest_por_tipo = {
                "admin": "admin_dashboard",
                "cooperado": "portal_cooperado",
                "restaurante": "portal_restaurante",
                "farmacia": "portal_restaurante",  # farmÃ¡cia usa o mesmo portal do restaurante
            }
            destino = dest_por_tipo.get(tipo)

            if not destino:
                # tenta inferir e corrigir tipo no banco
                tipo = _inferir_e_corrigir_tipo_usuario(u)
                destino = dest_por_tipo.get(tipo)

            if destino:
                return redirect(url_for(destino))

            # tipo desconhecido -> volta pro login
            flash("Tipo de usuÃ¡rio invÃ¡lido. Contate o administrador.", "danger")
            return redirect(url_for("login"))

        # falha de autenticaÃ§Ã£o
        erro_login = "UsuÃ¡rio/senha invÃ¡lidos."
        flash(erro_login, "danger")

    # Render do formulÃ¡rio
    login_tpl = os.path.join("templates", "login.html")
    if os.path.exists(login_tpl):
        return render_template("login.html", erro_login=erro_login)

    # fallback simples (se nÃ£o tiver template)
    return """
    <form method="POST" style="max-width:320px;margin:80px auto;font-family:Arial">
      <h3>Login</h3>
      <input name="usuario" placeholder="UsuÃ¡rio ou Restaurante" style="width:100%;padding:10px;margin:6px 0">
      <input name="senha" type="password" placeholder="Senha" style="width:100%;padding:10px;margin:6px 0">
      <button style="padding:10px 16px">Entrar</button>
    </form>
    """


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))
    

# =========================
# Admin Dashboard
# =========================
# app.py
from flask import jsonify, request

@app.post("/admin/cooperados/<int:id>/toggle-status")
def toggle_status_cooperado(id):
    try:
        coop = db.session.get(Cooperado, id)
        if not coop or not coop.usuario_ref:
            return jsonify(ok=False, error="Cooperado nÃ£o encontrado"), 404

        user = coop.usuario_ref

        # se o atributo nÃ£o existe (cÃ³digo antigo em produÃ§Ã£o), trate como True
        atual = bool(getattr(user, "ativo", True))
        if not hasattr(user, "ativo"):
            # tenta criar no ar (nÃ£o persiste coluna, sÃ³ evita quebrar)
            # mas avisa o caller para vocÃª rodar a migraÃ§Ã£o
            return jsonify(ok=False, error="Campo 'ativo' ausente no modelo/DB. FaÃ§a deploy com o modelo atualizado e a migraÃ§Ã£o."), 500

        user.ativo = not atual
        db.session.commit()
        return jsonify(ok=True, ativo=bool(user.ativo))
    except SQLAlchemyError:
        db.session.rollback()
        return jsonify(ok=False, error="Falha ao salvar no banco"), 500

@app.route("/admin", methods=["GET"])
@admin_required
def admin_dashboard():
    from collections import namedtuple
    import re
    args = request.args

    # --- Controle de abas
    active_tab = args.get("tab", "lancamentos")

    # --- Helper: escolhe a primeira data vÃ¡lida de uma lista de chaves
    def _pick_date(*keys):
        for k in keys:
            v = args.get(k)
            if v:
                d = _parse_date(v)
                if d:
                    return d
        return None

    # --- Datas unificadas para o RESUMO (prioriza resumo_inicio/fim)
    data_inicio = _pick_date("resumo_inicio", "data_inicio")
    data_fim = _pick_date("resumo_fim", "data_fim")

    restaurante_id = args.get("restaurante_id", type=int)
    cooperado_id = args.get("cooperado_id", type=int)
    considerar_periodo = bool(args.get("considerar_periodo"))
    dows = set(args.getlist("dow"))  # {"1","2",...}

    # ---- LanÃ§amentos (com filtros + DOW)
    q = Lancamento.query
    if restaurante_id:
        q = q.filter(Lancamento.restaurante_id == restaurante_id)
    if cooperado_id:
        q = q.filter(Lancamento.cooperado_id == cooperado_id)
    if data_inicio:
        q = q.filter(Lancamento.data >= data_inicio)
    if data_fim:
        q = q.filter(Lancamento.data <= data_fim)
    lanc_base = q.order_by(Lancamento.data.desc(), Lancamento.id.desc()).all()

    if dows:
        lancamentos = [l for l in lanc_base if l.data and _dow(l.data) in dows]
    else:
        lancamentos = lanc_base

    # Se marcar "considerar_periodo", sÃ³ mantemos dias do perÃ­odo do restaurante
    if considerar_periodo and restaurante_id:
        rest = db.session.get(Restaurante, restaurante_id)
        if rest:
            mapa = {
                "seg-dom": {"1", "2", "3", "4", "5", "6", "7"},
                "sab-sex": {"6", "7", "1", "2", "3", "4", "5"},
                "sex-qui": {"5", "6", "7", "1", "2", "3", "4"},
            }
            permitidos = mapa.get(rest.periodo, {"1", "2", "3", "4", "5", "6", "7"})
            lancamentos = [l for l in lancamentos if l.data and _dow(l.data) in permitidos]

    total_producoes = sum((l.valor or 0.0) for l in lancamentos)
    total_inss = total_producoes * 0.045

    # ---- Coop (institucional)
    rq = ReceitaCooperativa.query
    dq = DespesaCooperativa.query
    if data_inicio:
        rq = rq.filter(ReceitaCooperativa.data >= data_inicio)
        dq = dq.filter(DespesaCooperativa.data >= data_inicio)
    if data_fim:
        rq = rq.filter(ReceitaCooperativa.data <= data_fim)
        dq = dq.filter(DespesaCooperativa.data <= data_fim)

    receitas = rq.order_by(ReceitaCooperativa.data.desc().nullslast(), ReceitaCooperativa.id.desc()).all()
    despesas = dq.order_by(DespesaCooperativa.data.desc(), DespesaCooperativa.id.desc()).all()
    total_receitas = sum((r.valor_total or 0.0) for r in receitas)
    total_despesas = sum((d.valor or 0.0) for d in despesas)

    # ---- Cooperados (pessoa fÃ­sica)
    rq2 = ReceitaCooperado.query
    dq2 = DespesaCooperado.query

    # Receitas (pontuais): filtra por data simples
    if data_inicio:
        rq2 = rq2.filter(ReceitaCooperado.data >= data_inicio)
    if data_fim:
        rq2 = rq2.filter(ReceitaCooperado.data <= data_fim)

    # Despesas (semanais): usa sobreposiÃ§Ã£o do intervalo [data_inicio, data_fim]
    if data_inicio and data_fim:
        dq2 = dq2.filter(
            DespesaCooperado.data_inicio <= data_fim,
            DespesaCooperado.data_fim    >= data_inicio,
        )
    elif data_inicio:
        dq2 = dq2.filter(DespesaCooperado.data_fim >= data_inicio)
    elif data_fim:
        dq2 = dq2.filter(DespesaCooperado.data_inicio <= data_fim)

    receitas_coop = rq2.order_by(ReceitaCooperado.data.desc(), ReceitaCooperado.id.desc()).all()
    # vocÃª pode manter por .data (domingo) ou, se preferir, ordenar pelo fim real do perÃ­odo:
    despesas_coop = dq2.order_by(DespesaCooperado.data_fim.desc().nullslast(), DespesaCooperado.id.desc()).all()

    total_receitas_coop = sum((r.valor or 0.0) for r in receitas_coop)
    total_despesas_coop = sum((d.valor or 0.0) for d in despesas_coop)

    cfg = get_config()
    cooperados = Cooperado.query.order_by(Cooperado.nome).all()
    restaurantes = Restaurante.query.order_by(Restaurante.nome).all()

    # documentos OK?
    docinfo_map = {c.id: _build_docinfo(c) for c in cooperados}
    status_doc_por_coop = {
        c.id: {
            "cnh_ok": docinfo_map[c.id]["cnh"]["ok"],
            "placa_ok": docinfo_map[c.id]["placa"]["ok"],
        }
        for c in cooperados
    }

      # -------- Escalas agrupadas e contagem por cooperado ----------
    escalas_all = Escala.query.order_by(Escala.id.asc()).all()
    esc_by_int: dict[int, list] = defaultdict(list)
    esc_by_str: dict[str, list] = defaultdict(list)
    for e in escalas_all:
        k_int = e.cooperado_id if e.cooperado_id is not None else 0  # 0 = sem cadastro
        esc_item = {
            "data": e.data,
            "turno": e.turno,
            "horario": e.horario,
            "contrato": e.contrato,
            "cor": e.cor,
            "nome_planilha": e.cooperado_nome,
        }
        esc_by_int[k_int].append(esc_item)
        esc_by_str[str(k_int)].append(esc_item)

    cont_rows = dict(
        db.session.query(Escala.cooperado_id, func.count(Escala.id))
        .group_by(Escala.cooperado_id)
        .all()
    )
    qtd_escalas_map = {c.id: int(cont_rows.get(c.id, 0)) for c in cooperados}
    qtd_sem_cadastro = int(cont_rows.get(None, 0))

    # ---- GrÃ¡ficos (por mÃªs) â€” rÃ³tulo robusto "MM/YY"
    sums = {}
    for l in lancamentos:
        if not l.data:
            continue
        key = l.data.strftime("%Y-%m")  # sempre YYYY-MM
        sums[key] = sums.get(key, 0.0) + (l.valor or 0.0)

    labels_ord = sorted(sums.keys())

    def _fmt_label(k: str) -> str:
        # aceita "YYYY-MM" ou "YY-MM" e mostra "MM/YY"
        parts = k.split("-")
        if len(parts) == 2 and parts[0] and parts[1]:
            year, month = parts[0], parts[1]
            return f"{month}/{year[-2:]}"
        return k

    labels_fmt = [_fmt_label(k) for k in labels_ord]
    values = [round(sums[k], 2) for k in labels_ord]
    chart_data_lancamentos_coop = {"labels": labels_fmt, "values": values}
    chart_data_lancamentos_cooperados = chart_data_lancamentos_coop

    admin_user = Usuario.query.filter_by(tipo="admin").first()

    # ---- Folha (Ãºltimos 30 dias padrÃ£o)
    folha_inicio = _parse_date(args.get("folha_inicio")) or (date.today() - timedelta(days=30))
    folha_fim = _parse_date(args.get("folha_fim")) or date.today()
    FolhaItem = namedtuple("FolhaItem", "cooperado lancamentos receitas despesas bruto inss outras_desp liquido")

    folha_por_coop = []
    for c in cooperados:
        l = (
            Lancamento.query.filter(
                Lancamento.cooperado_id == c.id,
                Lancamento.data >= folha_inicio,
                Lancamento.data <= folha_fim,
            )
            .order_by(Lancamento.data.asc(), Lancamento.id.asc())
            .all()
        )
        r = (
            ReceitaCooperado.query.filter(
                ReceitaCooperado.cooperado_id == c.id,
                ReceitaCooperado.data >= folha_inicio,
                ReceitaCooperado.data <= folha_fim,
            )
            .order_by(ReceitaCooperado.data.asc(), ReceitaCooperado.id.asc())
            .all()
        )
        d = (
            DespesaCooperado.query.filter(
                (DespesaCooperado.cooperado_id == c.id) | (DespesaCooperado.cooperado_id.is_(None)),
                DespesaCooperado.data >= folha_inicio,
                DespesaCooperado.data <= folha_fim,
            )
            .order_by(DespesaCooperado.data.asc(), DespesaCooperado.id.asc())
            .all()
        )

        bruto_lanc = sum(x.valor or 0 for x in l)
        inss = round(bruto_lanc * 0.045, 2)
        outras_desp = sum(x.valor or 0 for x in d)
        bruto_total = bruto_lanc + sum(x.valor or 0 for x in r)
        liquido = bruto_total - inss - outras_desp

        # anotaÃ§Ãµes usadas no template
        for x in l:
            x.conta_inss = True
            x.isento_benef = False
            x.inss = round((x.valor or 0) * 0.045, 2)

        folha_por_coop.append(
            FolhaItem(
                cooperado=c,
                lancamentos=l,
                receitas=r,
                despesas=d,
                bruto=bruto_total,
                inss=inss,
                outras_desp=outras_desp,
                liquido=liquido,
            )
        )

    # ----------------------------
    # BenefÃ­cios para template (com filtros + id)
    # ----------------------------
    def _tokenize(s: str):
        return [x.strip() for x in re.split(r"[;,]", s or "") if x.strip()]

    def _d(s):
        if not s:
            return None
        s = s.strip()
        try:
            if "/" in s:  # dd/mm/yyyy
                d, m, y = s.split("/")
                return date(int(y), int(m), int(d))
            # yyyy-mm-dd
            y, m, d = s.split("-")
            return date(int(y), int(m), int(d))
        except Exception:
            return None

    # filtros vindos da querystring da prÃ³pria aba (jÃ¡ existem no seu HTML)
    b_ini = _d(request.args.get("b_ini"))
    b_fim = _d(request.args.get("b_fim"))
    coop_filter = request.args.get("coop_benef_id", type=int)

    q = BeneficioRegistro.query

    # sobreposiÃ§Ã£o de intervalo:
    # inclui o benefÃ­cio se [data_inicial, data_final] INTERSECTA o filtro
    if b_ini and b_fim:
        q = q.filter(
            BeneficioRegistro.data_inicial <= b_fim,
            BeneficioRegistro.data_final   >= b_ini,
        )
    elif b_ini:
        q = q.filter(BeneficioRegistro.data_final >= b_ini)
    elif b_fim:
        q = q.filter(BeneficioRegistro.data_inicial <= b_fim)

    historico_beneficios = q.order_by(BeneficioRegistro.id.desc()).all()

    beneficios_view = []
    for b in historico_beneficios:
        nomes = _tokenize(b.recebedores_nomes or "")
        ids   = _tokenize(b.recebedores_ids or "")

        recs = []
        for i, nome in enumerate(nomes):
            rid = None
            if i < len(ids) and str(ids[i]).isdigit():
                try:
                    rid = int(ids[i])
                except Exception:
                    rid = None

            # se o filtro por cooperado estiver ativo, sÃ³ mantÃ©m o recebedor alvo
            if coop_filter and (rid is not None) and (rid != coop_filter):
                continue

            recs.append({"id": rid, "nome": nome})

        # se o filtro por cooperado estiver ativo e nenhum recebedor bateu, pula o registro
        if coop_filter and not recs:
            continue

        beneficios_view.append({
            "id": b.id,  # <<< necessÃ¡rio para editar/excluir
            "data_inicial": b.data_inicial,
            "data_final": b.data_final,
            "data_lancamento": b.data_lancamento,
            "tipo": b.tipo,
            "valor_total": b.valor_total or 0.0,
            "recebedores": recs,
        })

    # ======== Trocas no admin ========
    def _escala_desc(e: Escala | None) -> str:
        return _escala_label(e)

    def _split_turno_horario(s: str) -> tuple[str, str]:
        if not s:
            return "", ""
        parts = [p.strip() for p in s.split("â€¢")]
        if len(parts) == 2:
            return parts[0], parts[1]
        return s.strip(), ""

    def _linha_from_escala(e: Escala, saiu: str, entrou: str) -> dict:
        return {
            "dia": _escala_label(e).split(" â€¢ ")[0],
            "turno_horario": " â€¢ ".join([x for x in [(e.turno or "").strip(), (e.horario or "").strip()] if x]),
            "contrato": (e.contrato or "").strip(),
            "saiu": saiu,
            "entrou": entrou,
        }

    trocas_all = TrocaSolicitacao.query.order_by(TrocaSolicitacao.id.desc()).all()
    trocas_pendentes, trocas_historico = [], []
    trocas_historico_flat = []

    for t in trocas_all:
        solicitante = db.session.get(Cooperado, t.solicitante_id)
        destinatario = db.session.get(Cooperado, t.destino_id)
        orig = db.session.get(Escala, t.origem_escala_id)

        linhas_afetadas = _parse_linhas_from_msg(t.mensagem) if t.status == "aprovada" else []

        if t.status == "aprovada" and not linhas_afetadas and orig and solicitante and destinatario:
            # linha 1 (origem)
            linhas_afetadas.append(_linha_from_escala(orig, saiu=solicitante.nome, entrou=destinatario.nome))
            # linha 2 (melhor candidata do solicitante no mesmo bucket)
            wd_o = _weekday_from_data_str(orig.data)
            buck_o = _turno_bucket(orig.turno, orig.horario)
            candidatas = Escala.query.filter_by(cooperado_id=solicitante.id).all()
            best = None
            for e in candidatas:
                if _weekday_from_data_str(e.data) == wd_o and _turno_bucket(e.turno, e.horario) == buck_o:
                    if (orig.contrato or "").strip().lower() == (e.contrato or "").strip().lower():
                        best = e
                        break
                    if best is None:
                        best = e
            if best:
                linhas_afetadas.append(_linha_from_escala(best, saiu=destinatario.nome, entrou=solicitante.nome))

        item = {
            "id": t.id,
            "status": t.status,
            "mensagem": t.mensagem,
            "criada_em": t.criada_em,
            "aplicada_em": t.aplicada_em,
            "solicitante": solicitante,
            "destinatario": destinatario,
            "origem": orig,
            "destino": destinatario,
            "origem_desc": _escala_desc(orig),
            "origem_weekday": _weekday_from_data_str(orig.data) if orig else None,
            "origem_turno_bucket": _turno_bucket(orig.turno if orig else None, orig.horario if orig else None),
            "linhas_afetadas": linhas_afetadas,
        }

        if t.status == "aprovada" and linhas_afetadas:
            itens = []
            for r in linhas_afetadas:
                turno_txt, horario_txt = _split_turno_horario(r.get("turno_horario", ""))
                itens.append(
                    {
                        "data": r.get("dia", ""),
                        "turno": turno_txt,
                        "horario": horario_txt,
                        "contrato": r.get("contrato", ""),
                        "saiu_nome": r.get("saiu", ""),
                        "entrou_nome": r.get("entrou", ""),
                    }
                )
                trocas_historico_flat.append(
                    {
                        "data": r.get("dia", ""),
                        "turno": turno_txt,
                        "horario": horario_txt,
                        "contrato": r.get("contrato", ""),
                        "saiu_nome": r.get("saiu", ""),
                        "entrou_nome": r.get("entrou", ""),
                        "aplicada_em": t.aplicada_em,
                    }
                )
            item["itens"] = itens

        (trocas_pendentes if t.status == "pendente" else trocas_historico).append(item)

    current_date = date.today()
    data_limite = date(current_date.year, 12, 31)

    # ---- Render
    return render_template(
        "admin_dashboard.html",
        tab=active_tab,  # <- mantÃ©m a aba ativa na UI
        total_producoes=total_producoes,
        total_inss=total_inss,
        total_receitas=total_receitas,
        total_despesas=total_despesas,
        total_receitas_coop=total_receitas_coop,
        total_despesas_coop=total_despesas_coop,
        salario_minimo=cfg.salario_minimo or 0.0,
        lancamentos=lancamentos,
        receitas=receitas,
        despesas=despesas,
        receitas_coop=receitas_coop,
        despesas_coop=despesas_coop,
        cooperados=cooperados,
        restaurantes=restaurantes,
        beneficios_view=beneficios_view,
        historico_beneficios=historico_beneficios,
        current_date=current_date,
        data_limite=data_limite,
        admin=admin_user,
        docinfo_map=docinfo_map,
        escalas_por_coop=esc_by_int,
        escalas_por_coop_json=esc_by_str,
        qtd_escalas_map=qtd_escalas_map,
        qtd_escalas_sem_cadastro=qtd_sem_cadastro,
        status_doc_por_coop=status_doc_por_coop,
        chart_data_lancamentos_coop=chart_data_lancamentos_coop,
        chart_data_lancamentos_cooperados=chart_data_lancamentos_cooperados,
        folha_inicio=folha_inicio,
        folha_fim=folha_fim,
        folha_por_coop=folha_por_coop,
        trocas_pendentes=trocas_pendentes,
        trocas_historico=trocas_historico,
        trocas_historico_flat=trocas_historico_flat,
    )

# =========================
# NavegaÃ§Ã£o/Export util
# =========================
@app.route("/filtrar_lancamentos")
@admin_required
def filtrar_lancamentos():
    qs = request.query_string.decode("utf-8")
    base = url_for("admin_dashboard", tab="lancamentos")
    joiner = "&" if qs else ""
    return redirect(f"{base}{joiner}{qs}")
    
@app.route("/exportar_lancamentos")
@admin_required
def exportar_lancamentos():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    args = request.args
    restaurante_id = args.get("restaurante_id", type=int)
    cooperado_id   = args.get("cooperado_id", type=int)
    data_inicio    = _parse_date(args.get("data_inicio"))
    data_fim       = _parse_date(args.get("data_fim"))
    dows           = set(args.getlist("dow"))

    q = Lancamento.query
    if restaurante_id:
        q = q.filter(Lancamento.restaurante_id == restaurante_id)
    if cooperado_id:
        q = q.filter(Lancamento.cooperado_id == cooperado_id)
    if data_inicio:
        q = q.filter(Lancamento.data >= data_inicio)
    if data_fim:
        q = q.filter(Lancamento.data <= data_fim)

    lancs = q.order_by(Lancamento.data.desc(), Lancamento.id.desc()).all()
    if dows:
        lancs = [l for l in lancs if l.data and _dow(l.data) in dows]

    # ---- Monta o Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "LanÃ§amentos"

    header = [
        "Restaurante", "Periodo", "Cooperado", "Descricao",
        "Valor", "Data", "HoraInicio", "HoraFim", "INSS", "Liquido",
    ]
    ws.append(header)

    # Estilo do header
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    fill = PatternFill("solid", fgColor="DDDDDD")
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, _ in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold
        cell.alignment = center
        cell.fill = fill
        cell.border = border

    # Linhas
    for l in lancs:
        v = float(l.valor or 0.0)
        inss = v * 0.045
        liq = v - inss

        row = [
            l.restaurante.nome if l.restaurante else "",
            l.restaurante.periodo if l.restaurante else "",
            l.cooperado.nome if l.cooperado else "",
            (l.descricao or ""),
            v,
            l.data,  # se None, vai vazio
            (l.hora_inicio or ""),
            (l.hora_fim or ""),
            inss,
            liq,
        ]
        ws.append(row)

    # Formatos de nÃºmero/data e largura das colunas
    # Ãndices das colunas (1-based): Valor=5, Data=6, INSS=9, Liquido=10
    currency_fmt = "0.00"
    date_fmt = "DD/MM/YYYY"

    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=5).number_format = currency_fmt
        ws.cell(row=r, column=6).number_format = date_fmt
        ws.cell(row=r, column=9).number_format = currency_fmt
        ws.cell(row=r, column=10).number_format = currency_fmt

    # Ajuste simples de largura (auto-fit aproximado)
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter]:
            try:
                txt = str(cell.value) if cell.value is not None else ""
            except Exception:
                txt = ""
            max_len = max(max_len, len(txt))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 50)

    # Congelar cabeÃ§alho
    ws.freeze_panes = "A2"

    # Borda nas cÃ©lulas de dados
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).border = border

    # ---- Retorna como download .xlsx
    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)

    return send_file(
        mem,
        as_attachment=True,
        download_name="lancamentos.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# =========================
# CRUD LanÃ§amentos (Admin)
# =========================
@app.route("/admin/lancamentos/add", methods=["POST"])
@admin_required
def admin_add_lancamento():
    f = request.form
    l = Lancamento(
        restaurante_id=f.get("restaurante_id", type=int),
        cooperado_id=f.get("cooperado_id", type=int),
        descricao=f.get("descricao", "").strip(),
        valor=f.get("valor", type=float),
        data=_parse_date(f.get("data")),
        hora_inicio=f.get("hora_inicio"),
        hora_fim=f.get("hora_fim"),
        qtd_entregas=f.get("qtd_entregas", type=int),
    )
    db.session.add(l)
    db.session.commit()
    flash("LanÃ§amento inserido.", "success")
    return redirect(url_for("admin_dashboard", tab="lancamentos"))

@app.route("/admin/lancamentos/<int:id>/edit", methods=["POST"])
@admin_required
def admin_edit_lancamento(id):
    l = Lancamento.query.get_or_404(id)
    f = request.form
    l.restaurante_id = f.get("restaurante_id", type=int)
    l.cooperado_id = f.get("cooperado_id", type=int)
    l.descricao = f.get("descricao", "").strip()
    l.valor = f.get("valor", type=float)
    l.data = _parse_date(f.get("data"))
    l.hora_inicio = f.get("hora_inicio")
    l.hora_fim = f.get("hora_fim")
    l.qtd_entregas = f.get("qtd_entregas", type=int)
    db.session.commit()
    flash("LanÃ§amento atualizado.", "success")
    return redirect(url_for("admin_dashboard", tab="lancamentos"))

@app.route("/admin/lancamentos/<int:id>/delete")
@admin_required
def admin_delete_lancamento(id):
    l = Lancamento.query.get_or_404(id)

    # ðŸ‘‡ LIMPEZA MANUAL: apaga avaliaÃ§Ãµes amarradas a este lanÃ§amento
    db.session.execute(sa_delete(AvaliacaoCooperado).where(AvaliacaoCooperado.lancamento_id == id))
    db.session.execute(sa_delete(AvaliacaoRestaurante).where(AvaliacaoRestaurante.lancamento_id == id))

    db.session.delete(l)
    db.session.commit()
    flash("LanÃ§amento excluÃ­do.", "success")
    return redirect(url_for("admin_dashboard", tab="lancamentos"))

# ===== IMPORTS =====
from flask import request, render_template, send_file, url_for
from sqlalchemy import func, literal, and_
from types import SimpleNamespace
import io, csv

# ===== IMPORTS =====
from flask import request, render_template, send_file, url_for
from sqlalchemy import func, literal, and_
from types import SimpleNamespace
import io, csv
@app.route("/admin/avaliacoes", methods=["GET"])
@admin_required
def admin_avaliacoes():
    # tipo=cooperado (padrÃ£o): Restaurante avalia Cooperado
    # tipo=restaurante: Cooperado avalia Restaurante
    tipo_raw = (request.args.get("tipo") or "cooperado").strip().lower()
    tipo = "restaurante" if tipo_raw == "restaurante" else "cooperado"

    restaurante_id = request.args.get("restaurante_id", type=int)
    cooperado_id   = request.args.get("cooperado_id", type=int)
    data_inicio    = (request.args.get("data_inicio") or "").strip()
    data_fim       = (request.args.get("data_fim") or "").strip()

    # Datas (aceita YYYY-MM-DD ou DD/MM/YYYY)
    di = _parse_date(data_inicio)
    df = _parse_date(data_fim)

    # Model por tipo
    Model = AvaliacaoRestaurante if (tipo == "restaurante") else AvaliacaoCooperado

    # Helper: checa se coluna existe (migraÃ§Ãµes)
    def col(*names):
        for n in names:
            if hasattr(Model, n):
                return getattr(Model, n)
        return None

    f_geral = col("estrelas_geral")

    if tipo == "restaurante":  # Cooperado -> Restaurante
        f_trat = col("estrelas_tratamento", "estrelas_pontualidade")
        f_amb  = col("estrelas_ambiente",   "estrelas_educacao")
        f_sup  = col("estrelas_suporte",    "estrelas_eficiencia")
    else:                       # Restaurante -> Cooperado
        f_pont  = col("estrelas_pontualidade")
        f_educ  = col("estrelas_educacao")
        f_efic  = col("estrelas_eficiencia")
        f_apres = col("estrelas_apresentacao")

    # Query base (sem relationships no template)
    base = (
        db.session.query(
            Model,
            Restaurante.id.label("rest_id"),
            Restaurante.nome.label("rest_nome"),
            Cooperado.id.label("coop_id"),
            Cooperado.nome.label("coop_nome"),
        )
        .join(Restaurante, Model.restaurante_id == Restaurante.id)
        .join(Cooperado,   Model.cooperado_id   == Cooperado.id)
    )

    # Filtros
    filtros = []
    if restaurante_id:
        filtros.append(Model.restaurante_id == restaurante_id)
    if cooperado_id:
        filtros.append(Model.cooperado_id == cooperado_id)
    # Para nÃ£o depender de datetime/time aqui, use func.date() (robusto e simples)
    if di:
        filtros.append(func.date(Model.criado_em) >= di.isoformat())
    if df:
        filtros.append(func.date(Model.criado_em) <= df.isoformat())

    if filtros:
        base = base.filter(and_(*filtros))

    # PaginaÃ§Ã£o
    page = max(1, request.args.get("page", type=int) or 1)
    per_page = min(200, max(1, request.args.get("per_page", type=int) or 50))
    offset = (page - 1) * per_page

    rows = base.order_by(Model.criado_em.desc()).limit(per_page).offset(offset).all()

    # Total
    cnt_q = db.session.query(func.count(Model.id))
    if filtros:
        cnt_q = cnt_q.filter(and_(*filtros))
    total = int(cnt_q.scalar() or 0)
    pages = max(1, (total + per_page - 1) // per_page)

    pager = SimpleNamespace(
        page=page, per_page=per_page, total=total, pages=pages,
        has_prev=(page > 1), has_next=(page < pages)
    )

    # Achata para o template â€” **sempre** com todas as chaves usadas na UI
    avaliacoes = []
    for a, rest_id, rest_nome, coop_id, coop_nome in rows:
        item = {
            "criado_em": a.criado_em,
            "rest_id":   rest_id,
            "rest_nome": rest_nome,
            "coop_id":   coop_id,
            "coop_nome": coop_nome,

            "geral":      getattr(a, "estrelas_geral", 0) or 0,
            "comentario": (getattr(a, "comentario", "") or "").strip(),
            "media":       getattr(a, "media_ponderada", None),
            "sentimento":  getattr(a, "sentimento", None),
            "temas":       getattr(a, "temas", None),
            "alerta":      bool(getattr(a, "alerta_crise", False)),

            # chaves "fixas" para nÃ£o quebrar JS/Jinja em nenhuma aba
            "tratamento": 0, "ambiente": 0, "suporte": 0,
            "trat": 0, "amb": 0, "sup": 0,  # retrocompat
            "pont": 0, "educ": 0, "efic": 0, "apres": 0,
        }

        if tipo == "restaurante":
            # Cooperado -> Restaurante: garante **Tratamento** visÃ­vel
            trat = getattr(a, "estrelas_tratamento", None)
            if trat is None:
                # fallback para bases antigas onde "pontualidade" era usado
                trat = getattr(a, "estrelas_pontualidade", 0)
            amb  = getattr(a, "estrelas_ambiente", None)
            if amb is None:
                amb = getattr(a, "estrelas_educacao", 0)
            sup  = getattr(a, "estrelas_suporte", None)
            if sup is None:
                sup = getattr(a, "estrelas_eficiencia", 0)

            item.update({
                "tratamento": trat or 0,
                "ambiente":   amb  or 0,
                "suporte":    sup  or 0,
                # retrocompat se o template usar nomes antigos:
                "trat": trat or 0, "amb": amb or 0, "sup": sup or 0,
            })
        else:
            # Restaurante -> Cooperado
            item.update({
                "pont":  getattr(a, "estrelas_pontualidade", 0) or 0,
                "educ":  getattr(a, "estrelas_educacao", 0) or 0,
                "efic":  getattr(a, "estrelas_eficiencia", 0) or 0,
                "apres": getattr(a, "estrelas_apresentacao", 0) or 0,
            })

        avaliacoes.append(SimpleNamespace(**item))

    # KPIs (funÃ§Ã£o local simples para evitar imports extras)
    def avg_or_zero(coluna):
        if coluna is None:
            return 0.0
        q = db.session.query(func.coalesce(func.avg(coluna), 0.0))
        if filtros:
            q = q.filter(and_(*filtros))
        return float(q.scalar() or 0.0)

    kpis = {"qtd": total, "geral": avg_or_zero(f_geral)}
    if tipo == "restaurante":
        kpis.update({
            "trat": avg_or_zero(f_trat),
            "amb":  avg_or_zero(f_amb),
            "sup":  avg_or_zero(f_sup),
        })
    else:
        kpis.update({
            "pont":  avg_or_zero(f_pont),
            "educ":  avg_or_zero(f_educ),
            "efic":  avg_or_zero(f_efic),
            "apres": avg_or_zero(f_apres),
        })

    # Ranking + chart (sempre definidos)
    ranking, chart_top = [], {"labels": [], "values": []}
    if tipo == "restaurante":
        q_rank = (
            db.session.query(
                Restaurante.id.label("id"),
                Restaurante.nome.label("nome"),
                func.count(Model.id).label("qtd"),
                func.coalesce(func.avg(f_geral), 0.0).label("m_geral"),
                (func.coalesce(func.avg(f_trat), 0.0) if f_trat is not None else literal(0.0)).label("m_trat"),
                (func.coalesce(func.avg(f_amb),  0.0) if f_amb  is not None else literal(0.0)).label("m_amb"),
                (func.coalesce(func.avg(f_sup),  0.0) if f_sup  is not None else literal(0.0)).label("m_sup"),
            )
            .join(Model, Model.restaurante_id == Restaurante.id)
        )
        if filtros:
            q_rank = q_rank.filter(and_(*filtros))
        ranking_rows = q_rank.group_by(Restaurante.id, Restaurante.nome).all()
        ranking = [{
            "rest_nome": r.nome, "qtd": int(r.qtd or 0),
            "m_geral": float(r.m_geral or 0),
            "m_trat":  float(r.m_trat or 0),
            "m_amb":   float(r.m_amb or 0),
            "m_sup":   float(r.m_sup or 0),
        } for r in ranking_rows]
        top = sorted([x for x in ranking if x["qtd"] >= 3], key=lambda x: x["m_geral"], reverse=True)[:10]
        chart_top = {"labels": [r["rest_nome"] for r in top], "values": [round(r["m_geral"], 2) for r in top]}
    else:
        q_rank = (
            db.session.query(
                Cooperado.id.label("id"),
                Cooperado.nome.label("nome"),
                func.count(Model.id).label("qtd"),
                func.coalesce(func.avg(f_geral), 0.0).label("m_geral"),
                (func.coalesce(func.avg(f_pont), 0.0) if f_pont is not None else literal(0.0)).label("m_pont"),
                (func.coalesce(func.avg(f_educ), 0.0) if f_educ is not None else literal(0.0)).label("m_educ"),
                (func.coalesce(func.avg(f_efic), 0.0) if f_efic is not None else literal(0.0)).label("m_efic"),
                (func.coalesce(func.avg(f_apres),0.0) if f_apres is not None else literal(0.0)).label("m_apres"),
            )
            .join(Model, Model.cooperado_id == Cooperado.id)
        )
        if filtros:
            q_rank = q_rank.filter(and_(*filtros))
        ranking_rows = q_rank.group_by(Cooperado.id, Cooperado.nome).all()
        ranking = [{
            "coop_nome": r.nome, "qtd": int(r.qtd or 0),
            "m_geral": float(r.m_geral or 0),
            "m_pont":  float(r.m_pont or 0),
            "m_educ":  float(r.m_educ or 0),
            "m_efic":  float(r.m_efic or 0),
            "m_apres": float(r.m_apres or 0),
        } for r in ranking_rows]
        top = sorted([x for x in ranking if x["qtd"] >= 3], key=lambda x: x["m_geral"], reverse=True)[:10]
        chart_top = {"labels": [r["coop_nome"] for r in top], "values": [round(r["m_geral"], 2) for r in top]}

    # Compatibilidade (mÃ©dia de geral por par coopÃ—rest) â€” nÃ£o quebra se vazio
    compat_map = {}
    for a in avaliacoes:
        key = (a.coop_id, a.rest_id)
        d = compat_map.get(key)
        if not d:
            d = {"coop": a.coop_nome, "rest": a.rest_nome, "sum": 0.0, "count": 0}
        d["sum"] += (a.geral or 0)
        d["count"] += 1
        compat_map[key] = d
    compat = []
    for d in compat_map.values():
        avg = (d["sum"] / d["count"]) if d["count"] else 0.0
        compat.append({"coop": d["coop"], "rest": d["rest"], "avg": avg, "count": d["count"]})
    compat.sort(key=lambda x: (-(x["avg"] or 0), -(x["count"] or 0), x["coop"], x["rest"]))

    # Filtros p/ repopular form + preserva args para paginaÃ§Ã£o
    _flt = SimpleNamespace(
        restaurante_id=restaurante_id,
        cooperado_id=cooperado_id,
        data_inicio=data_inicio or "",
        data_fim=data_fim or "",
    )
    preserve = request.args.to_dict(flat=True)
    preserve.pop("page", None)

    return render_template(
        "admin_dashboard.html",
        tab="avaliacoes",
        tipo=tipo,
        avaliacoes=avaliacoes,
        kpis=kpis,
        ranking=ranking,
        chart_top=chart_top,
        compat=compat,
        _flt=_flt,
        restaurantes=Restaurante.query.order_by(Restaurante.nome).all(),
        cooperados=Cooperado.query.order_by(Cooperado.nome).all(),
        pager=pager,
        page=pager.page,
        per_page=pager.per_page,
        preserve=preserve,
    )


@app.route("/admin/avaliacoes/export.csv", methods=["GET"])
@admin_required
def admin_export_avaliacoes_csv():
    tipo_raw = (request.args.get("tipo") or "restaurante").strip().lower()
    tipo = "restaurante" if tipo_raw == "restaurante" else "cooperado"

    restaurante_id = request.args.get("restaurante_id", type=int)
    cooperado_id   = request.args.get("cooperado_id", type=int)
    data_inicio    = (request.args.get("data_inicio") or "").strip()
    data_fim       = (request.args.get("data_fim") or "").strip()

    di = _parse_date(data_inicio)
    df = _parse_date(data_fim)

    Model = AvaliacaoRestaurante if (tipo == "restaurante") else AvaliacaoCooperado

    base = (
        db.session.query(
            Model,
            Restaurante.nome.label("rest_nome"),
            Cooperado.nome.label("coop_nome"),
        )
        .join(Restaurante, Model.restaurante_id == Restaurante.id)
        .join(Cooperado,   Model.cooperado_id   == Cooperado.id)
    )

    filtros = []
    if restaurante_id: filtros.append(Model.restaurante_id == restaurante_id)
    if cooperado_id:   filtros.append(Model.cooperado_id   == cooperado_id)
    if di: filtros.append(func.date(Model.criado_em) >= di.isoformat())
    if df: filtros.append(func.date(Model.criado_em) <= df.isoformat())
    if filtros: base = base.filter(and_(*filtros))

    rows = base.order_by(Model.criado_em.desc()).all()

    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";")

    if tipo == "restaurante":
        # Cooperado -> Restaurante: garante Tratamento no CSV
        w.writerow(["Data/Hora","Restaurante","Cooperado","Geral","Tratamento","Ambiente","Suporte",
                    "ComentÃ¡rio","MÃ©dia Ponderada","Sentimento","Temas","CrÃ­tico?"])
        for a, rest_nome, coop_nome in rows:
            trat = getattr(a, "estrelas_tratamento", None)
            if trat is None:
                trat = getattr(a, "estrelas_pontualidade", 0)
            amb  = getattr(a, "estrelas_ambiente", None)
            if amb is None:
                amb = getattr(a, "estrelas_educacao", 0)
            sup  = getattr(a, "estrelas_suporte", None)
            if sup is None:
                sup = getattr(a, "estrelas_eficiencia", 0)

            w.writerow([
                a.criado_em.strftime("%d/%m/%Y %H:%M") if a.criado_em else "",
                rest_nome, coop_nome,
                a.estrelas_geral or 0,
                (trat or 0), (amb or 0), (sup or 0),
                (getattr(a, "comentario", "") or "").strip(),
                (getattr(a, "media_ponderada", "") or ""),
                (getattr(a, "sentimento", "") or ""),
                (getattr(a, "temas", "") or ""),
                "SIM" if getattr(a, "alerta_crise", False) else "NÃƒO",
            ])
    else:
        # Restaurante -> Cooperado
        w.writerow(["Data/Hora","Restaurante","Cooperado","Geral","Pontualidade","EducaÃ§Ã£o","EficiÃªncia","ApresentaÃ§Ã£o",
                    "ComentÃ¡rio","MÃ©dia Ponderada","Sentimento","Temas","CrÃ­tico?"])
        for a, rest_nome, coop_nome in rows:
            w.writerow([
                a.criado_em.strftime("%d/%m/%Y %H:%M") if a.criado_em else "",
                rest_nome, coop_nome,
                a.estrelas_geral or 0,
                (getattr(a, "estrelas_pontualidade", 0) or 0),
                (getattr(a, "estrelas_educacao", 0) or 0),
                (getattr(a, "estrelas_eficiencia", 0) or 0),
                (getattr(a, "estrelas_apresentacao", 0) or 0),
                (getattr(a, "comentario", "") or "").strip(),
                (getattr(a, "media_ponderada", "") or ""),
                (getattr(a, "sentimento", "") or ""),
                (getattr(a, "temas", "") or ""),
                "SIM" if getattr(a, "alerta_crise", False) else "NÃƒO",
            ])

    mem = io.BytesIO(buf.getvalue().encode("utf-8-sig"))
    return send_file(mem, as_attachment=True,
                     download_name=f"avaliacoes_{tipo}.csv",
                     mimetype="text/csv")

# =========================
# CRUD Receitas/Despesas Coop (Admin)
# =========================
@app.route("/receitas/add", methods=["POST"])
@admin_required
def add_receita():
    f = request.form
    r = ReceitaCooperativa(
        descricao=f.get("descricao", "").strip(),
        valor_total=f.get("valor", type=float),
        data=_parse_date(f.get("data"))
    )
    db.session.add(r)
    db.session.commit()
    flash("Receita adicionada.", "success")
    return redirect(url_for("admin_dashboard", tab="receitas"))

@app.route("/receitas/<int:id>/edit", methods=["POST"])
@admin_required
def edit_receita(id):
    r = ReceitaCooperativa.query.get_or_404(id)
    f = request.form
    r.descricao = f.get("descricao", "").strip()
    # CORREÃ‡ÃƒO: campo correto Ã© valor_total
    r.valor_total = f.get("valor", type=float)
    r.data = _parse_date(f.get("data"))
    db.session.commit()
    flash("Receita atualizada.", "success")
    return redirect(url_for("admin_dashboard", tab="receitas"))

@app.route("/receitas/<int:id>/delete")
@admin_required
def delete_receita(id):
    r = ReceitaCooperativa.query.get_or_404(id)
    db.session.delete(r)
    db.session.commit()
    flash("Receita excluÃ­da.", "success")
    return redirect(url_for("admin_dashboard", tab="receitas"))

@app.route("/despesas/add", methods=["POST"])
@admin_required
def add_despesa():
    f = request.form
    d = DespesaCooperativa(
        descricao=f.get("descricao", "").strip(),
        valor=f.get("valor", type=float),
        data=_parse_date(f.get("data"))
    )
    db.session.add(d)
    db.session.commit()
    flash("Despesa adicionada.", "success")
    return redirect(url_for("admin_dashboard", tab="despesas"))

@app.route("/despesas/<int:id>/edit", methods=["POST"])
@admin_required
def edit_despesa(id):
    d = DespesaCooperativa.query.get_or_404(id)
    f = request.form
    d.descricao = f.get("descricao", "").strip()
    d.valor = f.get("valor", type=float)
    d.data = _parse_date(f.get("data"))
    db.session.commit()
    flash("Despesa atualizada.", "success")
    return redirect(url_for("admin_dashboard", tab="despesas"))

@app.route("/despesas/<int:id>/delete")
@admin_required
def delete_despesa(id):
    d = DespesaCooperativa.query.get_or_404(id)
    db.session.delete(d)
    db.session.commit()
    flash("Despesa excluÃ­da.", "success")
    return redirect(url_for("admin_dashboard", tab="despesas"))

# =========================
# Avisos (admin + pÃºblicos)
# =========================
import re
from flask import request, session, render_template, redirect, url_for, flash

@app.get("/avisos", endpoint="avisos_publicos")
def avisos_publicos():
    t = session.get("user_tipo")
    if t == "cooperado":
        return redirect(url_for("portal_cooperado_avisos"))
    if t == "restaurante":
        return redirect(url_for("portal_restaurante"))
    return redirect(url_for("login"))


@app.route("/admin/avisos", methods=["GET", "POST"])
@admin_required
def admin_avisos():
    cooperados = Cooperado.query.order_by(Cooperado.nome.asc()).all()
    restaurantes = Restaurante.query.order_by(Restaurante.nome.asc()).all()

    if request.method == "POST":
        f = request.form

        # ===== Alcance/destino vindos do form =====
        destino_tipo = (f.get("destino_tipo") or "").strip()  # 'cooperados' | 'restaurantes' | 'ambos'
        coop_alc = f.get("coop_alcance") or f.get("coop_alcance_ambos")   # 'todos' | 'selecionados'
        rest_alc = f.get("rest_alcance") or f.get("rest_alcance_ambos")   # 'todos' | 'selecionados'
        sel_coops = request.form.getlist("dest_cooperados[]") or request.form.getlist("dest_cooperados_ambos[]")
        sel_rests = request.form.getlist("dest_restaurantes[]") or request.form.getlist("dest_restaurantes_ambos[]")

        # ===== ConteÃºdo =====
        def _pick_msg(form):
            for key in (
                "corpo_html", "html", "mensagem_html", "conteudo_html", "descricao_html", "texto_html",
                "mensagem", "corpo", "conteudo", "descricao", "texto", "resumo", "body", "content"
            ):
                v = form.get(key)
                if v and v.strip():
                    return v.strip()
            return ""

        titulo = (f.get("titulo") or "").strip()
        msg    = _pick_msg(f)
        prioridade = (f.get("prioridade") or "normal")
        ativo = bool(f.get("ativo"))
        exigir_confirmacao = bool(f.get("exigir_confirmacao")) if hasattr(Aviso, "exigir_confirmacao") else False

        inicio_em = (lambda d=_parse_date(f.get("inicio_em")): datetime.combine(d, time()) if d else None)()
        fim_em    = (lambda d=_parse_date(f.get("fim_em")):    datetime.combine(d, time()) if d else None)()

        def _mk_aviso(tipo: str):
            a = Aviso(
                titulo=titulo,
                corpo=msg,
                tipo=tipo,  # 'global' | 'cooperado' | 'restaurante'
                prioridade=prioridade,
                fixado=False,
                ativo=ativo if hasattr(Aviso, "ativo") else True,
                criado_por_id=session.get("user_id"),
                inicio_em=inicio_em,
                fim_em=fim_em,
            )
            if hasattr(a, "exigir_confirmacao"):
                a.exigir_confirmacao = exigir_confirmacao
            return a

        # ===== Regras =====
        avisos_para_criar = []

        if destino_tipo == "cooperados":
            # sÃ³ cooperados
            if coop_alc == "selecionados":
                if not sel_coops:
                    flash("Selecione ao menos um cooperado.", "warning")
                    return redirect(url_for("admin_avisos"))
                try:
                    coop_id = int(sel_coops[0])  # modelo atual aceita 1 cooperado
                except Exception:
                    flash("SeleÃ§Ã£o de cooperado invÃ¡lida.", "warning")
                    return redirect(url_for("admin_avisos"))
                a = _mk_aviso("cooperado")
                a.destino_cooperado_id = coop_id
                avisos_para_criar.append(a)
            else:
                # todos cooperados = broadcast (destino_cooperado_id NULL)
                a = _mk_aviso("cooperado")
                a.destino_cooperado_id = None
                avisos_para_criar.append(a)

        elif destino_tipo == "restaurantes":
            # sÃ³ restaurantes
            if rest_alc == "selecionados":
                if not sel_rests:
                    flash("Selecione ao menos um restaurante.", "warning")
                    return redirect(url_for("admin_avisos"))
                try:
                    ids = [int(x) for x in sel_rests]
                except Exception:
                    flash("SeleÃ§Ã£o de restaurante invÃ¡lida.", "warning")
                    return redirect(url_for("admin_avisos"))
                a = _mk_aviso("restaurante")
                a.restaurantes = Restaurante.query.filter(Restaurante.id.in_(ids)).all()
                avisos_para_criar.append(a)
            else:
                # todos restaurantes = broadcast (lista vazia)
                a = _mk_aviso("restaurante")
                a.restaurantes = []
                avisos_para_criar.append(a)

        elif destino_tipo == "ambos":
            # cria DOIS avisos: um para cooperados e outro para restaurantes

            # Cooperados
            if coop_alc == "selecionados":
                if not sel_coops:
                    flash("Selecione ao menos um cooperado para o aviso dos cooperados.", "warning")
                    return redirect(url_for("admin_avisos"))
                try:
                    coop_id = int(sel_coops[0])
                except Exception:
                    flash("SeleÃ§Ã£o de cooperado invÃ¡lida.", "warning")
                    return redirect(url_for("admin_avisos"))
                a_coop = _mk_aviso("cooperado")
                a_coop.destino_cooperado_id = coop_id
            else:
                a_coop = _mk_aviso("cooperado")
                a_coop.destino_cooperado_id = None  # broadcast cooperados
            avisos_para_criar.append(a_coop)

            # Restaurantes
            if rest_alc == "selecionados":
                if not sel_rests:
                    flash("Selecione ao menos um restaurante para o aviso dos restaurantes.", "warning")
                    return redirect(url_for("admin_avisos"))
                try:
                    ids = [int(x) for x in sel_rests]
                except Exception:
                    flash("SeleÃ§Ã£o de restaurante invÃ¡lida.", "warning")
                    return redirect(url_for("admin_avisos"))
                a_rest = _mk_aviso("restaurante")
                a_rest.restaurantes = Restaurante.query.filter(Restaurante.id.in_(ids)).all()
            else:
                a_rest = _mk_aviso("restaurante")
                a_rest.restaurantes = []  # broadcast restaurantes
            avisos_para_criar.append(a_rest)

        else:
            # fallback antigo: se nada veio, publica global
            avisos_para_criar.append(_mk_aviso("global"))

        for a in avisos_para_criar:
            db.session.add(a)
        db.session.commit()

        flash("Aviso(s) publicado(s).", "success")
        return redirect(url_for("admin_avisos"))

    # GET
    avisos = Aviso.query.order_by(Aviso.fixado.desc(), Aviso.criado_em.desc()).all()
    return render_template(
        "admin_avisos.html",
        avisos=avisos,
        cooperados=cooperados,
        restaurantes=restaurantes,
    )


# --------- ROTAS QUE FALTAVAM (usadas no template) ---------

# Toggle VISIBILIDADE/FIXAÃ‡ÃƒO (aceita GET e POST)
@app.route("/admin/avisos/<int:aviso_id>/toggle", methods=["POST", "GET"], endpoint="admin_avisos_toggle")
@admin_required
def admin_avisos_toggle(aviso_id):
    a = Aviso.query.get_or_404(aviso_id)
    if hasattr(a, "ativo"):
        a.ativo = not bool(a.ativo)
    else:
        a.fixado = not bool(a.fixado)
    db.session.commit()
    flash("Aviso atualizado.", "success")
    return redirect(request.referrer or url_for("admin_avisos"))


# Excluir aviso (limpando relaÃ§Ãµes)
@app.route("/admin/avisos/<int:aviso_id>/excluir", methods=["POST"], endpoint="admin_avisos_excluir")
@admin_required
def admin_avisos_excluir(aviso_id):
    a = Aviso.query.get_or_404(aviso_id)

    # apaga confirmaÃ§Ãµes/leitorias
    try:
        AvisoLeitura.query.filter_by(aviso_id=aviso_id).delete(synchronize_session=False)
    except Exception:
        pass

    # limpa M2M com restaurantes, se existir
    try:
        if hasattr(a, "restaurantes"):
            a.restaurantes.clear()
    except Exception:
        pass

    db.session.delete(a)
    db.session.commit()
    flash("Aviso excluÃ­do.", "success")
    return redirect(url_for("admin_avisos"))


# Marcar aviso como lido (funciona com GET e POST)
@app.route("/avisos/<int:aviso_id>/lido", methods=["POST", "GET"])
def marcar_aviso_lido_universal(aviso_id: int):
    # se nÃ£o logado, bloqueia
    if "user_id" not in session:
        return redirect(url_for("login")) if request.method == "GET" else ("", 401)

    user_id = session.get("user_id")
    user_tipo = session.get("user_tipo")
    Aviso.query.get_or_404(aviso_id)

    def _ok_response():
        if request.method == "POST":
            return ("", 204)  # Ãºtil para fetch/AJAX
        return redirect(request.referrer or url_for("portal_cooperado_avisos"))

    if user_tipo == "cooperado":
        coop = Cooperado.query.filter_by(usuario_id=user_id).first()
        if not coop:
            return ("", 403) if request.method == "POST" else redirect(url_for("login"))
        if not AvisoLeitura.query.filter_by(aviso_id=aviso_id, cooperado_id=coop.id).first():
            db.session.add(AvisoLeitura(aviso_id=aviso_id, cooperado_id=coop.id, lido_em=datetime.now(timezone.utc)))
            db.session.commit()
        return _ok_response()

    if user_tipo == "restaurante":
        rest = Restaurante.query.filter_by(usuario_id=user_id).first()
        if not rest:
            return ("", 403) if request.method == "POST" else redirect(url_for("login"))
        if not AvisoLeitura.query.filter_by(aviso_id=aviso_id, restaurante_id=rest.id).first():
            db.session.add(AvisoLeitura(aviso_id=aviso_id, restaurante_id=rest.id, lido_em=datetime.now(timezone.utc)))
            db.session.commit()
        return _ok_response()

    return ("", 403) if request.method == "POST" else redirect(url_for("login"))

@app.route("/cooperados/add", methods=["POST"])
@admin_required
def add_cooperado():
    f = request.form
    nome = (f.get("nome") or "").strip()
    usuario_login = (f.get("usuario") or "").strip()
    senha = f.get("senha") or ""
    telefone = (f.get("telefone") or "").strip()   # <- NOVO
    foto = request.files.get("foto")

    # --- PEGAR TIPO ---
    tipo_usuario = f.get("tipo") or "cooperado"   # <- NOVO

    # evita usuÃ¡rio duplicado
    if Usuario.query.filter_by(usuario=usuario_login).first():
        flash("UsuÃ¡rio jÃ¡ existente.", "warning")
        return redirect(url_for("admin_dashboard", tab="cooperados"))

    # cria usuÃ¡rio (pode ser cooperado, restaurante, farmÃ¡cia, etc)
    u = Usuario(usuario=usuario_login, tipo=tipo_usuario, senha_hash="")
    u.set_password(senha)
    db.session.add(u)
    db.session.flush()  # garante u.id

    # cria o cooperado com telefone (opcional, depende do tipo)
    c = Cooperado(
        nome=nome,
        usuario_id=u.id,
        telefone=telefone,
        ultima_atualizacao=datetime.now(),
    )
    db.session.add(c)
    db.session.flush()  # garante c.id

    # salva foto (no banco)
    if foto and foto.filename:
        _save_foto_to_db(c, foto, is_cooperado=True)

    db.session.commit()
    flash("UsuÃ¡rio cadastrado.", "success")
    return redirect(url_for("admin_dashboard", tab="cooperados"))


@app.route("/cooperados/<int:id>/edit", methods=["POST"])
@admin_required
def edit_cooperado(id):
    c = Cooperado.query.get_or_404(id)
    f = request.form

    c.nome = (f.get("nome") or "").strip()
    c.usuario_ref.usuario = (f.get("usuario") or "").strip()
    c.telefone = (f.get("telefone") or "").strip()  # <- NOVO

    foto = request.files.get("foto")
    if foto and foto.filename:
        _save_foto_to_db(c, foto, is_cooperado=True)

    c.ultima_atualizacao = datetime.now()
    db.session.commit()
    flash("Cooperado atualizado.", "success")
    return redirect(url_for("admin_dashboard", tab="cooperados"))

@app.route("/cooperados/<int:id>/delete", methods=["POST"])
@admin_required
def delete_cooperado(id):
    c = Cooperado.query.get_or_404(id)
    u = c.usuario_ref

    try:
        # --- 1) Escalas do cooperado (e trocas ligadas a essas escalas) ---
        escala_ids = [eid for (eid,) in db.session.query(Escala.id)
                      .filter(Escala.cooperado_id == id).all()]

        if escala_ids:
            # Trocas que apontam para essas escalas
            db.session.execute(
                sa_delete(TrocaSolicitacao)
                .where(TrocaSolicitacao.origem_escala_id.in_(escala_ids))
            )
            # As prÃ³prias escalas
            db.session.execute(
                sa_delete(Escala)
                .where(Escala.id.in_(escala_ids))
            )

        # --- 2) Trocas onde o cooperado Ã© solicitante ou destino ---
        db.session.execute(
            sa_delete(TrocaSolicitacao)
            .where(or_(
                TrocaSolicitacao.solicitante_id == id,
                TrocaSolicitacao.destino_id == id
            ))
        )

        # --- 3) AvaliaÃ§Ãµes que guardam o cooperado_id ---
        db.session.execute(
            sa_delete(AvaliacaoCooperado).where(AvaliacaoCooperado.cooperado_id == id)
        )
        db.session.execute(
            sa_delete(AvaliacaoRestaurante).where(AvaliacaoRestaurante.cooperado_id == id)
        )

        # --- 4) LanÃ§amentos desse cooperado (se o CASCADE por lancamento_id nÃ£o estiver ativo) ---
        db.session.execute(
            sa_delete(Lancamento).where(Lancamento.cooperado_id == id)
        )

        # --- 5) MovimentaÃ§Ãµes financeiras do cooperado ---
        db.session.execute(
            sa_delete(ReceitaCooperado).where(ReceitaCooperado.cooperado_id == id)
        )
        db.session.execute(
            sa_delete(DespesaCooperado).where(DespesaCooperado.cooperado_id == id)
        )

        # --- 6) Leituras de avisos atreladas ao cooperado ---
        db.session.execute(
            sa_delete(AvisoLeitura).where(AvisoLeitura.cooperado_id == id)
        )

        # --- 7) Finalmente, o Cooperado e o Usuario vinculado ---
        db.session.delete(c)
        if u:
            db.session.delete(u)

        db.session.commit()
        flash("Cooperado excluÃ­do.", "success")

    except IntegrityError as e:
        db.session.rollback()
        current_app.logger.exception(e)
        flash("NÃ£o foi possÃ­vel excluir: existem vÃ­nculos ativos.", "danger")

    return redirect(url_for("admin_dashboard", tab="cooperados"))

@app.route("/cooperados/<int:id>/reset_senha", methods=["POST"])
@admin_required
def reset_senha_cooperado(id):
    c = Cooperado.query.get_or_404(id)
    ns = request.form.get("nova_senha") or ""
    cs = request.form.get("confirmar_senha") or ""
    if ns != cs:
        flash("As senhas nÃ£o conferem.", "warning")
        return redirect(url_for("admin_dashboard", tab="cooperados"))
    c.usuario_ref.set_password(ns)
    db.session.commit()
    flash("Senha do cooperado atualizada.", "success")
    return redirect(url_for("admin_dashboard", tab="cooperados"))

@app.route("/restaurantes/add", methods=["POST"])
@admin_required
def add_restaurante():
    f = request.form
    nome = (f.get("nome") or "").strip()
    periodo = (f.get("periodo") or "seg-dom").strip()
    usuario_login = (f.get("usuario") or "").strip()
    senha = f.get("senha") or ""
    foto = request.files.get("foto")
    # checkbox
    is_farmacia = "farmacia" in f  # True se marcado

    if not (nome and periodo and usuario_login and senha):
        flash("Preencha todos os campos obrigatÃ³rios.", "warning")
        return redirect(url_for("admin_dashboard", tab="restaurantes"))

    if Usuario.query.filter_by(usuario=usuario_login).first():
        flash("UsuÃ¡rio jÃ¡ existente.", "warning")
        return redirect(url_for("admin_dashboard", tab="restaurantes"))

    # tipo conforme checkbox
    u = Usuario(usuario=usuario_login, tipo=("farmacia" if is_farmacia else "restaurante"), senha_hash="")
    u.set_password(senha)
    db.session.add(u)
    db.session.flush()

    r = Restaurante(
        nome=nome,
        periodo=periodo,
        usuario_id=u.id,
        farmacia=is_farmacia,  # grava boolean
    )
    db.session.add(r)
    db.session.flush()  # garante r.id

    if foto and foto.filename:
        _save_foto_to_db(r, foto, is_cooperado=False)

    db.session.commit()
    flash("Estabelecimento cadastrado.", "success")
    return redirect(url_for("admin_dashboard", tab="restaurantes"))


# ------------------------------------------------------
# UPDATE: edita dados do restaurante (modal de ediÃ§Ã£o)
# Espera campos:
#   nome_edit, periodo_edit, farmacia_edit (checkbox),
#   logo_edit (file opcional)
# Endpoint usado no JS: update_restaurante
# ------------------------------------------------------
@app.route("/restaurantes/<int:id>/update", methods=["POST"])
@admin_required
def update_restaurante(id):
    r = Restaurante.query.get_or_404(id)
    f = request.form

    nome_edit = (f.get("nome_edit") or "").strip()
    periodo_edit = (f.get("periodo_edit") or "seg-dom").strip()
    is_farmacia = "farmacia_edit" in f  # checkbox do modal

    if not (nome_edit and periodo_edit):
        flash("Nome e PerÃ­odo sÃ£o obrigatÃ³rios.", "warning")
        return redirect(url_for("admin_dashboard", tab="restaurantes"))

    r.nome = nome_edit
    r.periodo = periodo_edit
    r.farmacia = is_farmacia

    # mantÃ©m o tipo do usuÃ¡rio alinhado ao flag farmÃ¡cia
    if r.usuario_ref:
        r.usuario_ref.tipo = "farmacia" if is_farmacia else "restaurante"

    foto = request.files.get("logo_edit")
    if foto and foto.filename:
        _save_foto_to_db(r, foto, is_cooperado=False)

    db.session.commit()
    flash("Estabelecimento atualizado.", "success")
    return redirect(url_for("admin_dashboard", tab="restaurantes"))


# ------------------------------------------------------
# DELETE: apaga vÃ­nculos (trocas/escalas/lancamentos),
# depois o restaurante e o usuÃ¡rio vinculado
# Endpoint usado no HTML: delete_restaurante
# ------------------------------------------------------
@app.route("/restaurantes/<int:id>/delete", methods=["POST"])
@admin_required
def delete_restaurante(id):
    r = Restaurante.query.get_or_404(id)
    u = r.usuario_ref
    try:
        # 1) IDs das escalas do restaurante
        escala_ids = [e.id for e in Escala.query.with_entities(Escala.id)
                      .filter(Escala.restaurante_id == id).all()]

        if escala_ids:
            # 2) Apagar trocas que referenciam essas escalas
            db.session.execute(
                sa_delete(TrocaSolicitacao).where(
                    TrocaSolicitacao.origem_escala_id.in_(escala_ids)
                )
            )
            # 3) Apagar as escalas
            db.session.execute(
                sa_delete(Escala).where(Escala.restaurante_id == id)
            )

        # 4) Apagar lanÃ§amentos do restaurante
        db.session.execute(
            sa_delete(Lancamento).where(Lancamento.restaurante_id == id)
        )

        # 5) Apagar restaurante e usuÃ¡rio
        db.session.delete(r)
        if u:
            db.session.delete(u)

        db.session.commit()
        flash("Estabelecimento excluÃ­do.", "success")
    except IntegrityError as e:
        db.session.rollback()
        current_app.logger.exception(e)
        flash("NÃ£o foi possÃ­vel excluir: existem vÃ­nculos ativos.", "danger")

    return redirect(url_for("admin_dashboard", tab="restaurantes"))


# ------------------------------------------------------
# RESET SENHA (ADMIN): modal "Senha" na tabela
# Espera campos: senha_nova, senha_conf
# Endpoint no HTML: reset_senha_restaurante
# ------------------------------------------------------
@app.route("/restaurantes/<int:id>/reset_senha", methods=["POST"])
@admin_required
def reset_senha_restaurante(id):
    r = Restaurante.query.get_or_404(id)
    user = r.usuario_ref
    if not user:
        flash("UsuÃ¡rio vinculado nÃ£o encontrado.", "danger")
        return redirect(url_for("admin_dashboard", tab="restaurantes"))

    ns = (request.form.get("senha_nova") or "").strip()
    cs = (request.form.get("senha_conf") or "").strip()

    if not ns or not cs:
        flash("Informe e confirme a nova senha.", "warning")
        return redirect(url_for("admin_dashboard", tab="restaurantes"))
    if ns != cs:
        flash("As senhas nÃ£o conferem.", "warning")
        return redirect(url_for("admin_dashboard", tab="restaurantes"))

    user.set_password(ns)
    db.session.commit()
    flash("Senha do estabelecimento redefinida.", "success")
    return redirect(url_for("admin_dashboard", tab="restaurantes"))


# ------------------------------------------------------
# PORTAL RESTAURANTE: alterar a prÃ³pria senha
# (permite trocar caso jÃ¡ tenha senha definida)
# ------------------------------------------------------
@app.route("/rest/alterar-senha", methods=["POST"], endpoint="rest_alterar_senha")
@role_required("restaurante", "farmacia")
def alterar_senha_rest():
    # identifica o estabelecimento (funciona tanto para restaurante quanto farmÃ¡cia)
    u_id = session.get("user_id")
    rest = Restaurante.query.filter_by(usuario_id=u_id).first_or_404()
    user = rest.usuario_ref  # Usuario vinculado ao estabelecimento

    # dados do formulÃ¡rio
    atual = (request.form.get("senha_atual") or "").strip()
    nova  = (request.form.get("senha_nova")  or "").strip()
    conf  = (request.form.get("senha_conf")  or "").strip()

    # validaÃ§Ãµes bÃ¡sicas
    if not (nova and conf):
        flash("Preencha todos os campos.", "warning")
        return redirect(url_for("portal_restaurante", view="config"))
    if nova != conf:
        flash("A confirmaÃ§Ã£o nÃ£o confere com a nova senha.", "warning")
        return redirect(url_for("portal_restaurante", view="config"))
    if len(nova) < 6:
        flash("A nova senha deve ter pelo menos 6 caracteres.", "warning")
        return redirect(url_for("portal_restaurante", view="config"))

    # se jÃ¡ existe hash salvo, exigir a senha atual correta
    if user.senha_hash:
        if not atual:
            flash("Informe a senha atual.", "warning")
            return redirect(url_for("portal_restaurante", view="config"))
        if not check_password_hash(user.senha_hash, atual):
            flash("Senha atual incorreta.", "danger")
            return redirect(url_for("portal_restaurante", view="config"))

    # aplica nova senha
    user.senha_hash = generate_password_hash(nova)
    db.session.commit()

    flash("Senha alterada com sucesso!", "success")
    return redirect(url_for("portal_restaurante", view="config"))

# ------------------------------------------------------
# CONFIGURAÃ‡Ã•ES GERAIS (mantive igual ao seu)
# ------------------------------------------------------
@app.route("/config/update", methods=["POST"])
@admin_required
def update_config():
    cfg = get_config()
    cfg.salario_minimo = request.form.get("salario_minimo", type=float) or 0.0
    db.session.commit()
    flash("ConfiguraÃ§Ã£o atualizada.", "success")
    return redirect(url_for("admin_dashboard", tab="config"))


@app.route("/admin/alterar_admin", methods=["POST"])
@admin_required
def alterar_admin():
    admin = Usuario.query.filter_by(tipo="admin").first()
    admin.usuario = request.form.get("usuario", admin.usuario).strip()
    nova = request.form.get("nova_senha", "")
    confirmar = request.form.get("confirmar_senha", "")
    if nova or confirmar:
        if nova != confirmar:
            flash("As senhas nÃ£o conferem.", "warning")
            return redirect(url_for("admin_dashboard", tab="config"))
        admin.set_password(nova)
    db.session.commit()
    flash("Conta do administrador atualizada.", "success")
    return redirect(url_for("admin_dashboard", tab="config"))
# =========================
# Receitas/Despesas Cooperado (Admin)
# =========================
@app.route("/coop/receitas/add", methods=["POST"])
@admin_required
def add_receita_coop():
    f = request.form
    rc = ReceitaCooperado(
        cooperado_id=f.get("cooperado_id", type=int),
        descricao=f.get("descricao", "").strip(),
        valor=f.get("valor", type=float),
        data=_parse_date(f.get("data"))
    )
    db.session.add(rc)
    db.session.commit()
    flash("Receita do cooperado adicionada.", "success")
    return redirect(url_for("admin_dashboard", tab="coop_receitas"))

@app.route("/coop/receitas/<int:id>/edit", methods=["POST"])
@admin_required
def edit_receita_coop(id):
    rc = ReceitaCooperado.query.get_or_404(id)
    f = request.form
    rc.cooperado_id = f.get("cooperado_id", type=int)
    rc.descricao = f.get("descricao", "").strip()
    rc.valor = f.get("valor", type=float)
    rc.data = _parse_date(f.get("data"))
    db.session.commit()
    flash("Receita do cooperado atualizada.", "success")
    return redirect(url_for("admin_dashboard", tab="coop_receitas"))

@app.route("/coop/receitas/<int:id>/delete")
@admin_required
def delete_receita_coop(id):
    rc = ReceitaCooperado.query.get_or_404(id)
    db.session.delete(rc)
    db.session.commit()
    flash("Receita do cooperado excluÃ­da.", "success")
    return redirect(url_for("admin_dashboard", tab="coop_receitas"))

@app.route("/coop/despesas/add", methods=["POST"])
@admin_required
def add_despesa_coop():
    f = request.form
    ids = request.form.getlist("cooperado_ids[]")
    descricao = f.get("descricao", "").strip()
    valor_total = f.get("valor", type=float) or 0.0
    d = _parse_date(f.get("data"))

    cooperados = Cooperado.query.order_by(Cooperado.nome).all()
    dest_ids = [c.id for c in cooperados] if "all" in ids else [int(i) for i in ids if i.isdigit()]
    if not dest_ids:
        flash("Selecione pelo menos um cooperado.", "warning")
        return redirect(url_for("admin_dashboard", tab="coop_despesas"))

    valor_unit = round(valor_total / max(1, len(dest_ids)), 2)
    for cid in dest_ids:
        db.session.add(DespesaCooperado(cooperado_id=cid, descricao=descricao, valor=valor_unit, data=d))
    db.session.commit()
    flash("Despesa(s) lanÃ§ada(s).", "success")
    return redirect(url_for("admin_dashboard", tab="coop_despesas"))

@app.route("/coop/despesas/<int:id>/edit", methods=["POST"])
@admin_required
def edit_despesa_coop(id):
    dc = DespesaCooperado.query.get_or_404(id)
    f = request.form
    dc.cooperado_id = f.get("cooperado_id", type=int)
    dc.descricao = f.get("descricao", "").strip()
    dc.valor = f.get("valor", type=float)
    dc.data = _parse_date(f.get("data"))
    db.session.commit()
    flash("Despesa do cooperado atualizada.", "success")
    return redirect(url_for("admin_dashboard", tab="coop_despesas"))

@app.route("/coop/despesas/<int:id>/delete")
@admin_required
def delete_despesa_coop(id):
    dc = DespesaCooperado.query.get_or_404(id)
    db.session.delete(dc)
    db.session.commit()
    flash("Despesa do cooperado excluÃ­da.", "success")
    return redirect(url_for("admin_dashboard", tab="coop_despesas"))

# =========================
# BenefÃ­cios â€” Editar / Excluir (Admin)
# =========================
def _split_field(form, key_list, key_str):
    """
    LÃª uma lista do form (key[]) ou uma string separada por , ; (key_str).
    Retorna lista de strings jÃ¡ stripadas e sem vazios.
    """
    vals = form.getlist(key_list) if key_list.endswith("[]") else []
    if not vals:
        raw = (form.get(key_str) or "").replace(",", ";")
        vals = [x.strip() for x in raw.split(";") if x.strip()]
    return vals

def _ensure_periodo_ok(di: date | None, df: date | None) -> tuple[date | None, date | None]:
    if di and df and df < di:
        di, df = df, di  # inverte para garantir di <= df
    return di, df

TIPO_MAP = {
    "hosp": "hospitalar", "hospitalar": "hospitalar",
    "farm": "farmaceutico","farmacÃªutico":"farmaceutico","farmaceutico":"farmaceutico",
    "alim": "alimentar",  "alimentar": "alimentar",
}

@app.post("/beneficios/<int:id>/edit")
@admin_required
def edit_beneficio(id):
    """
    Atualiza um registro de benefÃ­cio existente.
    Espera (via form):
      - data_inicial (YYYY-MM-DD ou DD/MM/YYYY)
      - data_final   (YYYY-MM-DD ou DD/MM/YYYY)
      - data_lancamento (opcional)
      - tipo  (hospitalar|farmaceutico|alimentar ou siglas: hosp|farm|alim)
      - valor_total (float)
      - recebedores_ids[]  ou recebedores_ids  (string: "1;2;3")
      - recebedores_nomes[] ou recebedores_nomes (string: "Ana;Bia;â€¦")
    """
    b = BeneficioRegistro.query.get_or_404(id)
    f = request.form

    # --- Datas ---
    di = _parse_date(f.get("data_inicial"))
    df = _parse_date(f.get("data_final"))
    di, df = _ensure_periodo_ok(di, df)

    if di: b.data_inicial = di
    if df: b.data_final   = df

    dl = _parse_date(f.get("data_lancamento"))
    if dl:
        b.data_lancamento = dl

    # --- Tipo ---
    tipo_in = (f.get("tipo") or "").strip().lower()
    if tipo_in in TIPO_MAP:
        b.tipo = TIPO_MAP[tipo_in]

    # --- Valor total ---
    val_raw = f.get("valor_total")
    if val_raw is not None and val_raw != "":
        try:
            b.valor_total = float(str(val_raw).replace(",", "."))
        except ValueError:
            flash("Valor total invÃ¡lido.", "warning")
            return redirect(url_for("admin_dashboard", tab="beneficios"))

    # --- Recebedores ---
    ids_list   = _split_field(f, "recebedores_ids[]",   "recebedores_ids")
    nomes_list = _split_field(f, "recebedores_nomes[]", "recebedores_nomes")

    # Se vierem sÃ³ IDs, tenta resolver nomes
    if ids_list and not nomes_list:
        ids_int = [int(x) for x in ids_list if str(x).isdigit()]
        if ids_int:
            coops = Cooperado.query.filter(Cooperado.id.in_(ids_int)).all()
            m = {str(c.id): c.nome for c in coops}
            nomes_list = [m.get(str(i), "") for i in ids_int]

    # Sanitiza; mantÃ©m alinhamento por Ã­ndice
    ids_sane   = [str(int(x)) for x in ids_list if str(x).isdigit()]
    nomes_sane = [n for n in nomes_list if n is not None]

    # Alinha tamanhos (corta o excedente do maior)
    n = min(len(ids_sane), len(nomes_sane)) if ids_sane and nomes_sane else max(len(ids_sane), len(nomes_sane))
    ids_sane   = ids_sane[:n]
    nomes_sane = (nomes_sane[:n] if nomes_sane else [""] * n)

    b.recebedores_ids   = ";".join(ids_sane)
    b.recebedores_nomes = ";".join(nomes_sane)

    db.session.commit()
    flash("BenefÃ­cio atualizado.", "success")
    return redirect(url_for("admin_dashboard", tab="beneficios"))


# 1) Excluir 1 (via modal, com hidden)
@app.post("/beneficios/delete-one", endpoint="excluir_beneficio_one")
@admin_required
def excluir_beneficio_one():
    bid = request.form.get("beneficio_id", type=int)
    if not bid:
        flash("ID invÃ¡lido.", "warning")
        return redirect(url_for("admin_dashboard", tab="beneficios"))
    b = BeneficioRegistro.query.get_or_404(bid)
    db.session.delete(b)
    db.session.commit()
    flash("Registro de benefÃ­cio excluÃ­do.", "info")
    return redirect(url_for("admin_dashboard", tab="beneficios"))

# 2) Excluir vÃ¡rios (bulk)
@app.post("/beneficios/delete-bulk", endpoint="excluir_beneficio_bulk")
@admin_required
def excluir_beneficio_bulk():
    ids = {int(x) for x in request.form.getlist("ids[]") if str(x).isdigit()}
    if not ids:
        flash("Selecione ao menos um benefÃ­cio.", "warning")
        return redirect(url_for("admin_dashboard", tab="beneficios"))
    qs = BeneficioRegistro.query.filter(BeneficioRegistro.id.in_(ids)).all()
    for b in qs:
        db.session.delete(b)
    db.session.commit()
    flash(f"{len(qs)} registro(s) excluÃ­do(s).", "info")
    return redirect(url_for("admin_dashboard", tab="beneficios"))

# =========================
# BenefÃ­cios â€” Criar/Ratear (Admin)
# =========================
@app.post("/beneficios/ratear", endpoint="ratear_beneficios")
@admin_required
def ratear_beneficios():
    """
    Cria um BeneficioRegistro a partir do form de 'Ratear benefÃ­cios'.
    Espera:
      - data_inicial, data_final, (opcional) data_lancamento
      - tipo (hospitalar|farmaceutico|alimentar ou hosp|farm|alim)
      - valor_total
      - recebedores_ids[]  ou recebedores_ids  ("1;2;3")
      - recebedores_nomes[] ou recebedores_nomes ("Ana;Bia;â€¦")
    """
    f = request.form

    di = _parse_date(f.get("data_inicial"))
    df = _parse_date(f.get("data_final"))
    if di and df and df < di:
        di, df = df, di

    dl = _parse_date(f.get("data_lancamento"))

    tipo_in = (f.get("tipo") or "").strip().lower()
    tipo = TIPO_MAP.get(tipo_in, tipo_in or "alimentar")  # default seguro

    # valor
    valor_total = None
    raw_val = f.get("valor_total")
    if raw_val not in (None, ""):
        try:
            valor_total = float(str(raw_val).replace(",", "."))
        except ValueError:
            flash("Valor total invÃ¡lido.", "warning")
            return redirect(url_for("admin_dashboard", tab="beneficios"))

    # recebedores
    ids_list   = _split_field(f, "recebedores_ids[]",   "recebedores_ids")
    nomes_list = _split_field(f, "recebedores_nomes[]", "recebedores_nomes")

    # se vier sÃ³ ID, tenta nomes
    if ids_list and not nomes_list:
        ids_int = [int(x) for x in ids_list if str(x).isdigit()]
        if ids_int:
            coops = Cooperado.query.filter(Cooperado.id.in_(ids_int)).all()
            m = {str(c.id): c.nome for c in coops}
            nomes_list = [m.get(str(i), "") for i in ids_int]

    ids_sane   = [str(int(x)) for x in ids_list if str(x).isdigit()]
    nomes_sane = [n for n in nomes_list if n is not None]

    n = min(len(ids_sane), len(nomes_sane)) if ids_sane and nomes_sane else max(len(ids_sane), len(nomes_sane))
    ids_sane   = ids_sane[:n]
    nomes_sane = (nomes_sane[:n] if nomes_sane else [""] * n)

    if not di or not df or not ids_sane:
        flash("Preencha perÃ­odo e pelo menos um recebedor.", "warning")
        return redirect(url_for("admin_dashboard", tab="beneficios"))

    b = BeneficioRegistro(
        data_inicial=di,
        data_final=df,
        data_lancamento=dl,
        tipo=tipo,
        valor_total=valor_total or 0.0,
        recebedores_ids=";".join(ids_sane),
        recebedores_nomes=";".join(nomes_sane),
    )
    db.session.add(b)
    db.session.commit()
    flash("BenefÃ­cio registrado/Rateado.", "success")
    return redirect(url_for("admin_dashboard", tab="beneficios"))

# =========================
# Escalas â€” Upload (substituiÃ§Ã£o TOTAL sempre)
# =========================
@app.route("/escalas/upload", methods=["POST"])
@admin_required
def upload_escala():
    import os, re as _re, unicodedata as _u, difflib as _dif

    file = request.files.get("file")
    if not file or not file.filename.lower().endswith(".xlsx"):
        flash("Envie um arquivo .xlsx vÃ¡lido.", "warning")
        return redirect(url_for("admin_dashboard", tab="escalas"))

    # salva o arquivo (o nome nÃ£o influencia a lÃ³gica)
    path = os.path.join(UPLOAD_DIR, secure_filename(file.filename))
    file.save(path)

    # abre com openpyxl
    try:
        import openpyxl
    except Exception:
        flash("Arquivo salvo, mas falta a biblioteca 'openpyxl' (pip install openpyxl).", "warning")
        return redirect(url_for("admin_dashboard", tab="escalas"))

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
    except Exception as e:
        flash(f"Erro ao abrir a planilha: {e}", "danger")
        return redirect(url_for("admin_dashboard", tab="escalas"))

    # ------- helpers -------
    def _norm_local(s: str) -> str:
        s = _u.normalize("NFD", str(s or "").strip().lower())
        s = "".join(ch for ch in s if _u.category(ch) != "Mn")
        return _re.sub(r"[^a-z0-9]+", " ", s).strip()

    def _norm_login_local(s: str) -> str:
        s = _u.normalize("NFD", s or "")
        s = "".join(ch for ch in s if _u.category(ch) != "Mn")
        return _re.sub(r"\s+", "", s.lower().strip())

    def to_css_color_local(v: str) -> str:
        t = str(v or "").strip()
        if not t: return ""
        if _re.fullmatch(r"[0-9a-fA-F]{8}", t):
            a = int(t[0:2], 16) / 255.0
            r = int(t[2:4], 16); g = int(t[4:6], 16); b = int(t[6:8], 16)
            return f"rgba({r},{g},{b},{a:.3f})"
        if _re.fullmatch(r"[0-9a-fA-F]{6}", t):
            return f"#{t}"
        if _re.fullmatch(r"#?[0-9a-fA-F]{6,8}", t):
            if not t.startswith("#"): t = f"#{t}"
            if len(t) == 9:
                a = int(t[1:3], 16) / 255.0
                r = int(t[3:5], 16); g = int(t[5:7], 16); b = int(t[7:9], 16)
                return f"rgba({r},{g},{b},{a:.3f})"
            return t
        m = _re.fullmatch(r"\s*(\d{1,3})\s*[,;]\s*(\d{1,3})\s*[,;]\s*(\d{1,3})\s*", t)
        if m:
            r, g, b = [max(0, min(255, int(x))) for x in m.groups()]
            return f"rgb({r},{g},{b})"
        mapa = {"azul":"blue","vermelho":"red","verde":"green","amarelo":"yellow",
                "cinza":"gray","preto":"black","branco":"white","laranja":"orange","roxo":"purple"}
        return mapa.get(t.lower(), t)

    def fmt_data_cell(v) -> str:
        if v is None or str(v).strip() == "": return ""
        if isinstance(v, datetime): return v.date().strftime("%d/%m/%Y")
        if isinstance(v, date):     return v.strftime("%d/%m/%Y")
        s = str(v).strip()
        m = _re.fullmatch(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", s)
        if m:
            y, mth, d = map(int, m.groups())
            try: return date(y, mth, d).strftime("%d/%m/%Y")
            except Exception: return s
        return s

    # ------- cabeÃ§alhos (detecÃ§Ã£o automÃ¡tica) -------
    def _score_header_row(cells):
        aliases = [
            "data","dia","data do plantao","turno","horario","horÃ¡rio","hora","periodo","perÃ­odo",
            "contrato","restaurante","unidade","local",
            "login","usuario","usuÃ¡rio","username","user","nome de usuario","nome de usuÃ¡rio",
            "nome","nome do cooperado","cooperado","motoboy","entregador",
            "cor","cores","cor da celula","cor celula",
        ]
        aliases_norm = {_norm_local(a) for a in aliases}
        score = 0
        seen = set()
        for c in cells:
            key = _norm_local(str(getattr(c, "value", "") or ""))
            if not key:
                continue
            score += 1
            for a in aliases_norm:
                if a and (a == key or a in key or key in a):
                    if (key, a) not in seen:
                        score += 2
                        seen.add((key, a))
        return score

    header_row_idx = 1
    best_score = -1
    last_row_to_check = min(ws.max_row, 10)
    for i in range(1, last_row_to_check + 1):
        row_cells = list(ws[i])
        s = _score_header_row(row_cells)
        if s > best_score:
            best_score = s
            header_row_idx = i

    headers_norm = { _norm_local(str(c.value or "")) : j for j, c in enumerate(ws[header_row_idx], start=1) }

    def find_col(*aliases):
        al = [_norm_local(a) for a in aliases]
        for a in al:
            if a in headers_norm: return headers_norm[a]
        for k_norm, j in headers_norm.items():
            for a in al:
                if a and a in k_norm: return j
        return None

    col_data     = find_col("data", "dia", "data do plantao")
    col_turno    = find_col("turno")
    col_horario  = find_col("horario", "horÃ¡rio", "hora", "periodo", "perÃ­odo")
    col_contrato = find_col("contrato", "restaurante", "unidade", "local")
    col_login    = find_col("login", "usuario", "usuÃ¡rio", "username", "user", "nome de usuario", "nome de usuÃ¡rio")
    col_nome     = find_col("nome", "nome do cooperado", "cooperado", "motoboy", "entregador")
    col_cor      = find_col("cor","cores","cor da celula","cor celula")

    app.logger.info(f"[ESCALAS] header_row={header_row_idx} headers_norm={headers_norm}")

    if not col_login and not col_nome:
        flash("NÃ£o encontrei a coluna de LOGIN nem a de NOME do cooperado na planilha.", "danger")
        app.logger.warning(f"[ESCALAS] Falha header: headers_norm={headers_norm} (linha {header_row_idx})")
        return redirect(url_for("admin_dashboard", tab="escalas"))

    # ------- cache entidades -------
    restaurantes = Restaurante.query.order_by(Restaurante.nome).all()
    cooperados   = Cooperado.query.order_by(Cooperado.nome).all()

    def match_restaurante_id(contrato_txt: str) -> int | None:
        a = _norm_local(contrato_txt)
        if not a: return None
        for r in restaurantes:
            b = _norm_local(r.nome)
            if a == b or a in b or b in a: return r.id
        try:
            nomes_norm = [_norm_local(r.nome) for r in restaurantes]
            close = _dif.get_close_matches(a, nomes_norm, n=1, cutoff=0.87)
            if close:
                alvo = close[0]
                for r in restaurantes:
                    if _norm_local(r.nome) == alvo: return r.id
        except Exception:
            pass
        return None

    def match_cooperado_by_login(login_txt: str) -> Cooperado | None:
        key = _norm_login_local(login_txt)
        if not key: return None
        for c in cooperados:
            login = getattr(c, "usuario_ref", None)
            login_val = getattr(login, "usuario", "") if login else ""
            if _norm_login_local(login_val) == key:
                return c
        return None

    # helper global por nome (jÃ¡ existe no seu arquivo)
    from_here_match_by_name = _match_cooperado_by_name

    # ------- parse linhas -------
    linhas_novas, total_linhas_planilha = [], 0
    start_row = header_row_idx + 1

    for i in range(start_row, ws.max_row + 1):
        login_txt = str(ws.cell(i, col_login).value).strip() if col_login else ""
        nome_txt  = str(ws.cell(i, col_nome ).value).strip() if col_nome  else ""

        # ignora linhas totalmente vazias
        if not login_txt and not nome_txt:
            vals = [
                (ws.cell(i, col_data).value     if col_data     else None),
                (ws.cell(i, col_turno).value    if col_turno    else None),
                (ws.cell(i, col_horario).value  if col_horario  else None),
                (ws.cell(i, col_contrato).value if col_contrato else None),
            ]
            if all((v is None or str(v).strip() == "") for v in vals):
                continue

        total_linhas_planilha += 1

        data_v     = ws.cell(i, col_data).value     if col_data     else None
        turno_v    = ws.cell(i, col_turno).value    if col_turno    else None
        horario_v  = ws.cell(i, col_horario).value  if col_horario  else None
        contrato_v = ws.cell(i, col_contrato).value if col_contrato else None
        cor_v      = ws.cell(i, col_cor).value      if col_cor      else None

        contrato_txt = (str(contrato_v).strip() if contrato_v is not None else "")
        rest_id      = match_restaurante_id(contrato_txt)

        coop_match = match_cooperado_by_login(login_txt) if login_txt else None
        if not coop_match and nome_txt:
            coop_match = from_here_match_by_name(nome_txt, cooperados)

        nome_fallback = (nome_txt or login_txt)

        linhas_novas.append({
            "cooperado_id":   (coop_match.id if coop_match else None),
            "cooperado_nome": (None if coop_match else nome_fallback),
            "data":           fmt_data_cell(data_v),
            "turno":          (str(turno_v).strip() if turno_v is not None else ""),
            "horario":        (str(horario_v).strip() if horario_v is not None else ""),
            "contrato":       contrato_txt,
            "cor":            to_css_color_local(cor_v),
            "restaurante_id": rest_id,
        })

    if not linhas_novas:
        app.logger.warning(f"[ESCALAS] Nenhuma linha importada. header_row={header_row_idx} headers_norm={headers_norm}")
        flash("Nada importado: nenhum registro vÃ¡lido encontrado. Verifique a linha dos cabeÃ§alhos e os nomes das colunas.", "warning")
        return redirect(url_for("admin_dashboard", tab="escalas"))

    # ------- SUBSTITUIÃ‡ÃƒO TOTAL -------
    try:
        from sqlalchemy import text as sa_text, delete as sa_delete

        # Remove todas as escalas antigas de forma segura (sem violar FKs de trocas)
        if _is_sqlite():
            # Em dev/local com SQLite nÃ£o hÃ¡ TRUNCATE CASCADE â€” apaga dependentes e depois escalas
            db.session.execute(sa_delete(TrocaSolicitacao))
            db.session.execute(sa_delete(Escala))
        else:
            # Em produÃ§Ã£o (Postgres): TRUNCATE com CASCADE limpa escalas e dependentes de uma vez
            db.session.execute(sa_text("TRUNCATE TABLE escalas RESTART IDENTITY CASCADE"))

        # insere novas linhas
        for row in linhas_novas:
            db.session.add(Escala(**row))

        # marca atualizaÃ§Ã£o para cooperados reconhecidos
        ids_reconhecidos = {int(r["cooperado_id"]) for r in linhas_novas if r.get("cooperado_id")}
        for cid in ids_reconhecidos:
            c = db.session.get(Cooperado, cid)
            if c:
                c.ultima_atualizacao = datetime.now()

        db.session.commit()
        flash(f"Escala substituÃ­da com sucesso. {len(linhas_novas)} linha(s) importada(s) (de {total_linhas_planilha}).", "success")

    except Exception as e:
        db.session.rollback()
        app.logger.exception("Erro ao importar a escala")
        flash(f"Erro ao importar a escala: {e}", "danger")

    return redirect(url_for("admin_dashboard", tab="escalas"))


# =========================
# AÃ§Ãµes de exclusÃ£o de escalas
# =========================
@app.post("/escalas/purge_all")
@admin_required
def escalas_purge_all():
    res = db.session.execute(sa_delete(Escala))
    db.session.commit()
    flash(f"Todas as escalas foram excluÃ­das ({res.rowcount or 0}).", "info")
    return redirect(url_for("admin_dashboard", tab="escalas"))

@app.post("/escalas/purge_cooperado/<int:coop_id>")
@admin_required
def escalas_purge_cooperado(coop_id):
    res = db.session.execute(sa_delete(Escala).where(Escala.cooperado_id == coop_id))
    db.session.commit()
    flash(f"Escalas do cooperado #{coop_id} excluÃ­das ({res.rowcount or 0}).", "info")
    return redirect(url_for("admin_dashboard", tab="escalas"))

@app.post("/escalas/purge_restaurante/<int:rest_id>")
@admin_required
def escalas_purge_restaurante(rest_id):
    res = db.session.execute(sa_delete(Escala).where(Escala.restaurante_id == rest_id))
    db.session.commit()
    flash(f"Escalas do restaurante #{rest_id} excluÃ­das ({res.rowcount or 0}).", "info")
    return redirect(url_for("admin_dashboard", tab="escalas"))

# =========================
# Trocas (Admin aprovar/recusar)
# =========================
@app.post("/admin/trocas/<int:id>/aprovar")
@admin_required
def admin_aprovar_troca(id):
    t = TrocaSolicitacao.query.get_or_404(id)
    if t.status != "pendente":
        flash("Esta solicitaÃ§Ã£o jÃ¡ foi tratada.", "warning")
        return redirect(url_for("admin_dashboard", tab="escalas"))

    orig_e = db.session.get(Escala, t.origem_escala_id)
    if not orig_e:
        flash("PlantÃ£o de origem invÃ¡lido.", "danger")
        return redirect(url_for("admin_dashboard", tab="escalas"))

    solicitante = db.session.get(Cooperado, t.solicitante_id)
    destinatario = db.session.get(Cooperado, t.destino_id)
    if not solicitante or not destinatario:
        flash("Cooperado(s) invÃ¡lido(s) na solicitaÃ§Ã£o.", "danger")
        return redirect(url_for("admin_dashboard", tab="escalas"))

    wd_o = _weekday_from_data_str(orig_e.data)
    buck_o = _turno_bucket(orig_e.turno, orig_e.horario)
    minhas = (Escala.query
              .filter_by(cooperado_id=destinatario.id)
              .order_by(Escala.id.asc()).all())
    candidatas = [e for e in minhas
                  if _weekday_from_data_str(e.data) == wd_o
                  and _turno_bucket(e.turno, e.horario) == buck_o]

    if len(candidatas) != 1:
        if len(candidatas) == 0:
            flash("Destino nÃ£o possui plantÃµes compatÃ­veis (mesmo dia da semana e mesmo turno).", "danger")
        else:
            flash("Mais de um plantÃ£o compatÃ­vel encontrado para o destino. Aprove pelo portal do cooperado (onde Ã© possÃ­vel escolher).", "warning")
        return redirect(url_for("admin_dashboard", tab="escalas"))

    dest_e = candidatas[0]

    linhas = [
        {
            "dia": _escala_label(orig_e).split(" â€¢ ")[0],
            "turno_horario": " â€¢ ".join([x for x in [(orig_e.turno or "").strip(), (orig_e.horario or "").strip()] if x]),
            "contrato": (orig_e.contrato or "").strip(),
            "saiu": solicitante.nome,
            "entrou": destinatario.nome,
        },
        {
            "dia": _escala_label(dest_e).split(" â€¢ ")[0],
            "turno_horario": " â€¢ ".join([x for x in [(dest_e.turno or "").strip(), (dest_e.horario or "").strip()] if x]),
            "contrato": (dest_e.contrato or "").strip(),
            "saiu": destinatario.nome,
            "entrou": solicitante.nome,
        }
    ]
    afetacao_json = {"linhas": linhas}

    solicitante_id = orig_e.cooperado_id
    destino_id = dest_e.cooperado_id
    orig_e.cooperado_id, dest_e.cooperado_id = destino_id, solicitante_id

    t.status = "aprovada"
    t.aplicada_em = datetime.now(timezone.utc)
    prefix = "" if not (t.mensagem and t.mensagem.strip()) else (t.mensagem.rstrip() + "\n")
    t.mensagem = prefix + "__AFETACAO_JSON__:" + json.dumps(afetacao_json, ensure_ascii=False)

    db.session.commit()
    flash("Troca aprovada e aplicada com sucesso!", "success")
    return redirect(url_for("admin_dashboard", tab="escalas"))

@app.post("/admin/trocas/<int:id>/recusar")
@admin_required
def admin_recusar_troca(id):
    t = TrocaSolicitacao.query.get_or_404(id)
    if t.status != "pendente":
        flash("Esta solicitaÃ§Ã£o jÃ¡ foi tratada.", "warning")
        return redirect(url_for("admin_dashboard", tab="escalas"))
    t.status = "recusada"
    db.session.commit()
    flash("SolicitaÃ§Ã£o recusada.", "info")
    return redirect(url_for("admin_dashboard", tab="escalas"))


# --- Admin tool: aplicar ON DELETE CASCADE nas FKs (Postgres) ---
@app.get("/admin/tools/apply_fk_cascade")
@admin_required
def apply_fk_cascade():
    """
    Aplica/garante ON DELETE CASCADE nas FKs relevantes (Postgres).
    Tudo estÃ¡ dentro de uma string SQL, evitando SyntaxError no deploy.
    """
    from sqlalchemy import text as sa_text

    sql = """
BEGIN;

-- =========================
-- AVALIAÃ‡Ã•ES (jÃ¡ existia)
-- =========================
-- ajusta FK de avaliacoes.lancamento_id
ALTER TABLE public.avaliacoes
  DROP CONSTRAINT IF EXISTS avaliacoes_lancamento_id_fkey;
ALTER TABLE public.avaliacoes
  ADD CONSTRAINT avaliacoes_lancamento_id_fkey
  FOREIGN KEY (lancamento_id)
  REFERENCES public.lancamentos (id)
  ON DELETE CASCADE;

-- cria/garante CASCADE para avaliacoes_restaurante.lancamento_id
DO $do$
BEGIN
    IF NOT EXISTS (
        SELECT 1
        FROM information_schema.table_constraints
        WHERE constraint_name = 'av_rest_lancamento_id_fkey'
          AND table_name = 'avaliacoes_restaurante'
          AND table_schema = 'public'
    ) THEN
        ALTER TABLE public.avaliacoes_restaurante
          ADD CONSTRAINT av_rest_lancamento_id_fkey
          FOREIGN KEY (lancamento_id)
          REFERENCES public.lancamentos (id)
          ON DELETE CASCADE;
    ELSE
        -- garante o CASCADE (drop/add)
        EXECUTE $$ALTER TABLE public.avaliacoes_restaurante
                 DROP CONSTRAINT IF EXISTS av_rest_lancamento_id_fkey$$;
        EXECUTE $$ALTER TABLE public.avaliacoes_restaurante
                 ADD CONSTRAINT av_rest_lancamento_id_fkey
                 FOREIGN KEY (lancamento_id)
                 REFERENCES public.lancamentos (id)
                 ON DELETE CASCADE$$;
    END IF;
END
$do$;

-- =========================
-- ESCALAS
-- =========================
-- cooperado_id -> cooperados(id) ON DELETE CASCADE
ALTER TABLE public.escalas
  DROP CONSTRAINT IF EXISTS escalas_cooperado_id_fkey;
ALTER TABLE public.escalas
  ADD CONSTRAINT escalas_cooperado_id_fkey
  FOREIGN KEY (cooperado_id)
  REFERENCES public.cooperados (id)
  ON DELETE CASCADE;

-- restaurante_id -> restaurantes(id) ON DELETE CASCADE
ALTER TABLE public.escalas
  DROP CONSTRAINT IF EXISTS escalas_restaurante_id_fkey;
ALTER TABLE public.escalas
  ADD CONSTRAINT escalas_restaurante_id_fkey
  FOREIGN KEY (restaurante_id)
  REFERENCES public.restaurantes (id)
  ON DELETE CASCADE;

-- =========================
-- TROCAS
-- =========================
-- solicitante_id -> cooperados(id) ON DELETE CASCADE
ALTER TABLE public.trocas
  DROP CONSTRAINT IF EXISTS trocas_solicitante_id_fkey;
ALTER TABLE public.trocas
  ADD CONSTRAINT trocas_solicitante_id_fkey
  FOREIGN KEY (solicitante_id)
  REFERENCES public.cooperados (id)
  ON DELETE CASCADE;

-- destino_id -> cooperados(id) ON DELETE CASCADE
ALTER TABLE public.trocas
  DROP CONSTRAINT IF EXISTS trocas_destino_id_fkey;
ALTER TABLE public.trocas
  ADD CONSTRAINT trocas_destino_id_fkey
  FOREIGN KEY (destino_id)
  REFERENCES public.cooperados (id)
  ON DELETE CASCADE;

-- origem_escala_id -> escalas(id) ON DELETE CASCADE
ALTER TABLE public.trocas
  DROP CONSTRAINT IF EXISTS trocas_origem_escala_id_fkey;
ALTER TABLE public.trocas
  ADD CONSTRAINT trocas_origem_escala_id_fkey
  FOREIGN KEY (origem_escala_id)
  REFERENCES public.escalas (id)
  ON DELETE CASCADE;

COMMIT;
"""

    try:
        if _is_sqlite():
            flash("SQLite local: esta operaÃ§Ã£o Ã© especÃ­fica de Postgres (sem efeito aqui).", "warning")
            return redirect(url_for("admin_dashboard", tab="config"))

        db.session.execute(sa_text(sql))
        db.session.commit()
        flash("FKs com ON DELETE CASCADE aplicadas com sucesso.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao aplicar FKs: {e}", "danger")
    return redirect(url_for("admin_dashboard", tab="config"))


# =========================
# Documentos (Admin)
# =========================
@app.route("/documentos/<int:coop_id>", methods=["GET", "POST"])
@admin_required
def editar_documentos(coop_id):
    c = Cooperado.query.get_or_404(coop_id)

    if request.method == "POST":
        f = request.form
        c.cnh_numero = f.get("cnh_numero")
        c.placa = f.get("placa")
        def parse_date_local(s):
            try:
                return datetime.strptime(s, "%Y-%m-%d").date() if s else None
            except Exception:
                return None
        c.cnh_validade = parse_date_local(f.get("cnh_validade"))
        c.placa_validade = parse_date_local(f.get("placa_validade"))
        c.ultima_atualizacao = datetime.now()
        db.session.commit()
        flash("Documentos atualizados.", "success")
        return redirect(url_for("admin_dashboard", tab="escalas"))

    tpl = os.path.join("templates", "editar_documentos.html")
    hoje = date.today()
    prazo_final = date(hoje.year, 12, 31)
    if os.path.exists(tpl):
        docinfo = {
            "prazo_final": prazo_final,
            "dias_ate_prazo": max(0, (prazo_final - hoje).days),
            "cnh": {
                "numero": c.cnh_numero,
                "validade": c.cnh_validade,
                "prox_validade": _prox_ocorrencia_anual(c.cnh_validade),
                "ok": (c.cnh_validade is not None and c.cnh_validade >= hoje),
                "modo": "auto",
            },
            "placa": {
                "numero": c.placa,
                "validade": c.placa_validade,
                "prox_validade": _prox_ocorrencia_anual(c.placa_validade),
                "ok": (c.placa_validade is not None and c.placa_validade >= hoje),
                "modo": "auto",
            }
        }
        return render_template("editar_documentos.html", cooperado=c, docinfo=docinfo)

    return f"""
    <div style="max-width:560px;margin:30px auto;font-family:Arial">
      <h3>Documentos â€” {c.nome}</h3>
      <form method="POST">
        <label>CNH (nÃºmero)</label><br>
        <input name="cnh_numero" value="{c.cnh_numero or ''}" style="width:100%;padding:8px"><br><br>
        <label>Validade CNH</label><br>
        <input type="date" name="cnh_validade" value="{c.cnh_validade.strftime('%Y-%m-%d') if c.cnh_validade else ''}" style="width:100%;padding:8px"><br><br>
        <label>Placa</label><br>
        <input name="placa" value="{c.placa or ''}" style="width:100%;padding:8px"><br><br>
        <label>Validade da Placa</label><br>
        <input type="date" name="placa_validade" value="{c.placa_validade.strftime('%Y-%m-%d') if c.placa_validade else ''}" style="width:100%;padding:8px"><br><br>
        <button style="padding:10px 16px">Salvar</button>
        <a href="{url_for('admin_dashboard', tab='escalas')}" style="margin-left:10px">Voltar</a>
      </form>
    </div>
    """

# =========================
# PORTAL COOPERADO
# =========================
@app.route("/portal/cooperado")
@role_required("cooperado")
def portal_cooperado():
    u_id = session.get("user_id")
    coop = Cooperado.query.filter_by(usuario_id=u_id).first()
    if not coop:
        return "<p style='font-family:Arial;margin:40px'>Seu usuÃ¡rio nÃ£o estÃ¡ vinculado a um cooperado. Avise o administrador.</p>"

    try:
        coop.usuario = coop.usuario_ref.usuario
    except Exception:
        coop.usuario = ""

    # ---------- FILTRO POR DATA (padrÃ£o = HOJE) ----------
    di = _parse_date(request.args.get("data_inicio"))
    df = _parse_date(request.args.get("data_fim"))

    # padrÃ£o: mostrar SOMENTE a data do lanÃ§amento (hoje)
    if di and not df:
        df = di
    if df and not di:
        di = df
    if not di and not df:
        di = df = date.today()

    def in_range(qs, col):
        return qs.filter(col >= di, col <= df)

    ql = in_range(Lancamento.query.filter_by(cooperado_id=coop.id), Lancamento.data)
    producoes = ql.order_by(Lancamento.data.desc(), Lancamento.id.desc()).all()

    # --- Marca se o cooperado jÃ¡ avaliou cada produÃ§Ã£o ---
    ids = [l.id for l in producoes]
    minhas = {}
    if ids:
        rows = (
            db.session.query(
                AvaliacaoRestaurante.lancamento_id,
                AvaliacaoRestaurante.estrelas_geral
            )
            .filter(
                AvaliacaoRestaurante.lancamento_id.in_(ids),
                AvaliacaoRestaurante.cooperado_id == coop.id
            )
            .all()
        )
        minhas = {lid: nota for lid, nota in rows}

    for l in producoes:
        l.minha_avaliacao = minhas.get(l.id)

    qr = in_range(ReceitaCooperado.query.filter_by(cooperado_id=coop.id), ReceitaCooperado.data)
    receitas_coop = qr.order_by(ReceitaCooperado.data.desc(), ReceitaCooperado.id.desc()).all()

    qd = in_range(DespesaCooperado.query.filter_by(cooperado_id=coop.id), DespesaCooperado.data)
    despesas_coop = qd.order_by(DespesaCooperado.data.desc(), DespesaCooperado.id.desc()).all()

    # INSS calculado por lanÃ§amento e somado APENAS dentro do perÃ­odo filtrado
    total_bruto = sum((l.valor or 0.0) for l in producoes) + sum((r.valor or 0.0) for r in receitas_coop)
    inss_valor = sum((l.valor or 0.0) * 0.045 for l in producoes)
    total_descontos = sum((d.valor or 0.0) for d in despesas_coop)
    total_liquido = total_bruto - inss_valor - total_descontos

    cfg = get_config()
    salario_minimo = cfg.salario_minimo or 0.0
    inss_complemento = salario_minimo * 0.20

    today = date.today()
    def dias_para_3112():
        alvo = date(today.year, 12, 31)
        if today > alvo:
            alvo = date(today.year + 1, 12, 31)
        return (alvo - today).days

    doc_cnh = {
        "numero": coop.cnh_numero,
        "vencimento": coop.cnh_validade,
        "ok": (coop.cnh_validade is not None and coop.cnh_validade >= today),
        "dias_para_prazo": dias_para_3112(),
    }
    doc_placa = {
        "numero": coop.placa,
        "vencimento": coop.placa_validade,
        "ok": (coop.placa_validade is not None and coop.placa_validade >= today),
        "dias_para_prazo": dias_para_3112(),
    }

    # ---------- ESCALA (dedupe + ordenaÃ§Ã£o cronolÃ³gica robusta) ----------
    raw_escala = (Escala.query
                  .filter_by(cooperado_id=coop.id)
                  .order_by(Escala.id.asc())
                  .all())

    import unicodedata as _u, re as _re
    def _norm_c(s: str) -> str:
        s = _u.normalize("NFD", str(s or "").lower())
        s = "".join(ch for ch in s if _u.category(ch) != "Mn")
        return _re.sub(r"[^a-z0-9]+", " ", s).strip()

    def _score(e):
        h = (e.horario or "").strip()
        return (1 if h else 0, len(h), e.id)

    # helper: parse data "dd/mm/aaaa"
    def _to_date_from_str(s: str):
        m = _re.search(r'(\d{1,2})/(\d{1,2})/(\d{2,4})', str(s or ''))
        if not m:
            return None
        d_, mth, y = map(int, m.groups())
        if y < 100:
            y += 2000
        try:
            return date(y, mth, d_)
        except Exception:
            return None

    # helper: minutos do horÃ¡rio "HH:MM"
    def _mins(h):
        m = _re.search(r'(\d{1,2}):(\d{2})', str(h or ''))
        if not m:
            return 24*60 + 59  # empurra vazios pro fim do dia
        hh, mm = map(int, m.groups())
        return hh*60 + mm

    # helper: bucket para ordenar dia antes de noite
    def _bucket_idx(turno, horario):
        b = (_turno_bucket(turno, horario) or "").lower()
        if "dia" in b:
            return 1
        if "noite" in b:
            return 2
        # fallback pelo horÃ¡rio
        mins = _mins(horario)
        return 2 if (mins >= 17*60 or mins <= 6*60) else 1

    # Escolhe â€œmelhorâ€ registro por (data/turno/contrato)
    best = {}
    for e in raw_escala:
        key = (_norm_c(e.data), _norm_c(e.turno), _norm_c(e.contrato))
        cur = best.get(key)
        if not cur or _score(e) > _score(cur):
            best[key] = e

    # Ordena cronologicamente (nÃ£o por id!)
    cand = list(best.values())
    for e in cand:
        d = _to_date_from_str(e.data) or date.min
        mins = _mins(e.horario or "")
        bidx = _bucket_idx(e.turno, e.horario)
        e._ord = (d.toordinal(), bidx, mins, (e.contrato or ""), e.id)

    minha_escala = sorted(cand, key=lambda x: x._ord)

    # ---------- Status/cores por data ----------
    for e in minha_escala:
        dt = _to_date_from_str(e.data)
        if dt is None:
            status = 'unknown'
        else:
            if dt < today:
                status = 'past'
            elif dt == today:
                status = 'today'
            elif dt == today + timedelta(days=1):
                status = 'tomorrow'
            else:
                status = 'future'
        e.status = status
        e.status_color = (
            '#ef4444' if status == 'past' else
            '#22c55e' if status == 'today' else
            '#3b82f6' if status in ('tomorrow', 'future') else
            'transparent'
        )

    # ---------- versÃ£o JSON jÃ¡ na MESMA ORDEM ----------
    minha_escala_json = []
    for e in minha_escala:
        minha_escala_json.append({
            "id": e.id,
            "data": e.data or "",
            "turno": e.turno or "",
            "horario": e.horario or "",
            "contrato": e.contrato or "",
            "weekday": _weekday_from_data_str(e.data),
            "turno_bucket": _turno_bucket(e.turno, e.horario),
        })

    # ---------- Trocas ----------
    coops = (Cooperado.query
             .filter(Cooperado.id != coop.id)
             .order_by(Cooperado.nome.asc())
             .all())
    cooperados_json = [
        {"id": c.id, "nome": c.nome, "foto_url": (c.foto_url or "")}
        for c in coops
    ]

    def _escala_desc(e: Escala | None) -> str:
        return _escala_label(e)

    rx = (TrocaSolicitacao.query
          .filter(TrocaSolicitacao.destino_id == coop.id)
          .order_by(TrocaSolicitacao.id.desc())
          .all())

    trocas_recebidas_pendentes = []
    trocas_recebidas_historico = []
    for t in rx:
        solicitante = db.session.get(Cooperado, t.solicitante_id)
        orig = db.session.get(Escala, t.origem_escala_id)

        mensagem_limpa = _strip_afetacao_blob(t.mensagem)
        linhas_afetadas = _parse_linhas_from_msg(t.mensagem) if t.status == "aprovada" else []

        item = {
            "id": t.id,
            "status": t.status,
            "mensagem": mensagem_limpa,
            "criada_em": t.criada_em,
            "aplicada_em": t.aplicada_em,
            "solicitante": solicitante,
            "origem": orig,
            "origem_desc": _escala_desc(orig),
            "linhas_afetadas": linhas_afetadas,
            "origem_weekday": _weekday_from_data_str(orig.data) if orig else None,
            "origem_turno_bucket": _turno_bucket(orig.turno if orig else None, orig.horario if orig else None),
        }

        (trocas_recebidas_pendentes if t.status == "pendente" else trocas_recebidas_historico).append(item)

    ex = (TrocaSolicitacao.query
          .filter(TrocaSolicitacao.solicitante_id == coop.id)
          .order_by(TrocaSolicitacao.id.desc())
          .all())

    trocas_enviadas = []
    for t in ex:
        destino = db.session.get(Cooperado, t.destino_id)
        orig = db.session.get(Escala, t.origem_escala_id)
        mensagem_limpa = _strip_afetacao_blob(t.mensagem)
        linhas_afetadas = _parse_linhas_from_msg(t.mensagem) if t.status == "aprovada" else []
        trocas_enviadas.append({
            "id": t.id,
            "status": t.status,
            "mensagem": mensagem_limpa,
            "criada_em": t.criada_em,
            "aplicada_em": t.aplicada_em,
            "destino": destino,
            "origem": orig,
            "origem_desc": _escala_desc(orig),
            "linhas_afetadas": linhas_afetadas,
        })

    return render_template(
        "painel_cooperado.html",
        cooperado=coop,
        producoes=producoes,
        receitas_coop=receitas_coop,
        despesas_coop=despesas_coop,
        total_bruto=total_bruto,
        inss_valor=inss_valor,
        total_descontos=total_descontos,
        total_liquido=total_liquido,
        inss_complemento=inss_complemento,
        salario_minimo=salario_minimo,
        current_year=today.year,
        doc_cnh=doc_cnh,
        doc_placa=doc_placa,
        minha_escala=minha_escala,
        minha_escala_json=minha_escala_json,
        cooperados_json=cooperados_json,
        trocas_recebidas_pendentes=trocas_recebidas_pendentes,
        trocas_recebidas_historico=trocas_recebidas_historico,
        trocas_enviadas=trocas_enviadas,
    )
from flask import request, redirect, url_for, flash, abort, session

@app.post("/coop/avaliar/restaurante/<int:lanc_id>")
@role_required("cooperado")
def coop_avaliar_restaurante(lanc_id):
    # 1) Cooperado logado
    u_id = session.get("user_id")
    coop = Cooperado.query.filter_by(usuario_id=u_id).first_or_404()

    # 2) LanÃ§amento existe e Ã© dele
    lanc = Lancamento.query.get_or_404(lanc_id)
    if lanc.cooperado_id != coop.id:
        abort(403)

    # 3) JÃ¡ existe avaliaÃ§Ã£o DESTE cooperado para ESTE lanÃ§amento?
    ja = (AvaliacaoRestaurante.query
          .filter_by(lancamento_id=lanc.id, cooperado_id=coop.id)
          .first())
    if ja:
        flash("VocÃª jÃ¡ avaliou esta produÃ§Ã£o.", "info")
        return redirect(request.referrer or url_for("coop_dashboard"))

    # 4) LÃª os campos do form (suporta 'nota' simples OU 'av_geral')
    f  = request.form
    g  = _clamp_star(f.get("nota") or f.get("av_geral"))   # obrigatÃ³rio
    p  = _clamp_star(f.get("av_pontualidade"))
    ed = _clamp_star(f.get("av_educacao"))
    ef = _clamp_star(f.get("av_eficiencia"))
    ap = _clamp_star(f.get("av_apresentacao"))
    tx = (f.get("av_comentario") or "").strip()

    if not g:
        flash("Selecione uma nota de 1 a 5.", "warning")
        return redirect(request.referrer or url_for("coop_dashboard"))

    # 5) Cria e salva
    a = AvaliacaoRestaurante(
        restaurante_id=lanc.restaurante_id,
        cooperado_id=coop.id,
        lancamento_id=lanc.id,
        estrelas_geral=g,
        estrelas_pontualidade=p,
        estrelas_educacao=ed,
        estrelas_eficiencia=ef,
        estrelas_apresentacao=ap,
        comentario=tx,
        media_ponderada=_media_ponderada(g, p, ed, ef, ap),
        sentimento=_analise_sentimento(tx),
        temas="; ".join(_identifica_temas(tx)),
        alerta_crise=_sinaliza_crise(g, tx),
        criado_em=datetime.now(timezone.utc),
    )
    db.session.add(a)
    db.session.commit()

    flash("AvaliaÃ§Ã£o do restaurante registrada.", "success")
    return redirect(request.referrer or url_for("coop_dashboard"))

# Alias para manter compatibilidade com o action do formulÃ¡rio
@app.post("/producoes/<int:lanc_id>/avaliar", endpoint="producoes_avaliar")
@role_required("cooperado")
def producoes_avaliar(lanc_id):
    # IMPORTANTÃSSIMO: retornar o que a funÃ§Ã£o real retorna
    return coop_avaliar_restaurante(lanc_id)

@app.route("/painel/cooperado")
@role_required("cooperado")
def coop_dashboard():
    return portal_cooperado()

@app.route("/escala/solicitar_troca", methods=["POST"])
@role_required("cooperado")
def solicitar_troca():
    u_id = session.get("user_id")
    me = Cooperado.query.filter_by(usuario_id=u_id).first()
    if not me:
        abort(403)

    from_escala_id = request.form.get("from_escala_id", type=int)
    to_cooperado_id = request.form.get("to_cooperado_id", type=int)
    mensagem = (request.form.get("mensagem") or "").strip()

    if not from_escala_id or not to_cooperado_id:
        flash("Selecione a escala e o cooperado de destino.", "warning")
        return redirect(url_for("portal_cooperado"))

    origem = db.session.get(Escala, from_escala_id)
    if not origem or origem.cooperado_id != me.id:
        flash("Escala invÃ¡lida para solicitaÃ§Ã£o.", "danger")
        return redirect(url_for("portal_cooperado"))

    destino = db.session.get(Cooperado, to_cooperado_id)
    if not destino or destino.id == me.id:
        flash("Cooperado de destino invÃ¡lido.", "danger")
        return redirect(url_for("portal_cooperado"))

    t = TrocaSolicitacao(
        solicitante_id=me.id,
        destino_id=destino.id,
        origem_escala_id=origem.id,
        mensagem=mensagem or None,
        status="pendente",
    )
    db.session.add(t)
    db.session.commit()
    flash("SolicitaÃ§Ã£o de troca enviada ao administrador.", "success")
    return redirect(url_for("portal_cooperado"))

@app.post("/trocas/<int:troca_id>/aceitar")
@role_required("cooperado")
def aceitar_troca(troca_id):
    u_id = session.get("user_id")
    me = Cooperado.query.filter_by(usuario_id=u_id).first()
    t = TrocaSolicitacao.query.get_or_404(troca_id)
    if not me or t.destino_id != me.id:
        abort(403)
    if t.status != "pendente":
        flash("Esta solicitaÃ§Ã£o jÃ¡ foi tratada.", "warning")
        return redirect(url_for("portal_cooperado"))

    destino_escala_id = request.form.get("destino_escala_id", type=int)
    orig_e = db.session.get(Escala, t.origem_escala_id)
    if not orig_e:
        flash("PlantÃ£o de origem invÃ¡lido.", "danger")
        return redirect(url_for("portal_cooperado"))

    if not destino_escala_id:
        minhas = Escala.query.filter_by(cooperado_id=me.id).order_by(Escala.id.asc()).all()
        wd_o = _weekday_from_data_str(orig_e.data)
        buck_o = _turno_bucket(orig_e.turno, orig_e.horario)
        candidatas = [e for e in minhas
                      if _weekday_from_data_str(e.data) == wd_o and _turno_bucket(e.turno, e.horario) == buck_o]
        if len(candidatas) == 1:
            destino_escala_id = candidatas[0].id
        elif len(candidatas) == 0:
            flash("VocÃª nÃ£o tem plantÃµes compatÃ­veis (mesmo dia da semana e turno).", "danger")
            return redirect(url_for("portal_cooperado"))
        else:
            flash("Selecione no modal qual dos seus plantÃµes compatÃ­veis deseja usar.", "warning")
            return redirect(url_for("portal_cooperado"))

    dest_e = db.session.get(Escala, destino_escala_id)
    if not dest_e or dest_e.cooperado_id != me.id:
        flash("SeleÃ§Ã£o de escala invÃ¡lida.", "danger")
        return redirect(url_for("portal_cooperado"))

    wd_orig = _weekday_from_data_str(orig_e.data)
    wd_dest = _weekday_from_data_str(dest_e.data)
    buck_orig = _turno_bucket(orig_e.turno, orig_e.horario)
    buck_dest = _turno_bucket(dest_e.turno, dest_e.horario)
    if wd_orig is None or wd_dest is None or wd_orig != wd_dest or buck_orig != buck_dest:
        flash("Troca incompatÃ­vel: precisa ser mesmo dia da semana e mesmo turno (dia/noite).", "danger")
        return redirect(url_for("portal_cooperado"))

    solicitante = db.session.get(Cooperado, t.solicitante_id)
    destinatario = me
    linhas = [
        {
            "dia": _escala_label(orig_e).split(" â€¢ ")[0],
            "turno_horario": " â€¢ ".join([x for x in [(orig_e.turno or "").strip(), (orig_e.horario or "").strip()] if x]),
            "contrato": (orig_e.contrato or "").strip(),
            "saiu": solicitante.nome,
            "entrou": destinatario.nome,
        },
        {
            "dia": _escala_label(dest_e).split(" â€¢ ")[0],
            "turno_horario": " â€¢ ".join([x for x in [(dest_e.turno or "").strip(), (dest_e.horario or "").strip()] if x]),
            "contrato": (dest_e.contrato or "").strip(),
            "saiu": destinatario.nome,
            "entrou": solicitante.nome,
        }
    ]
    afetacao_json = {"linhas": linhas}

    solicitante_id = orig_e.cooperado_id
    destino_id = dest_e.cooperado_id
    orig_e.cooperado_id, dest_e.cooperado_id = destino_id, solicitante_id

    t.status = "aprovada"
    t.aplicada_em = datetime.now(timezone.utc)
    prefix = "" if not (t.mensagem and t.mensagem.strip()) else (t.mensagem.rstrip() + "\n")
    t.mensagem = prefix + "__AFETACAO_JSON__:" + json.dumps(afetacao_json, ensure_ascii=False)

    db.session.commit()
    flash("Troca aplicada com sucesso!", "success")
    return redirect(url_for("portal_cooperado"))

@app.post("/trocas/<int:troca_id>/recusar")
@role_required("cooperado")
def recusar_troca(troca_id):
    u_id = session.get("user_id")
    me = Cooperado.query.filter_by(usuario_id=u_id).first()
    t = TrocaSolicitacao.query.get_or_404(troca_id)
    if not me or t.destino_id != me.id:
        abort(403)
    if t.status != "pendente":
        flash("Esta solicitaÃ§Ã£o jÃ¡ foi tratada.", "warning")
        return redirect(url_for("portal_cooperado"))

    t.status = "recusada"
    db.session.commit()
    flash("SolicitaÃ§Ã£o recusada.", "info")
    return redirect(url_for("portal_cooperado"))

# =========================
# PORTAL RESTAURANTE / FARMÃCIA
# =========================
@app.route("/portal/restaurante")
@role_required("restaurante", "farmacia")  # aceita mÃºltiplos tipos
def portal_restaurante():
    u_id = session.get("user_id")
    if not u_id:
        abort(401)

    # Encontra o estabelecimento vinculado a este usuÃ¡rio
    rest = Restaurante.query.filter_by(usuario_id=u_id).first()
    if not rest:
        return (
            "<p style='font-family:Arial;margin:40px'>Seu usuÃ¡rio nÃ£o estÃ¡ vinculado a um "
            "estabelecimento. Avise o administrador.</p>"
        )

    # Detecta se o login Ã© de farmÃ¡cia
    user_tipo = (rest.usuario_ref.tipo if rest.usuario_ref else None) or session.get("user_tipo")
    eh_farmacia = (str(user_tipo or "").lower() == "farmacia")

    # Abas/visÃµes: 'lancar', 'escalas', 'lancamentos', 'config', 'avisos'
    view = (request.args.get("view", "lancar") or "lancar").strip().lower()

    # Bloqueia as visÃµes que farmÃ¡cia nÃ£o deve ver
    if eh_farmacia and view in {"escalas", "lancamentos"}:
        view = "lancar"

    # Flags para o template (mostrar/ocultar abas)
    show_cooperados = not eh_farmacia
    show_lancamentos = not eh_farmacia

    # ---- helper mÃªs YYYY-MM
    import re
    def _parse_yyyy_mm_local(s: str):
        if not s:
            return None, None
        m = re.fullmatch(r"(\d{4})-(\d{2})", s.strip())
        if not m:
            return None, None
        y = int(m.group(1)); mth = int(m.group(2))
        try:
            di_ = date(y, mth, 1)
            if mth == 12:
                df_ = date(y + 1, 1, 1) - timedelta(days=1)
            else:
                df_ = date(y, mth + 1, 1) - timedelta(days=1)
            return di_, df_
        except Exception:
            return None, None

    # -------------------- LANÃ‡AMENTOS (totais por perÃ­odo) --------------------
    di = _parse_date(request.args.get("data_inicio"))
    df = _parse_date(request.args.get("data_fim"))

    # Filtro por mÃªs (?mes=YYYY-MM)
    mes = (request.args.get("mes") or "").strip()
    periodo_desc = None
    if mes:
        di_mes, df_mes = _parse_yyyy_mm_local(mes)
        if di_mes and df_mes:
            di, df = di_mes, df_mes
            periodo_desc = "mÃªs"

    if not di or not df:
        # Sem filtro => janela semanal baseada no perÃ­odo do restaurante
        wd_map = {"seg-dom": 0, "sab-sex": 5, "sex-qui": 4}  # seg=0 ... dom=6
        start_wd = wd_map.get(rest.periodo, 0)
        hoje = date.today()
        delta = (hoje.weekday() - start_wd) % 7
        di_auto = hoje - timedelta(days=delta)
        df_auto = di_auto + timedelta(days=6)
        di = di or di_auto
        df = df or df_auto
        periodo_desc = periodo_desc or rest.periodo
    else:
        periodo_desc = periodo_desc or "personalizado"

    cooperados = Cooperado.query.order_by(Cooperado.nome).all()

    total_bruto = 0.0
    total_qtd = 0
    total_entregas = 0
    for c in cooperados:
        q = (
            Lancamento.query
            .filter_by(restaurante_id=rest.id, cooperado_id=c.id)
            .filter(Lancamento.data >= di, Lancamento.data <= df)
            .order_by(Lancamento.data.desc(), Lancamento.id.desc())
        )
        c.lancamentos = q.all()
        c.total_periodo = sum((l.valor or 0.0) for l in c.lancamentos)
        total_bruto += c.total_periodo
        total_qtd += len(c.lancamentos)
        total_entregas += sum((l.qtd_entregas or 0) for l in c.lancamentos)

    total_inss = total_bruto * 0.045
    total_liquido = total_bruto - total_inss

    # -------------------- ESCALA --------------------
    def contrato_bate_restaurante(contrato: str, rest_nome: str) -> bool:
        a = " ".join(_normalize_name(contrato or ""))
        b = " ".join(_normalize_name(rest_nome or ""))
        if not a or not b:
            return False
        return a == b or a in b or b in a

    ref = _parse_date(request.args.get("ref")) or date.today()
    modo = request.args.get("modo", "semana")  # 'semana' ou 'dia'
    if modo == "dia":
        dias_list = [ref]
    else:
        semana_inicio = ref - timedelta(days=ref.weekday())
        dias_list = [semana_inicio + timedelta(days=i) for i in range(7)]

    escalas_all = Escala.query.order_by(Escala.id.asc()).all()
    eff_map = _carry_forward_contrato(escalas_all)

    escalas_rest = [
        e for e in escalas_all
        if contrato_bate_restaurante(eff_map.get(e.id, e.contrato or ""), rest.nome)
    ]
    if not escalas_rest:
        escalas_rest = [e for e in escalas_all if (e.contrato or "").strip() == rest.nome.strip()]

    agenda = {d: [] for d in dias_list}
    seen = {d: set() for d in dias_list}  # evita duplicar

    for e in escalas_rest:
        dt = _parse_data_escala_str(e.data)       # date | None
        wd = _weekday_from_data_str(e.data)       # 1..7 | None
        for d in dias_list:
            hit = (dt and dt == d) or (wd and wd == ((d.weekday() % 7) + 1))
            if not hit:
                continue

            coop = db.session.get(Cooperado, e.cooperado_id) if e.cooperado_id else None
            nome_fallback = (e.cooperado_nome or "").strip()
            nome_show = (coop.nome if coop else nome_fallback) or "â€”"
            contrato_eff = (eff_map.get(e.id, e.contrato or "") or "").strip()

            key = (
                (coop.id if coop else _norm(nome_show)),
                _norm(e.turno),
                _norm(e.horario),
                _norm(contrato_eff),
            )
            if key in seen[d]:
                break
            seen[d].add(key)

            agenda[d].append({
                "coop": coop,
                "cooperado_nome": nome_fallback or None,
                "nome_planilha": nome_show,
                "turno": (e.turno or "").strip(),
                "horario": (e.horario or "").strip(),
                "contrato": contrato_eff,
                "cor": (e.cor or "").strip(),
            })
            break

    for d in dias_list:
        agenda[d].sort(
            key=lambda x: (
                (x["contrato"] or "").lower(),
                (x.get("nome_planilha") or (x["coop"].nome if x["coop"] else "")).lower()
            )
        )

    # -------------------- Lista de lanÃ§amentos (aba "lancamentos") --------------------
    lancamentos_periodo = []
    total_lanc_valor = 0.0
    total_lanc_entregas = 0

    if (not eh_farmacia) and view == "lancamentos":
        q = (
            db.session.query(Lancamento, Cooperado)
            .join(Cooperado, Cooperado.id == Lancamento.cooperado_id)
            .filter(
                Lancamento.restaurante_id == rest.id,
                Lancamento.data >= di,
                Lancamento.data <= df,
            )
            .order_by(Lancamento.data.asc(), Lancamento.id.asc())
        )
        for lanc, coop in q.all():
            item = {
                "id": lanc.id,
                "data": lanc.data.strftime("%d/%m/%Y") if lanc.data else "",
                "hora_inicio": (
                    lanc.hora_inicio if isinstance(lanc.hora_inicio, str)
                    else (lanc.hora_inicio.strftime("%H:%M") if lanc.hora_inicio else "")
                ),
                "hora_fim": (
                    lanc.hora_fim if isinstance(lanc.hora_fim, str)
                    else (lanc.hora_fim.strftime("%H:%M") if lanc.hora_fim else "")
                ),
                "qtd_entregas": lanc.qtd_entregas or 0,
                "valor": float(lanc.valor or 0.0),
                "descricao": (lanc.descricao or ""),
                "cooperado_id": coop.id,
                "cooperado_nome": coop.nome,
                "contrato_nome": rest.nome,
                "status_entrega": getattr(lanc, "status_entrega", None),
                "motivo_nao_entregue": getattr(lanc, "motivo_nao_entregue", None),
                "recebido_por": getattr(lanc, "recebido_por", None),
            }
            lancamentos_periodo.append(item)

        total_lanc_valor = sum(x["valor"] for x in lancamentos_periodo)
        total_lanc_entregas = sum(x["qtd_entregas"] for x in lancamentos_periodo)

    # ---- URLs/flags para template
    from werkzeug.routing import BuildError
    try:
        url_lancar_producao = url_for("lancar_producao")
    except BuildError:
        url_lancar_producao = "/restaurante/lancar_producao"

    has_editar_lanc = ("editar_lancamento" in app.view_functions)

    # -------------------- Render --------------------
    return render_template(
        "restaurante_dashboard.html",
        rest=rest,
        cooperados=cooperados,
        filtro_inicio=di,
        filtro_fim=df,
        filtro_mes=(mes or ""),
        periodo_desc=periodo_desc,
        total_bruto=total_bruto,
        total_inss=total_inss,
        total_liquido=total_liquido,
        total_qtd=total_qtd,
        total_entregas=total_entregas,
        view=view,
        agenda=agenda,
        dias_list=dias_list,
        ref_data=ref,
        modo=modo,
        lancamentos_periodo=(lancamentos_periodo if (not eh_farmacia and view == "lancamentos") else []),
        total_lanc_valor=total_lanc_valor,
        total_lanc_entregas=total_lanc_entregas,
        url_lancar_producao=url_lancar_producao,
        has_editar_lanc=has_editar_lanc,
        show_cooperados=show_cooperados,
        show_lancamentos=show_lancamentos,
    )



# =========================
# Documentos (Admin + PÃºblico)
# =========================
@app.route("/admin/documentos")
@admin_required
def admin_documentos():
    documentos = Documento.query.order_by(Documento.enviado_em.desc()).all()
    return render_template("admin_documentos.html", documentos=documentos)

@app.post("/admin/documentos/upload")
@admin_required
def admin_upload_documento():
    f = request.form
    titulo = (f.get("titulo") or "").strip()
    categoria = (f.get("categoria") or "outro").strip()
    descricao = (f.get("descricao") or "").strip()
    arquivo = request.files.get("arquivo")

    if not titulo or not (arquivo and arquivo.filename):
        flash("Preencha o tÃ­tulo e selecione o arquivo.", "warning")
        return redirect(url_for("admin_documentos"))

    # === NOVO: salva em diretÃ³rio persistente e retorna NOME ÃšNICO ===
    nome_unico = salvar_documento_upload(arquivo)
    if not nome_unico:
        flash("Falha ao salvar o arquivo.", "danger")
        return redirect(url_for("admin_documentos"))

    # compat: tambÃ©m guardamos um URL que aponta para /docs/<nome>
    d = Documento(
        titulo=titulo,
        categoria=categoria,
        descricao=descricao,
        arquivo_url=url_for("serve_documento", nome=nome_unico),  # compat com templates antigos
        arquivo_nome=nome_unico,  # agora guardamos o NOME ÃšNICO persistido
        enviado_em=datetime.now(timezone.utc),
    )
    db.session.add(d)
    db.session.commit()
    flash("Documento enviado.", "success")
    return redirect(url_for("admin_documentos"))


@app.get("/admin/documentos/<int:doc_id>/delete")
@admin_required
def admin_delete_documento(doc_id):
    d = Documento.query.get_or_404(doc_id)
    try:
        # === NOVO: tenta deletar do armazenamento persistente pelo nome salvo ===
        p = resolve_document_path(d.arquivo_nome)
        if p and os.path.exists(p):
            os.remove(p)
        # Fallback (legado): se sobrou um caminho em /static/uploads/docs/ no arquivo_url, remove tambÃ©m
        if d.arquivo_url and d.arquivo_url.startswith("/static/uploads/docs/"):
            legacy_path = os.path.join(BASE_DIR, d.arquivo_url.lstrip("/"))
            if os.path.exists(legacy_path):
                os.remove(legacy_path)
    except Exception:
        pass

    db.session.delete(d)
    db.session.commit()
    flash("Documento removido.", "success")
    return redirect(url_for("admin_documentos"))


@app.route("/documentos")
def documentos_publicos():
    uid = session.get("user_id")
    if not uid:
        return redirect(url_for("login"))
    documentos = Documento.query.order_by(Documento.enviado_em.desc()).all()
    return render_template("documentos_publicos.html", documentos=documentos)


@app.route('/documentos/<int:doc_id>/baixar')
def baixar_documento(doc_id):
    doc = Documento.query.get_or_404(doc_id)
    # === NOVO: resolve caminho persistente pelo nome salvo ===
    path = resolve_document_path(doc.arquivo_nome)
    if not path or not os.path.exists(path):
        abort(404)
    # forÃ§a download (independente do tipo)
    return send_file(
        path,
        as_attachment=True,
        download_name=os.path.basename(doc.arquivo_nome)
    )

# routes/portal_restaurante_avisos.py
from flask import Blueprint, render_template, jsonify, request, abort
from flask_login import login_required
import sqlalchemy as sa
from sqlalchemy import select
from sqlalchemy.dialects.postgresql import insert as pg_insert

from app import db
from utils.auth import role_required
from utils.context import _restaurante_atual  # implemente/importe seu helper
restaurante_bp = Blueprint("restaurante", __name__, url_prefix="/restaurante")

# ------------ Helper defensivo p/ RESTAURANTE ------------
def get_avisos_for_restaurante(rest):
    """
    Tenta cobrir os esquemas mais comuns:
      - Aviso.ativo (bool)
      - Aviso.restaurante_id (avisos direcionados)
      - Aviso.publico_geral ou Aviso.global (avisos gerais)
      - Aviso.destinatario_tipo / Aviso.alvo ('todos'/'restaurante')
    Ajuste os nomes dos campos conforme seu modelo real.
    """
    q = db.session.query(Aviso)

    # ativo?
    if hasattr(Aviso, "ativo"):
        q = q.filter(Aviso.ativo.is_(True))

    # CondiÃ§Ãµes "diretas" (avisos para este restaurante)
    direct_conds = []
    if hasattr(Aviso, "restaurante_id"):
        direct_conds.append(Aviso.restaurante_id == rest.id)

    # CondiÃ§Ãµes "globais"
    global_conds = []
    if hasattr(Aviso, "publico_geral"):
        global_conds.append(Aviso.publico_geral.is_(True))
    if hasattr(Aviso, "global_"):
        global_conds.append(getattr(Aviso, "global_").is_(True))  # se usar nome reservado
    if hasattr(Aviso, "destinatario_tipo"):
        global_conds.append(Aviso.destinatario_tipo.in_(["todos", "restaurante"]))
    if hasattr(Aviso, "alvo"):
        global_conds.append(Aviso.alvo.in_(["todos", "restaurante"]))

    if direct_conds or global_conds:
        q = q.filter(sa.or_(*(direct_conds + global_conds)))

    # OrdenaÃ§Ã£o padrÃ£o
    if hasattr(Aviso, "criado_em"):
        q = q.order_by(Aviso.criado_em.desc())
    else:
        q = q.order_by(Aviso.id.desc())

    return q.all()

# ------------ LISTA HTML ------------
@restaurante_bp.get("/avisos", endpoint="portal_restaurante_avisos")
@role_required("restaurante")
def avisos_list_restaurante():
    rest = _restaurante_atual() or abort(403)
    avisos = get_avisos_for_restaurante(rest)

    lidos_ids = {
        r.aviso_id
        for r in db.session.execute(
            select(AvisoLeitura.aviso_id).where(AvisoLeitura.restaurante_id == rest.id)
        ).scalars().all()
    }

    for a in avisos:
        a.lido = (a.id in lidos_ids)

    avisos_nao_lidos_count = sum(1 for a in avisos if not getattr(a, "lido", False))
    current_year = datetime.now().year

    return render_template(
        "portal_restaurante_avisos.html",
        avisos=avisos,
        avisos_nao_lidos_count=avisos_nao_lidos_count,
        current_year=current_year
    )

# ------------ CONTADOR JSON ------------
@restaurante_bp.get("/avisos/unread_count", endpoint="portal_restaurante_avisos_unread_count")
@role_required("restaurante")
@login_required
def avisos_unread_count_restaurante():
    rest = _restaurante_atual() or abort(403)
    avisos = get_avisos_for_restaurante(rest)
    ids = [a.id for a in avisos]

    if not ids:
        resp = jsonify({"unread": 0})
        resp.headers["Cache-Control"] = "no-store"
        return resp

    lidos_ids = {
        r.aviso_id
        for r in db.session.execute(
            select(AvisoLeitura.aviso_id)
            .where(
                AvisoLeitura.restaurante_id == rest.id,
                AvisoLeitura.aviso_id.in_(ids)
            )
        ).scalars().all()
    }
    unread = len(ids) - len(lidos_ids)
    resp = jsonify({"unread": unread})
    resp.headers["Cache-Control"] = "no-store"
    return resp

# ------------ MARCAR LIDO (bulk / idempotente) ------------
@restaurante_bp.post("/avisos/marcar_lido")
@role_required("restaurante")
@login_required
def restaurante_marcar_lido():
    rest = _restaurante_atual() or abort(403)
    data = request.get_json(silent=True) or {}
    ids = [int(i) for i in (data.get("ids") or [])]
    if not ids:
        return {"ok": True, "count": 0}

    values = [{"aviso_id": i, "cooperado_id": None, "restaurante_id": rest.id} for i in ids]
    stmt = pg_insert(AvisoLeitura).values(values).on_conflict_do_nothing(
        constraint="uq_aviso_dest"
    )
    db.session.execute(stmt)
    db.session.commit()
    return {"ok": True, "count": len(ids)}


# =========================
# InicializaÃ§Ã£o automÃ¡tica do DB em servidores (Gunicorn/Render)
# =========================
try:
    with app.app_context():
        init_db()
except Exception as _e:
    # Evita crash no import; logs Ãºteis no servidor
    try:
        app.logger.warning(f"Falha ao inicializar DB: {_e}")
    except Exception:
        pass


@app.errorhandler(413)
def too_large(e):
    flash("Arquivo excede o tamanho mÃ¡ximo permitido (32MB).", "danger")
    return redirect(url_for('admin_documentos'))

# =========================
# InicializaÃ§Ã£o automÃ¡tica do DB em servidores (Gunicorn/Render)
# =========================
try:
    with app.app_context():
        init_db()
except Exception as _e:
    ...

# ==== TABELAS (upload/abrir/baixar) =========================================
from flask import (
    render_template, request, redirect, url_for, flash, session,
    send_file, abort, current_app, jsonify
)
from werkzeug.utils import secure_filename
from pathlib import Path
import os, re, unicodedata, mimetypes, logging

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# BASE_DIR e TABELAS_DIR (sempre salva/serve de static/uploads/tabelas)
# ---------------------------------------------------------------------------
try:
    BASE_DIR  # type: ignore[name-defined]
except NameError:
    BASE_DIR = Path(__file__).resolve().parent

# SEMPRE neste local:
TABELAS_DIR = str(Path(BASE_DIR) / "static" / "uploads" / "tabelas")

# Requer no app principal:
# - db (SQLAlchemy)
# - modelos: Tabela(id, titulo, descricao?, arquivo_url, arquivo_nome?, enviado_em)
#            Restaurante(usuario_id, nome?, usuario?, usuario_ref?.usuario?)
# - decorators: admin_required, role_required (se usar avisos/portais)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _tabelas_base_dir() -> Path:
    p = Path(TABELAS_DIR)
    p.mkdir(parents=True, exist_ok=True)
    return p

def _norm_txt(s: str) -> str:
    s = unicodedata.normalize("NFD", (s or "").strip())
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = re.sub(r"\s+", " ", s)
    return s.lower()

def _guess_mimetype_from_path(path: str) -> str:
    mt, _ = mimetypes.guess_type(path)
    return mt or "application/octet-stream"

def _enforce_restaurante_titulo(tabela, restaurante):
    """
    Regra: restaurante sÃ³ acessa a tabela cujo TÃTULO == NOME/LOGIN do restaurante (normalizado).
    """
    login_nome = (
        getattr(getattr(restaurante, "usuario_ref", None), "usuario", None)
        or getattr(restaurante, "usuario", None)
        or (restaurante.nome or "")
    )
    if _norm_txt(tabela.titulo) != _norm_txt(login_nome):
        abort(403)

def _serve_tabela_or_redirect(tabela, *, as_attachment: bool):
    """
    Resolve e serve o arquivo da Tabela:
    - http(s) => redirect
    - sempre tenta primeiro static/uploads/tabelas/<arquivo>
    - aceita absoluto, relativo e sÃ³ o nome
    - ignora querystring/fragments (ex.: foo.pdf?v=123#x)
    """
    url = (tabela.arquivo_url or "").strip()
    if not url:
        abort(404)

    # URL externa
    if url.startswith(("http://", "https://")):
        return redirect(url)

    base_dir    = Path(BASE_DIR)
    tabelas_dir = _tabelas_base_dir()

    # normaliza: remove "/" inicial, query e fragment
    raw = url.lstrip("/")
    raw_no_q = raw.split("?", 1)[0].split("#", 1)[0]
    fname = (raw_no_q.split("/")[-1] if raw_no_q else "").strip()

    candidates = []

    # 1) SEMPRE prioriza nosso diretÃ³rio oficial
    if fname:
        candidates.append(tabelas_dir / fname)

    # 2) Como veio, relativo ao BASE_DIR (compat c/ legado: static/uploads/tabelas/...)
    candidates.append(base_dir / raw_no_q)

    # 3) Absoluto (se alguÃ©m gravou caminho completo por engano)
    p = Path(url)
    if p.is_absolute():
        candidates.append(p)

    # 4) Mais dois legados comuns
    if fname:
        candidates.append(base_dir / "uploads" / "tabelas" / fname)
        candidates.append(base_dir / "static" / "uploads" / "tabelas" / fname)

    file_path = next((c for c in candidates if c.exists() and c.is_file()), None)
    if not file_path:
        try:
            log.warning(
                "Arquivo de Tabela nÃ£o encontrado. id=%s titulo=%r arquivo_url=%r tents=%r",
                getattr(tabela, "id", None),
                getattr(tabela, "titulo", None),
                tabela.arquivo_url,
                [str(c) for c in candidates],
            )
        except Exception:
            pass
        abort(404)

    return send_file(
        str(file_path),
        as_attachment=as_attachment,
        download_name=(tabela.arquivo_nome or file_path.name),
        mimetype=_guess_mimetype_from_path(str(file_path)),
    )

# ---------------------------------------------------------------------------
# Admin: listar / upload / delete
# ---------------------------------------------------------------------------
@app.get("/admin/tabelas", endpoint="admin_tabelas")
@admin_required
def admin_tabelas():
    tabelas = Tabela.query.order_by(Tabela.enviado_em.desc(), Tabela.id.desc()).all()
    return render_template("admin_tabelas.html", tabelas=tabelas)

@app.post("/admin/tabelas/upload", endpoint="admin_upload_tabela")
@admin_required
def admin_upload_tabela():
    f = request.form
    titulo    = (f.get("titulo") or "").strip()
    descricao = (f.get("descricao") or "").strip() or None

    # aceita vÃ¡rios nomes possÃ­veis do input file
    arquivo = (
        request.files.get("arquivo")
        or request.files.get("file")
        or request.files.get("tabela")
    )

    if not titulo or not (arquivo and arquivo.filename):
        flash("Preencha o tÃ­tulo e selecione o arquivo.", "warning")
        return redirect(url_for("admin_tabelas"))

    base_dir = _tabelas_base_dir()

    # nome seguro + timestamp pra nÃ£o colidir
    raw = secure_filename(arquivo.filename)
    stem, ext = os.path.splitext(raw)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    safe_stem = re.sub(r"[^A-Za-z0-9_.-]+", "-", stem) or "arquivo"
    final_name = f"{safe_stem}_{ts}{ext or ''}"

    dest = base_dir / final_name
    arquivo.save(str(dest))

    t = Tabela(
        titulo=titulo,
        descricao=descricao,
        # Importante: gravar apenas o NOME, nÃ£o o caminho.
        # O _serve_tabela_or_redirect vai resolver para TABELAS_DIR.
        arquivo_url=final_name,
        arquivo_nome=arquivo.filename,
        enviado_em=datetime.now(timezone.utc),
    )
    db.session.add(t)
    db.session.commit()

    flash("Tabela publicada.", "success")
    return redirect(url_for("admin_tabelas"))

@app.get("/admin/tabelas/<int:tab_id>/delete", endpoint="admin_delete_tabela")
@admin_required
def admin_delete_tabela(tab_id: int):
    t = Tabela.query.get_or_404(tab_id)
    try:
        url = (t.arquivo_url or "").strip()
        if url and not url.startswith(("http://", "https://")):
            path = _tabelas_base_dir() / (url.split("/")[-1])
            if path.exists():
                path.unlink()
    except Exception:
        pass
    db.session.delete(t)
    db.session.commit()
    flash("Tabela excluÃ­da.", "success")
    return redirect(url_for("admin_tabelas"))

# ---------------------------------------------------------------------------
# Cooperado/Admin/Restaurante: listagem (cooperado vÃª TODAS)
# MantÃ©m endpoints esperados pelo HTML: 'tabela_abrir' e 'baixar_tabela'
# ---------------------------------------------------------------------------
@app.get("/tabelas", endpoint="tabelas_publicas")
def tabelas_publicas():
    if session.get("user_tipo") not in {"admin", "cooperado", "restaurante"}:
        return redirect(url_for("login"))

    tabs = Tabela.query.order_by(Tabela.enviado_em.desc(), Tabela.id.desc()).all()

    items = [{
        "id": t.id,
        "titulo": t.titulo,
        "descricao": getattr(t, "descricao", None),
        "enviado_em": t.enviado_em,
        "arquivo_nome": getattr(t, "arquivo_nome", None),
        "abrir_url":  url_for("tabela_abrir",  tab_id=t.id),
        "baixar_url": url_for("baixar_tabela", tab_id=t.id),
    } for t in tabs]

    back_href = url_for("portal_cooperado") if (
        session.get("user_tipo") == "cooperado" and "portal_cooperado" in current_app.view_functions
    ) else ""

    return render_template("tabelas_publicas.html", tabelas=tabs, items=items, back_href=back_href)

# ---------------------------------------------------------------------------
# Abrir / Baixar compartilhado (cooperado/admin/restaurante)
# Endpoints compatÃ­veis com HTML existente:
#   - 'tabela_abrir'
#   - 'baixar_tabela'
# ---------------------------------------------------------------------------
@app.get("/tabelas/<int:tab_id>/abrir", endpoint="tabela_abrir")
def tabela_abrir(tab_id: int):
    if session.get("user_tipo") not in {"admin", "cooperado", "restaurante"}:
        return redirect(url_for("login"))
    t = Tabela.query.get_or_404(tab_id)

    if session.get("user_tipo") == "restaurante":
        rest = Restaurante.query.filter_by(usuario_id=session.get("user_id")).first_or_404()
        _enforce_restaurante_titulo(t, rest)

    return _serve_tabela_or_redirect(t, as_attachment=False)

@app.get("/tabelas/<int:tab_id>/baixar", endpoint="baixar_tabela")
def tabela_baixar(tab_id: int):
    if session.get("user_tipo") not in {"admin", "cooperado", "restaurante"}:
        return redirect(url_for("login"))
    t = Tabela.query.get_or_404(tab_id)

    if session.get("user_tipo") == "restaurante":
        rest = Restaurante.query.filter_by(usuario_id=session.get("user_id")).first_or_404()
        _enforce_restaurante_titulo(t, rest)

    return _serve_tabela_or_redirect(t, as_attachment=True)

# ---------------------------------------------------------------------------
# Restaurante: vÃª/abre/baixa SOMENTE a prÃ³pria tabela
# Endpoints compatÃ­veis com HTML existente:
#   - 'rest_tabela_abrir'
#   - 'rest_tabela_download'
# ---------------------------------------------------------------------------
@app.get("/rest/tabelas", endpoint="rest_tabelas")
def rest_tabelas():
    if session.get("user_tipo") != "restaurante":
        return redirect(url_for("login"))

    u_id = session.get("user_id")
    rest = Restaurante.query.filter_by(usuario_id=u_id).first_or_404()

    login_nome = (
        getattr(getattr(rest, "usuario_ref", None), "usuario", None)
        or getattr(rest, "usuario", None)
        or (rest.nome or "")
    )
    alvo_norm = _norm_txt(login_nome)

    candidatos = Tabela.query.order_by(Tabela.enviado_em.desc()).all()
    tabela_exata = next((t for t in candidatos if _norm_txt(t.titulo) == alvo_norm), None)

    has_portal_restaurante = ("portal_restaurante" in current_app.view_functions)

    return render_template(
        "restaurantes_tabelas.html",
        restaurante=rest,
        login_nome=login_nome,
        tabela=tabela_exata,  # o template deve lidar com None (sem tabela)
        has_portal_restaurante=has_portal_restaurante,
        back_href=url_for("portal_restaurante") if has_portal_restaurante else url_for("rest_tabelas"),
        current_year=datetime.now(timezone.utc).year,
    )

@app.get("/rest/tabelas/<int:tabela_id>/abrir", endpoint="rest_tabela_abrir")
def rest_tabela_abrir(tabela_id: int):
    if session.get("user_tipo") != "restaurante":
        return redirect(url_for("login"))
    rest = Restaurante.query.filter_by(usuario_id=session.get("user_id")).first_or_404()
    t = Tabela.query.get_or_404(tabela_id)
    _enforce_restaurante_titulo(t, rest)
    return _serve_tabela_or_redirect(t, as_attachment=False)

@app.get("/rest/tabelas/<int:tabela_id>/download", endpoint="rest_tabela_download")
def rest_tabela_download(tabela_id: int):
    if session.get("user_tipo") != "restaurante":
        return redirect(url_for("login"))
    rest = Restaurante.query.filter_by(usuario_id=session.get("user_id")).first_or_404()
    t = Tabela.query.get_or_404(tabela_id)
    _enforce_restaurante_titulo(t, rest)
    return _serve_tabela_or_redirect(t, as_attachment=True)

# ---------------------------------------------------------------------------
# DiagnÃ³stico rÃ¡pido (admin)
# ---------------------------------------------------------------------------
@app.get("/admin/tabelas/scan", endpoint="admin_tabelas_scan")
@admin_required
def admin_tabelas_scan():
    base = _tabelas_base_dir()
    items = []
    for t in Tabela.query.order_by(Tabela.enviado_em.desc()).all():
        url = (t.arquivo_url or "").strip()
        # sempre tentar resolver como gravamos hoje (fname dentro de static/uploads/tabelas)
        fname = (url.split("?", 1)[0].split("#", 1)[0]).split("/")[-1] if url else ""
        resolved = str(base / fname) if fname else url
        exists = (Path(resolved).exists() if fname else False)
        items.append({
            "id": t.id,
            "titulo": t.titulo,
            "arquivo_nome": t.arquivo_nome,
            "arquivo_url": t.arquivo_url,
            "resolved": resolved,
            "exists": bool(exists),
        })
    return jsonify({"tabelas_dir": str(base), "items": items})

# ---------------------------------------------------------------------------
# (Opcional) Normalizador: deixa arquivo_url sÃ³ com o NOME do arquivo
# Use se no seu banco ficaram caminhos tipo "static/uploads/tabelas/foo.pdf"
# ---------------------------------------------------------------------------
@app.post("/admin/tabelas/normalize-arquivo-url", endpoint="admin_tabelas_normalize_arquivo_url")
@admin_required
def admin_tabelas_normalize_arquivo_url():
    alterados = 0
    for t in Tabela.query.all():
        url = (t.arquivo_url or "").strip()
        if not url or url.startswith(("http://", "https://")):
            continue
        fname = url.split("?", 1)[0].split("#", 1)[0].split("/")[-1]
        if fname and fname != url:
            t.arquivo_url = fname
            alterados += 1
    if alterados:
        db.session.commit()
    return jsonify({"ok": True, "alterados": alterados})

# =========================
# AVISOS â€” AÃ§Ãµes (cooperado)
# =========================

@app.post("/avisos/<int:aviso_id>/lido", endpoint="marcar_aviso_lido")
@role_required("cooperado")
def marcar_aviso_lido(aviso_id: int):
    u_id = session.get("user_id")
    coop = Cooperado.query.filter_by(usuario_id=u_id).first_or_404()

    aviso = Aviso.query.get_or_404(aviso_id)

    ja_lido = AvisoLeitura.query.filter_by(
        cooperado_id=coop.id, aviso_id=aviso.id
    ).first()

    if not ja_lido:
        db.session.add(AvisoLeitura(
            cooperado_id=coop.id,
            aviso_id=aviso.id,
            lido_em=datetime.now(timezone.utc),
        ))
        db.session.commit()

    # volta para a lista; se quiser voltar ancorado: + f"#aviso-{aviso.id}"
    return redirect(url_for("portal_cooperado_avisos"))

@app.post("/avisos/marcar-todos", endpoint="marcar_todos_avisos_lidos")
@role_required("cooperado")
def marcar_todos_avisos_lidos():
    u_id = session.get("user_id")
    coop = Cooperado.query.filter_by(usuario_id=u_id).first_or_404()

    # todos avisos visÃ­veis ao cooperado
    avisos = get_avisos_for_cooperado(coop)

    # ids jÃ¡ lidos
    lidos_ids = {
        a_id for (a_id,) in db.session.query(AvisoLeitura.aviso_id)
        .filter(AvisoLeitura.cooperado_id == coop.id).all()
    }

    # persiste sÃ³ os que faltam
    now = datetime.now(timezone.utc)
    for a in avisos:
        if a.id not in lidos_ids:
            db.session.add(AvisoLeitura(
                cooperado_id=coop.id,
                aviso_id=a.id,
                lido_em=now,
            ))

    db.session.commit()
    return redirect(url_for("portal_cooperado_avisos"))

# =========================
# PORTAL: Avisos (Restaurante/FarmÃ¡cia)
# =========================
@app.get("/portal/restaurante/avisos")
@role_required("restaurante", "farmacia")
def portal_restaurante_avisos():
    u_id = session.get("user_id")
    rest = Restaurante.query.filter_by(usuario_id=u_id).first_or_404()

    # Carrega avisos aplicÃ¡veis (helper acima)
    avisos_db = get_avisos_for_restaurante(rest)

    # Quais jÃ¡ foram lidos por este rest?
    lidos_ids = {
        a_id for (a_id,) in db.session.query(AvisoLeitura.aviso_id)
        .filter(AvisoLeitura.restaurante_id == rest.id).all()
    }

    # Monta payload para o template
    avisos = [{
        "id": a.id,
        "titulo": a.titulo or "Aviso",
        "criado_em": a.criado_em,
        "lido": (a.id in lidos_ids),
        "prioridade_alta": (str(a.prioridade or "").lower() == "alta"),
        # seu template espera 'corpo_html'; aqui passamos Aviso.corpo
        "corpo_html": a.corpo or "",
    } for a in avisos_db]

    avisos_nao_lidos_count = sum(1 for x in avisos if not x["lido"])

    return render_template(
        "portal_restaurante_avisos.html",
        avisos=avisos,
        avisos_nao_lidos_count=avisos_nao_lidos_count,
        current_year=datetime.now(timezone.utc).year,
    )


# Marcar todos como lidos (Restaurante/FarmÃ¡cia)
@app.post("/avisos-restaurante/marcar-todos", endpoint="marcar_todos_avisos_lidos_restaurante")
@role_required("restaurante", "farmacia")
def marcar_todos_avisos_lidos_restaurante():
    u_id = session.get("user_id")
    rest = Restaurante.query.filter_by(usuario_id=u_id).first_or_404()

    avisos = get_avisos_for_restaurante(rest)

    # ids jÃ¡ lidos
    lidos_ids = {
        a_id for (a_id,) in db.session.query(AvisoLeitura.aviso_id)
        .filter(AvisoLeitura.restaurante_id == rest.id).all()
    }

    now = datetime.now(timezone.utc)
    for a in avisos:
        if a.id not in lidos_ids:
            db.session.add(AvisoLeitura(
                restaurante_id=rest.id,
                aviso_id=a.id,
                lido_em=now,
            ))
    db.session.commit()
    return redirect(url_for("portal_restaurante_avisos"))

# =========================
# Main
# =========================
if __name__ == "__main__":
    with app.app_context():
        init_db()
    app.run(debug=True, host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
    


